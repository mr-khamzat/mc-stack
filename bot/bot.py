#!/usr/bin/env python3
"""
MeshCentral Monitoring & Inventory Telegram Bot — v4
All improvements: .env config, async HTTP, cached DB export, pagination,
search, device alerts, change history, uptime graphs, device compare,
WoL, software inventory, MC config backup, graceful shutdown,
remote commands via MeshCentral API, PDF inventory export.
"""

import asyncio
import json
import os
import subprocess
import shutil
import logging
import time
import io
import csv
import re
import ssl
import signal
import struct
import socket
import ipaddress
from datetime import datetime, timezone, timedelta
from pathlib import Path

import aiohttp
import psutil
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from dotenv import load_dotenv
from fpdf import FPDF
import pyzipper
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
logging.getLogger("fontTools.subset").setLevel(logging.WARNING)
from aiogram import Bot, Dispatcher, F, Router
from aiogram.types import (
    Message, CallbackQuery, BufferedInputFile,
    InlineKeyboardMarkup, InlineKeyboardButton,
    ReplyKeyboardMarkup, KeyboardButton,
)
from aiogram.filters import Command
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext

# ─── Config ───────────────────────────────────────────────────────────

load_dotenv("/opt/meshcentral-bot/.env")

BOT_TOKEN = os.getenv("BOT_TOKEN", "")
MC_URL = os.getenv("MC_URL", "https://hub.office.mooo.com")
MC_DATA = os.getenv("MC_DATA", "/opt/meshcentral/meshcentral-data")
MC_DIR = os.getenv("MC_DIR", "/opt/meshcentral")
MC_WSS = os.getenv("MC_WSS", "wss://hub.office.mooo.com:443")
MESHCTRL = f"{MC_DIR}/node_modules/meshcentral/meshctrl.js"
ADMIN_FILE = "/opt/meshcentral-bot/admin.json"
DATA_DIR = Path("/opt/meshcentral-bot")
HISTORY_FILE = DATA_DIR / "history.json"
UPTIME_FILE = DATA_DIR / "uptime.json"
ALERTS_FILE = DATA_DIR / "alerts_cfg.json"
SNAPSHOTS_FILE = DATA_DIR / "snapshots.json"
SCRIPTS_FILE = DATA_DIR / "scripts.json"
MUTE_FILE              = DATA_DIR / "mute.json"
WIFI_FILE              = DATA_DIR / "wifi_clients.json"
KEENETIC_PROBES_FILE   = DATA_DIR / "keenetic_probes.json"
KEENETIC_PROBE_SCRIPT  = DATA_DIR / "keenetic_probe.ps1"
NETMAP_FILE            = DATA_DIR / "public" / "netmap.html"
NOTES_FILE             = DATA_DIR / "notes.json"
DISK_HISTORY_FILE      = DATA_DIR / "disk_history.json"
SNAP_HISTORY_FILE      = DATA_DIR / "snap_history.json"
SCHEDULER_FILE         = DATA_DIR / "cmd_scheduler.json"
SNMP_DATA_FILE         = DATA_DIR / "snmp_data.json"
SNMP_PROBE_SCRIPT      = DATA_DIR / "snmp_probe.ps1"
SNMP_POLL_INTERVAL     = 300  # seconds (5 min)
PRINTERS_FILE          = DATA_DIR / "printers.json"
PRINTER_INK_PS1        = DATA_DIR / "printer_ink.ps1"
INK_ALERTS_FILE        = DATA_DIR / "ink_alerts.json"
INK_WARN_PCT           = 20   # % threshold for low ink alert
NETMAP_INTERVAL        = 60   # seconds
WIFI_POLL_INTERVAL     = 8 * 3600  # seconds (8 hours)
# ── New features ──
HW_INVENTORY_FILE  = DATA_DIR / "hw_inventory.json"
HW_INVENTORY_PS1   = DATA_DIR / "hw_inventory.ps1"
TEMP_DATA_FILE     = DATA_DIR / "temp_data.json"
TEMP_PROBE_PS1     = DATA_DIR / "temp_probe.ps1"
STATUS_HTML_FILE   = DATA_DIR / "public" / "status.html"
HW_POLL_INTERVAL   = 4 * 3600   # 4 hours
TEMP_POLL_INTERVAL = 900         # 15 minutes
TEMP_WARN_C        = 75          # °C alert threshold

HEALTH_CHECK_INTERVAL = 60
DEVICE_CHECK_INTERVAL = 45
INVENTORY_HOUR = 8
DAILY_REPORT_HOUR = 9
WEEKLY_DIGEST_HOUR = 10   # Sunday 10:00 UTC
UPDATE_CHECK_HOUR = 11
DB_CACHE_TTL = 60  # seconds — снижено для более актуального статуса
SSL_DOMAINS = [d.strip() for d in os.getenv("SSL_DOMAINS", "hub.office.mooo.com,panelwin.mooo.com,subwin.mooo.com").split(",") if d.strip()]
# HTTP services to monitor: "Name|url" pairs comma-separated in HTTP_SERVICES env var
_HTTP_SERVICES_RAW = os.getenv(
    "HTTP_SERVICES",
    "MeshCentral|https://hub.office.mooo.com,"
    "Remnawave Panel|https://panelwin.mooo.com/api/"
)
HTTP_SERVICES: list[tuple[str, str]] = []
for _item in _HTTP_SERVICES_RAW.split(","):
    _item = _item.strip()
    if "|" in _item:
        _sn, _su = _item.split("|", 1)
        HTTP_SERVICES.append((_sn.strip(), _su.strip()))
SSL_WARN_DAYS = int(os.getenv("SSL_WARN_DAYS", "30"))
SSL_CRIT_DAYS = int(os.getenv("SSL_CRIT_DAYS", "7"))
SSL_CHECK_HOUR = 11  # час UTC для ежедневной проверки

PAGE_SIZE = 5

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("mc-bot")

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(storage=MemoryStorage())
router = Router()
dp.include_router(router)

# ─── State ────────────────────────────────────────────────────────────

_known_devices: dict[str, dict] = {}
_last_inventory_date = ""
_last_daily_report = ""
_last_weekly_digest = ""
_last_update_check = ""
_last_ssl_check = ""
_ssl_cache: list = []  # [{domain, days_left, expires, ok, error}]
_mc_was_down = False
_http_down: dict[str, bool] = {}   # service name → was_down flag
_db_cache: list = []
_db_cache_time: float = 0
_online_cache: set = set()       # node IDs currently online (from meshctrl)
_online_cache_time: float = 0
_shutdown_event = asyncio.Event()
_background_tasks: list[asyncio.Task] = []
_wifi_clients: dict = {}  # {agent_name: {ok, router, updated, count, clients: [...]}}
_snmp_data:    dict = {}  # {agent_name: {ok, router, updated, data: {...}, prev: {...}}}
_hw_inventory: dict = {}  # {device_name: {hostname, cpu_name, ram_total_gb, disks, ...}}
_temp_data:    dict = {}  # {device_name: {temps, cpu_load_pct, updated}}

# ─── Keyboard ─────────────────────────────────────────────────────────

BTN_STATUS    = "🖥 Статус"
BTN_DEVICES   = "📋 Устройства"
BTN_INVENTORY = "📦 Инвентарь"
BTN_HEALTH    = "❤️ Здоровье"
BTN_TOOLS     = "🔧 Инструменты"
BTN_WIFI      = "📡 WiFi сети"

MAIN_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=BTN_STATUS), KeyboardButton(text=BTN_DEVICES)],
        [KeyboardButton(text=BTN_INVENTORY), KeyboardButton(text=BTN_HEALTH)],
        [KeyboardButton(text=BTN_TOOLS), KeyboardButton(text=BTN_WIFI)],
    ],
    resize_keyboard=True, is_persistent=True,
)

# ─── Persistence helpers ─────────────────────────────────────────────

def _load_json(path: Path, default=None):
    if default is None:
        default = {}
    if path.exists():
        try:
            with open(path) as f:
                return json.load(f)
        except Exception:
            pass
    return default

def _save_json(path: Path, data):
    tmp = path.with_suffix(".tmp")
    with open(tmp, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)
    tmp.replace(path)


# ─── Admin ────────────────────────────────────────────────────────────

def load_admin() -> dict:
    return _load_json(Path(ADMIN_FILE))

def save_admin(d: dict):
    _save_json(Path(ADMIN_FILE), d)

def get_admin_id() -> int | None:
    return load_admin().get("admin_id")

def is_admin(uid: int) -> bool:
    d = load_admin()
    return True if not d.get("admin_id") else d["admin_id"] == uid

def lock_admin(uid: int, uname: str) -> bool:
    d = load_admin()
    if not d.get("admin_id"):
        d.update(admin_id=uid, username=uname)
        save_admin(d)
        return True
    return False


# ─── Alerts config ───────────────────────────────────────────────────

DEFAULT_ALERTS = {
    "disk_pct": 90,
    "av_off": True,
    "offline_hours": 24,
    "new_device": True,
}

def load_alerts_cfg() -> dict:
    cfg = _load_json(ALERTS_FILE, DEFAULT_ALERTS.copy())
    for k, v in DEFAULT_ALERTS.items():
        cfg.setdefault(k, v)
    return cfg

def save_alerts_cfg(cfg: dict):
    _save_json(ALERTS_FILE, cfg)


# ─── Quick Scripts ───────────────────────────────────────────────────

DEFAULT_SCRIPTS = {
    # ─── 🖥 СИСТЕМА ────────────────────────────────────────────────
    "sys_info": {
        "cmd": "systeminfo | findstr /C:\"OS Name\" /C:\"OS Version\" /C:\"System Boot Time\" /C:\"Total Physical Memory\" /C:\"Available Physical Memory\" /C:\"Domain\"",
        "ps": False, "cat": "system",
        "desc": "🖥 Версия ОС, домен, время загрузки, память",
    },
    "uptime": {
        "cmd": "$b=(gcim Win32_OperatingSystem).LastBootUpTime; $u=(Get-Date)-$b; Write-Host \"Boot: $($b.ToString('dd.MM.yyyy HH:mm'))`nUptime: $($u.Days)д $($u.Hours)ч $($u.Minutes)м\"",
        "ps": True, "cat": "system",
        "desc": "⏱ Время работы без перезагрузки",
    },
    "top_cpu": {
        "cmd": "Get-Process | Sort-Object CPU -Desc | Select-Object -First 10 Name,@{N='CPU';E={[Math]::Round($_.CPU,1)}},@{N='RAM_MB';E={[Math]::Round($_.WS/1MB,0)}},Id | Format-Table -Auto",
        "ps": True, "cat": "system",
        "desc": "📊 Топ-10 процессов по CPU",
    },
    "top_ram": {
        "cmd": "Get-Process | Sort-Object WS -Desc | Select-Object -First 10 Name,@{N='RAM_MB';E={[Math]::Round($_.WS/1MB,0)}},Id | Format-Table -Auto",
        "ps": True, "cat": "system",
        "desc": "💾 Топ-10 процессов по памяти",
    },
    "disk_space": {
        "cmd": "Get-PSDrive -PSProvider FileSystem | Select-Object Name,@{N='Used_GB';E={[Math]::Round($_.Used/1GB,1)}},@{N='Free_GB';E={[Math]::Round($_.Free/1GB,1)}},@{N='Total_GB';E={[Math]::Round(($_.Used+$_.Free)/1GB,1)}} | Format-Table -Auto",
        "ps": True, "cat": "system",
        "desc": "💿 Свободное место на всех дисках",
    },
    "logged_users": {
        "cmd": "query user 2>&1",
        "ps": False, "cat": "system",
        "desc": "👤 Залогиненные пользователи на ПК",
    },
    "last_errors": {
        "cmd": "Get-EventLog -LogName System -EntryType Error -Newest 10 | Select-Object TimeWritten,Source,@{N='Msg';E={$_.Message.Substring(0,[Math]::Min(100,$_.Message.Length))}} | Format-List",
        "ps": True, "cat": "system",
        "desc": "⚠️ Последние 10 ошибок системного журнала",
    },
    "autorun": {
        "cmd": "Get-CimInstance Win32_StartupCommand | Select-Object Name,Command,User | Format-Table -Auto",
        "ps": True, "cat": "system",
        "desc": "🚀 Программы в автозапуске",
    },
    # ─── 🌐 СЕТЬ ───────────────────────────────────────────────────
    "net_config": {
        "cmd": "ipconfig /all",
        "ps": False, "cat": "network",
        "desc": "🌐 Полная сетевая конфигурация (ipconfig /all)",
    },
    "net_adapters": {
        "cmd": "Get-NetAdapter | Where-Object Status -eq 'Up' | Select-Object Name,InterfaceDescription,LinkSpeed,MacAddress | Format-Table -Auto",
        "ps": True, "cat": "network",
        "desc": "🔌 Активные сетевые адаптеры",
    },
    "connections": {
        "cmd": "Get-NetTCPConnection -State Established | Select-Object LocalAddress,LocalPort,RemoteAddress,RemotePort | Sort-Object RemoteAddress | Format-Table -Auto",
        "ps": True, "cat": "network",
        "desc": "🔗 Активные TCP-соединения",
    },
    "ping_inet": {
        "cmd": "Test-Connection 8.8.8.8 -Count 4 | Select-Object Address,ResponseTime | Format-Table",
        "ps": True, "cat": "network",
        "desc": "🌍 Пинг до 8.8.8.8 (проверка интернета)",
    },
    "flush_dns": {
        "cmd": "ipconfig /flushdns",
        "ps": False, "cat": "network",
        "desc": "🧹 Сброс DNS-кэша",
    },
    # ─── 🔧 ОБСЛУЖИВАНИЕ ───────────────────────────────────────────
    "clear_temp": {
        "cmd": "del /q/f/s %TEMP%\\* 2>nul & del /q/f/s C:\\Windows\\Temp\\* 2>nul & echo Done",
        "ps": False, "cat": "maintenance",
        "desc": "🗑 Очистка TEMP-папок (пользователь + система)",
    },
    "gpupdate": {
        "cmd": "gpupdate /force",
        "ps": False, "cat": "maintenance",
        "desc": "🔄 Принудительное обновление групповых политик",
    },
    "sfc_scan": {
        "cmd": "sfc /scannow",
        "ps": False, "cat": "maintenance",
        "desc": "🛡 Проверка целостности системных файлов Windows",
    },
    "check_updates": {
        "cmd": "(New-Object -ComObject Microsoft.Update.Session).CreateUpdateSearcher().Search('IsInstalled=0').Updates | Select-Object Title | Format-List",
        "ps": True, "cat": "maintenance",
        "desc": "🆙 Доступные обновления Windows",
    },
    "installed_soft": {
        "cmd": "Get-ItemProperty 'HKLM:\\Software\\Microsoft\\Windows\\CurrentVersion\\Uninstall\\*','HKLM:\\Software\\Wow6432Node\\Microsoft\\Windows\\CurrentVersion\\Uninstall\\*' -EA 0 | Where-Object DisplayName | Select-Object DisplayName,DisplayVersion | Sort-Object DisplayName | Format-Table -Auto",
        "ps": True, "cat": "maintenance",
        "desc": "📦 Список установленного ПО",
    },
    # ─── 🖨 ПРИНТЕРЫ ───────────────────────────────────────────────
    "printers": {
        "cmd": "Get-Printer | Select-Object Name,PrinterStatus,PortName,Shared,Default | Format-Table -Auto",
        "ps": True, "cat": "printers",
        "desc": "🖨 Список принтеров и их статус",
    },
    "restart_spooler": {
        "cmd": "net stop spooler && net start spooler",
        "ps": False, "cat": "printers",
        "desc": "🔄 Перезапуск службы печати (спулер)",
    },
    "clear_print_queue": {
        "cmd": "net stop spooler & del /q/f/s %systemroot%\\System32\\spool\\PRINTERS\\* 2>nul & net start spooler & echo Done",
        "ps": False, "cat": "printers",
        "desc": "🗑 Очистка очереди печати + перезапуск спулера",
    },
    # ─── 🛡 БЕЗОПАСНОСТЬ ───────────────────────────────────────────
    "local_admins": {
        "cmd": "net localgroup administrators",
        "ps": False, "cat": "security",
        "desc": "👥 Локальные администраторы",
    },
    "local_users": {
        "cmd": "Get-LocalUser | Select-Object Name,Enabled,LastLogon | Format-Table -Auto",
        "ps": True, "cat": "security",
        "desc": "👤 Локальные пользователи и статус учётных записей",
    },
    "firewall": {
        "cmd": "netsh advfirewall show allprofiles state",
        "ps": False, "cat": "security",
        "desc": "🔥 Статус брандмауэра Windows (все профили)",
    },
    "shares": {
        "cmd": "net share",
        "ps": False, "cat": "security",
        "desc": "📁 Общие папки (сетевые шары)",
    },
}

def load_scripts() -> dict:
    scripts = _load_json(SCRIPTS_FILE, {})
    for k, v in DEFAULT_SCRIPTS.items():
        scripts.setdefault(k, v)
    return scripts

def save_scripts(scripts: dict):
    _save_json(SCRIPTS_FILE, scripts)


# ─── Maintenance Mode (Mutes) ───────────────────────────────────────

def load_mutes() -> dict:
    return _load_json(MUTE_FILE, {})

def save_mutes(mutes: dict):
    _save_json(MUTE_FILE, mutes)

def cleanup_expired_mutes():
    mutes = load_mutes()
    now = time.time()
    changed = False
    for key in list(mutes.keys()):
        if mutes[key].get("until", 0) and mutes[key]["until"] < now:
            del mutes[key]
            changed = True
    if changed:
        save_mutes(mutes)

def is_muted(device_name: str, group: str) -> bool:
    cleanup_expired_mutes()
    mutes = load_mutes()
    now = time.time()
    # Check exact device mute
    if device_name in mutes:
        m = mutes[device_name]
        if m.get("until", 0) == 0 or m["until"] > now:
            return True
    # Check group mute
    if group in mutes:
        m = mutes[group]
        if m.get("until", 0) == 0 or m["until"] > now:
            return True
    # Check __all__ mute
    if "__all__" in mutes:
        m = mutes["__all__"]
        if m.get("until", 0) == 0 or m["until"] > now:
            return True
    return False

def parse_duration(s: str) -> int | None:
    """Parse duration string like 30m, 2h, 1d into seconds. Returns None on error."""
    s = s.strip().lower()
    m = re.match(r'^(\d+)\s*([mhd])$', s)
    if not m:
        return None
    val = int(m.group(1))
    unit = m.group(2)
    if unit == 'm':
        return val * 60
    elif unit == 'h':
        return val * 3600
    elif unit == 'd':
        return val * 86400
    return None


# ─── Alert Help System ───────────────────────────────────────────────

ALERT_HELP = {
    "disk": {
        "title": "💿 Алерт: Диск заполнен",
        "what": "Срабатывает когда использование диска превышает заданный порог.",
        "when": "Проверка каждые {interval}с. Алерт отправляется 1 раз в день на устройство.",
        "config": "🔧 Инструменты → 🔔 Алерты → 💿 Порог\nТекущий порог: {threshold}%\nВарианты: 80%, 85%, 90%, 95%",
        "action": (
            "1. Проверьте что занимает место: /run ИмяПК -ps Get-ChildItem C:\\ -Recurse | Sort Length -Desc | Select -First 20 FullName,Length\n"
            "2. Очистите временные файлы: /run ИмяПК cleanmgr /d C\n"
            "3. Проверьте корзину, загрузки, логи\n"
            "4. Используйте /mute ИмяПК 2h для временного отключения"
        ),
        "example": "💿 Диск заполнен: PC-OFFICE\nC: 95%, D: 91%",
    },
    "av": {
        "title": "🛡 Алерт: Антивирус выключен",
        "what": "Срабатывает когда антивирус на устройстве отключён.",
        "when": "Проверка каждые {interval}с. Алерт 1 раз в день.",
        "config": "🔧 Инструменты → 🔔 Алерты → 🛡 AV\nВкл/Выкл переключатель",
        "action": (
            "1. Проверьте статус: /run ИмяПК -ps Get-MpComputerStatus | Select AntivirusEnabled,RealTimeProtectionEnabled\n"
            "2. Включите защиту: /run ИмяПК -ps Set-MpPreference -DisableRealtimeMonitoring $false\n"
            "3. Убедитесь что AV обновлён: /run ИмяПК -ps Update-MpSignature"
        ),
        "example": "🛡 Антивирус выключен: LAPTOP-USER\nWindows Defender (off)",
    },
    "offline": {
        "title": "⏰ Алерт: Устройство долго офлайн",
        "what": "Срабатывает когда устройство не подключалось дольше заданного времени.",
        "when": "Проверка каждые {interval}с. Алерт 1 раз в день.",
        "config": "🔧 Инструменты → 🔔 Алерты → ⏰ Офлайн\nТекущий порог: {offline_hours}ч\nВарианты: 6ч, 12ч, 24ч, 48ч, 72ч",
        "action": (
            "1. Проверьте физическое подключение устройства\n"
            "2. Попробуйте Wake-on-LAN из карточки устройства\n"
            "3. Проверьте сеть на стороне устройства\n"
            "4. Убедитесь что служба Mesh Agent запущена"
        ),
        "example": "⏰ Долго офлайн: SERVER-BACKUP (72ч)",
    },
    "new_device": {
        "title": "🆕 Алерт: Новое устройство",
        "what": "Срабатывает при обнаружении нового устройства в MeshCentral.",
        "when": "Проверка каждые {interval}с. Алерт при первом появлении.",
        "config": "🔧 Инструменты → 🔔 Алерты → 🆕 Новое\nВкл/Выкл переключатель",
        "action": (
            "1. Проверьте карточку нового устройства\n"
            "2. Убедитесь что оно в нужной группе\n"
            "3. Проверьте установленное ПО и безопасность\n"
            "4. При необходимости переместите в нужную группу в MC"
        ),
        "example": "🆕 Новое устройство: NEW-PC\n💻 Windows 11 Pro\n🌐 192.168.1.50",
    },
}


# ─── DB Export & Parse (cached, async subprocess) ────────────────────

async def _export_db_async() -> list:
    global _db_cache, _db_cache_time
    now = time.time()
    if _db_cache and (now - _db_cache_time) < DB_CACHE_TTL:
        return _db_cache

    try:
        proc = await asyncio.create_subprocess_exec(
            "node", f"{MC_DIR}/node_modules/meshcentral/meshcentral.js", "--dbexport",
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
            cwd=MC_DIR,
        )
        await asyncio.wait_for(proc.wait(), timeout=30)
    except asyncio.TimeoutError:
        log.error("DB export timed out")
        return _db_cache or []
    except Exception as e:
        log.error(f"DB export error: {e}")
        return _db_cache or []

    db_file = f"{MC_DATA}/meshcentral.db.json"
    if not os.path.exists(db_file):
        return _db_cache or []
    try:
        with open(db_file) as f:
            data = json.load(f)
        os.remove(db_file)
        _db_cache = data
        _db_cache_time = now
    except Exception as e:
        log.error(f"DB parse error: {e}")
    return _db_cache


async def _list_agents_quick() -> list[dict]:
    """Fast agent list via meshctrl ListDevices (used for WiFi office FSM picker).
    Returns [{name, group, online, id}]. Much faster than get_full_devices().
    """
    key = await _get_login_key()
    if not key:
        return []
    try:
        proc = await asyncio.create_subprocess_exec(
            "node", MESHCTRL, "ListDevices",
            "--url", MC_WSS,
            "--loginkey", key,
            "--json",
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
        )
        stdout, _ = await asyncio.wait_for(proc.communicate(), timeout=20)
        raw = stdout.decode(errors="replace")
        idx = raw.find("[")
        if idx == -1:
            return []
        data = json.loads(raw[idx:])
        return [
            {
                "id":     d.get("_id", ""),
                "name":   d.get("name", "?"),
                "group":  d.get("groupname", ""),
                "online": bool(d.get("conn", 0) & 1),
            }
            for d in data if d.get("name")
        ]
    except Exception as e:
        log.error(f"_list_agents_quick: {e}")
        return []


async def _get_realtime_online_ids() -> set:
    """Get set of node IDs currently connected via meshctrl ListDevices.
    Falls back to empty set on error (caller will use lastconnect fallback).
    Cached for 45 seconds.
    """
    global _online_cache, _online_cache_time
    now = time.time()
    if _online_cache and (now - _online_cache_time) < 45:
        return _online_cache

    try:
        login_key = await _get_login_key()
        if not login_key:
            return _online_cache
        proc = await asyncio.create_subprocess_exec(
            "node", MESHCTRL, "ListDevices",
            "--url", MC_WSS,
            "--loginkey", login_key,
            "--json",
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
        )
        stdout, _ = await asyncio.wait_for(proc.communicate(), timeout=20)
        raw = stdout.decode(errors="replace").strip()
        # meshctrl may print log lines before JSON — find the JSON array
        json_start = raw.find("[")
        if json_start == -1:
            return _online_cache
        data = json.loads(raw[json_start:])
        online_ids: set = set()
        for dev in data:
            # conn flag 1 = agent connected
            if dev.get("conn", 0) & 1:
                nid = dev.get("_id", "")
                if nid:
                    online_ids.add(nid)
        _online_cache = online_ids
        _online_cache_time = now
        return _online_cache
    except Exception as e:
        log.warning(f"realtime online ids error: {e}")
        return _online_cache


def _extract_node_id(full_id: str) -> str:
    for prefix in ("sinode//", "ifnode//", "lcnode//"):
        if full_id.startswith(prefix):
            return "node//" + full_id[len(prefix):]
    return full_id


def _fmt_size(b) -> str:
    b = int(b) if b else 0
    if b >= 1024**4:
        return f"{b / 1024**4:.1f} TB"
    if b >= 1024**3:
        return f"{b / 1024**3:.0f} GB"
    if b >= 1024**2:
        return f"{b / 1024**2:.0f} MB"
    return f"{b} B"


async def get_full_devices() -> list[dict]:
    """Parse DB into rich device objects with full hardware info."""
    raw, realtime_online = await asyncio.gather(
        _export_db_async(),
        _get_realtime_online_ids(),
    )
    if not raw:
        return []

    meshes = {r["_id"]: r.get("name", "?") for r in raw if r.get("type") == "mesh"}
    nodes = {r["_id"]: r for r in raw if r.get("type") == "node"}

    sysinfos = {}
    ifinfos = {}
    lastconns = {}
    for r in raw:
        rid = r.get("_id", "")
        rtype = r.get("type", "")
        if rtype == "sysinfo":
            sysinfos[_extract_node_id(rid)] = r
        elif rtype == "ifinfo":
            ifinfos[_extract_node_id(rid)] = r
        elif rtype == "lastconnect":
            lastconns[_extract_node_id(rid)] = r

    devices = []
    for nid, n in nodes.items():
        si = sysinfos.get(nid, {})
        ii = ifinfos.get(nid, {})
        lc = lastconns.get(nid, {})

        hw = si.get("hardware", {})
        win = hw.get("windows", {})
        ident = hw.get("identifiers", {})
        tpm = hw.get("tpm", {})
        net_info = hw.get("network", {})

        # CPU
        cpus = win.get("cpu", [])
        cpu_str = ", ".join(c.get("Name", "?").strip() for c in cpus) if cpus else ident.get("cpu_name", "-")

        # RAM
        ram_modules = win.get("memory", [])
        ram_total = sum(int(m.get("Capacity", 0)) for m in ram_modules)
        ram_details = []
        for m in ram_modules:
            cap = _fmt_size(m.get("Capacity", 0))
            pn = m.get("PartNumber", "").strip()
            spd = m.get("Speed", "")
            slot = m.get("DeviceLocator", "")
            ram_details.append(f"{slot}: {cap} {pn} {spd}MHz")

        # GPU
        gpus = win.get("gpu", [])
        gpu_str = ", ".join(g.get("Name", "?") for g in gpus) if gpus else ", ".join(ident.get("gpu_name", []))

        # Drives
        drives = win.get("drives", []) or ident.get("storage_devices", [])
        drive_details = []
        for d in drives:
            model = d.get("Model", d.get("Caption", "?"))
            size = _fmt_size(d.get("Size", 0))
            drive_details.append(f"{model} ({size})")

        # Volumes
        volumes = win.get("volumes", {})
        vol_details = []
        vol_alerts = []
        for letter, v in volumes.items():
            vname = v.get("name", "")
            vtype = v.get("type", "")
            vsize = v.get("size", 0)
            vfree = v.get("sizeremaining", 0)
            vsize_s = _fmt_size(vsize)
            vfree_s = _fmt_size(vfree)
            label = f" [{vname}]" if vname else ""
            vol_details.append(f"{letter}:{label} {vtype} {vfree_s}/{vsize_s} free")
            if vsize and vfree:
                used_pct = (1 - int(vfree) / int(vsize)) * 100
                if used_pct >= 90:
                    vol_alerts.append(f"{letter}: {used_pct:.0f}%")

        # Motherboard & BIOS
        board = f"{ident.get('board_vendor', '')} {ident.get('board_name', '')}".strip() or "-"
        board_sn = ident.get("board_serial", "-")
        bios_date = ident.get("bios_date", "")
        bios = f"{ident.get('bios_vendor', '')} v{ident.get('bios_version', '')} ({bios_date[:8]})".strip()
        bios_mode = ident.get("bios_mode", "-")

        # OS details
        osinfo = win.get("osinfo", {})
        os_full = osinfo.get("Caption", n.get("osdesc", ""))
        os_arch = osinfo.get("OSArchitecture", "")
        os_build = osinfo.get("BuildNumber", "")
        os_sn = osinfo.get("SerialNumber", "-")
        os_install = osinfo.get("InstallDate", "")[:8] if osinfo.get("InstallDate") else "-"
        os_domain = osinfo.get("Domain", "WORKGROUP")

        # Antivirus
        av_list = n.get("av", [])
        av_str = ", ".join(
            f"{a.get('product', '?')} ({'on' if a.get('enabled') else 'off'})"
            for a in av_list
        ) if av_list else "-"
        av_disabled = any(not a.get("enabled", True) for a in av_list) if av_list else False
        wsc = n.get("wsc", {})

        # TPM
        tpm_str = f"v{tpm.get('SpecVersion', '?')} {tpm.get('ManufacturerId', '')}" if tpm else "-"

        # Network interfaces
        netifs = ii.get("netif2", {})
        nic_details = []
        for iname, addrs in netifs.items():
            if "Loopback" in iname:
                continue
            ipv4s = [a["address"] for a in addrs if a.get("family") == "IPv4" and not a["address"].startswith("169.254")]
            mac = addrs[0].get("mac", "") if addrs else ""
            status = addrs[0].get("status", "") if addrs else ""
            speed = addrs[0].get("speed", 0) if addrs else 0
            speed_str = f"{speed // 1_000_000}Mbps" if speed and speed < 9e18 else ""
            if ipv4s:
                nic_details.append({"name": iname, "ips": ipv4s, "mac": mac, "status": status, "speed": speed_str})

        # Online status: prefer real-time from meshctrl (conn flag), fallback to lastconnect
        lc_time = lc.get("time")
        lc_addr = lc.get("addr", "-")
        if realtime_online:
            # meshctrl gave us live data — use it as source of truth
            online = nid in realtime_online
            offline_hours = 0 if online else (
                (time.time() * 1000 - lc_time) / 3_600_000 if lc_time else 0
            )
        else:
            # fallback: use lastconnect timestamp (less accurate for stable connections)
            online = False
            offline_hours = 0
            if lc_time:
                diff_ms = time.time() * 1000 - lc_time
                online = diff_ms < 300_000
                offline_hours = diff_ms / 3_600_000 if not online else 0

        # Users
        users = n.get("users", [])
        last_boot = n.get("lastbootuptime")
        boot_str = datetime.fromtimestamp(last_boot / 1000, tz=timezone.utc).strftime("%d.%m.%Y %H:%M") if last_boot else "-"

        # Resolution
        res_str = "-"
        if gpus:
            g = gpus[0]
            h = g.get("CurrentHorizontalResolution")
            v = g.get("CurrentVerticalResolution")
            if h and v:
                res_str = f"{h}x{v}"

        # Software
        software = win.get("software", {})
        sw_list = []
        if isinstance(software, dict):
            for sw_name, sw_info in software.items():
                ver = sw_info.get("version", "") if isinstance(sw_info, dict) else ""
                sw_list.append({"name": sw_name, "version": ver})
        elif isinstance(software, list):
            for sw in software:
                sw_list.append({"name": sw.get("name", "?"), "version": sw.get("version", "")})

        nic_str_list = []
        for nic in nic_details[:5]:
            nic_str_list.append(f"{nic['name']}: {', '.join(nic['ips'])} ({nic['mac']}) {nic['speed']} [{nic['status']}]")

        devices.append({
            "id": nid,
            "name": n.get("name", "?"),
            "group": meshes.get(n.get("meshid", ""), "?"),
            "online": online,
            "ip": n.get("ip", ""),
            "lc_addr": lc_addr,
            "offline_hours": offline_hours,
            # OS
            "os": os_full,
            "os_arch": os_arch,
            "os_build": os_build,
            "os_sn": os_sn,
            "os_install": os_install,
            "os_domain": os_domain,
            # Hardware
            "cpu": cpu_str,
            "ram_total": _fmt_size(ram_total),
            "ram_details": ram_details,
            "gpu": gpu_str or "-",
            "resolution": res_str,
            "drives": drive_details,
            "volumes": vol_details,
            "vol_alerts": vol_alerts,
            "volumes_raw": {lt: {"total": int(v.get("size", 0)), "free": int(v.get("sizeremaining", 0))} for lt, v in volumes.items() if v.get("size", 0) > 0},
            "board": board,
            "board_sn": board_sn,
            "bios": bios,
            "bios_mode": bios_mode,
            "tpm": tpm_str,
            # Security
            "antivirus": av_str,
            "av_disabled": av_disabled,
            "firewall": wsc.get("firewall", "-"),
            "auto_update": wsc.get("autoUpdate", "-"),
            # Network
            "nics": nic_str_list,
            "nic_details": nic_details,
            "dns": net_info.get("dns", []),
            # Users
            "users": users,
            "last_boot": boot_str,
            # Agent
            "agent_ver": str(n.get("agent", {}).get("ver", "")),
            "agent_core": n.get("agent", {}).get("core", ""),
            # Software
            "software": sw_list,
        })

    return devices


# ─── Device card ─────────────────────────────────────────────────────

def build_device_card(d: dict) -> str:
    icon = "🟢" if d["online"] else "⚪"
    lines = [
        f"━━━━━━━━━━━━━━━━━━━━━━",
        f"📱 <b>{d['name']}</b>  {icon}",
        f"━━━━━━━━━━━━━━━━━━━━━━",
        f"",
        f"📁 Группа: <b>{d['group']}</b>",
        f"🌐 IP: <code>{d['ip']}</code>",
        f"🔗 Endpoint: <code>{d['lc_addr']}</code>",
        f"",
        f"<b>── ОС ──</b>",
        f"💻 {d['os']} {d['os_arch']}",
        f"🔢 Build: {d['os_build']}  SN: <code>{d['os_sn']}</code>",
        f"📅 Установлена: {d['os_install']}  Domain: {d['os_domain']}",
        f"",
        f"<b>── Железо ──</b>",
        f"🧠 CPU: {d['cpu']}",
        f"💾 RAM: <b>{d['ram_total']}</b>",
    ]
    for rm in d["ram_details"]:
        lines.append(f"   • {rm}")

    lines += [
        f"🎮 GPU: {d['gpu']}",
        f"🖥 Разрешение: {d['resolution']}",
        f"🔧 Плата: {d['board']} (SN: {d['board_sn']})",
        f"⚙️ BIOS: {d['bios']} [{d['bios_mode']}]",
        f"🔐 TPM: {d['tpm']}",
        f"",
        f"<b>── Диски ──</b>",
    ]
    for dr in d["drives"]:
        lines.append(f"   💿 {dr}")
    if d["volumes"]:
        lines.append(f"<b>── Тома ──</b>")
        for v in d["volumes"]:
            lines.append(f"   📂 {v}")

    lines += [
        f"",
        f"<b>── Безопасность ──</b>",
        f"🛡 AV: {d['antivirus']}",
        f"🧱 Firewall: {d['firewall']}  Updates: {d['auto_update']}",
        f"",
        f"<b>── Сеть ──</b>",
    ]
    for nic in d["nics"][:5]:
        lines.append(f"   🔌 {nic}")
    if d["dns"]:
        lines.append(f"   DNS: {', '.join(d['dns'])}")

    lines += [
        f"",
        f"👤 Пользователи: {', '.join(d['users']) or '-'}",
        f"🔄 Последняя загрузка: {d['last_boot']}",
        f"🤖 Агент: v{d['agent_ver']} ({d['agent_core'][:20]})",
    ]

    if d.get("vol_alerts"):
        lines.append(f"")
        lines.append(f"⚠️ <b>Диск заполнен:</b> {', '.join(d['vol_alerts'])}")
    if d.get("av_disabled"):
        lines.append(f"⚠️ <b>Антивирус выключен!</b>")

    return "\n".join(lines)


def build_inventory_csv(devices: list[dict]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";")
    w.writerow([
        "Имя", "Группа", "Online", "IP", "ОС", "Архитектура", "Build", "OS SN", "Domain",
        "CPU", "RAM", "RAM модули", "GPU", "Разрешение",
        "Материнская плата", "Board SN", "BIOS", "TPM",
        "Диски", "Тома",
        "Антивирус", "Firewall", "Auto Update",
        "Сетевые адаптеры", "DNS", "Пользователи", "Последняя загрузка",
    ])
    for d in devices:
        w.writerow([
            d["name"], d["group"], "Yes" if d["online"] else "No", d["ip"],
            d["os"], d["os_arch"], d["os_build"], d["os_sn"], d["os_domain"],
            d["cpu"], d["ram_total"], " | ".join(d["ram_details"]), d["gpu"], d["resolution"],
            d["board"], d["board_sn"], d["bios"], d["tpm"],
            " | ".join(d["drives"]), " | ".join(d["volumes"]),
            d["antivirus"], d["firewall"], d["auto_update"],
            " | ".join(d["nics"]), ", ".join(d["dns"]),
            ", ".join(d["users"]), d["last_boot"],
        ])
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def build_inventory_pdf(devices: list[dict], title: str = "MeshCentral Inventory") -> bytes:
    """Generate a PDF report with device inventory."""
    FONT = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
    FONT_B = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
    FONT_M = "/usr/share/fonts/truetype/dejavu/DejaVuSansMono.ttf"

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_font("ds", "", FONT, uni=True)
    pdf.add_font("ds", "B", FONT_B, uni=True)
    pdf.add_font("dm", "", FONT_M, uni=True)

    # ── Title page ──
    pdf.add_page()
    pdf.set_font("ds", "B", 22)
    pdf.cell(0, 40, "", ln=True)
    pdf.cell(0, 15, title, ln=True, align="C")
    pdf.set_font("ds", "", 12)
    ts = datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M UTC")
    pdf.cell(0, 10, ts, ln=True, align="C")
    pdf.cell(0, 8, f"{len(devices)} devices", ln=True, align="C")

    online = sum(1 for d in devices if d["online"])
    offline = len(devices) - online
    pdf.cell(0, 8, f"Online: {online}  |  Offline: {offline}", ln=True, align="C")
    pdf.cell(0, 20, "", ln=True)

    # ── Summary table ──
    pdf.set_font("ds", "B", 14)
    pdf.cell(0, 10, "Summary", ln=True)
    pdf.set_font("ds", "", 9)

    col_w = [8, 45, 30, 12, 40, 25, 22, 12]
    headers = ["#", "Name", "Group", "Status", "OS", "CPU", "RAM", "Agent"]
    pdf.set_fill_color(41, 128, 185)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("ds", "B", 8)
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 7, h, border=1, fill=True, align="C")
    pdf.ln()
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("ds", "", 7)

    for idx, d in enumerate(sorted(devices, key=lambda x: x["name"]), 1):
        if pdf.get_y() > 270:
            pdf.add_page()
        fill = idx % 2 == 0
        if fill:
            pdf.set_fill_color(235, 245, 251)
        status = "ON" if d["online"] else "OFF"
        os_short = d["os"][:25] if len(d["os"]) > 25 else d["os"]
        cpu_short = d["cpu"][:16] if len(d["cpu"]) > 16 else d["cpu"]
        vals = [str(idx), d["name"][:28], d["group"][:18], status, os_short, cpu_short, d["ram_total"], d["agent_ver"][:8]]
        for i, v in enumerate(vals):
            pdf.cell(col_w[i], 6, v, border=1, fill=fill, align="C" if i in (0, 3) else "L")
        pdf.ln()

    # ── Device cards ──
    for d in sorted(devices, key=lambda x: x["name"]):
        pdf.add_page()
        icon = "[ON]" if d["online"] else "[OFF]"

        # Header bar
        pdf.set_fill_color(41, 128, 185)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("ds", "B", 14)
        pdf.cell(0, 10, f"  {d['name']}  {icon}", ln=True, fill=True)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(0, 3, "", ln=True)

        def section(label):
            pdf.set_font("ds", "B", 10)
            pdf.set_fill_color(230, 230, 230)
            pdf.cell(0, 7, f"  {label}", ln=True, fill=True)

        def row(key, val):
            pdf.set_font("ds", "B", 8)
            pdf.cell(45, 5, f"  {key}:", align="L")
            pdf.set_font("ds", "", 8)
            pdf.cell(0, 5, str(val)[:90], ln=True)

        section("General")
        row("Group", d["group"])
        row("IP", d["ip"])
        row("Endpoint", d["lc_addr"])
        row("Last boot", d["last_boot"])
        row("Users", ", ".join(d["users"]) or "-")

        section("OS")
        row("OS", f"{d['os']} {d['os_arch']}")
        row("Build", d["os_build"])
        row("OS SN", d["os_sn"])
        row("Install date", d["os_install"])
        row("Domain", d["os_domain"])

        section("Hardware")
        row("CPU", d["cpu"])
        row("RAM", d["ram_total"])
        for rm in d["ram_details"][:4]:
            row("  Module", rm)
        row("GPU", d["gpu"])
        row("Resolution", d["resolution"])
        row("Motherboard", f"{d['board']} (SN: {d['board_sn']})")
        row("BIOS", f"{d['bios']} [{d['bios_mode']}]")
        row("TPM", d["tpm"])

        section("Storage")
        for dr in d["drives"][:6]:
            row("Drive", dr)
        for v in d["volumes"][:6]:
            row("Volume", v)

        section("Security")
        row("Antivirus", d["antivirus"])
        row("Firewall", d["firewall"])
        row("Auto Update", d["auto_update"])

        section("Network")
        for nic in d["nics"][:4]:
            row("NIC", nic[:85])
        if d["dns"]:
            row("DNS", ", ".join(d["dns"]))

        row("Agent", f"v{d['agent_ver']}")

        # Alerts
        if d.get("vol_alerts") or d.get("av_disabled"):
            pdf.cell(0, 3, "", ln=True)
            pdf.set_font("ds", "B", 9)
            pdf.set_text_color(200, 0, 0)
            if d.get("vol_alerts"):
                pdf.cell(0, 6, f"  WARNING: Disk full — {', '.join(d['vol_alerts'])}", ln=True)
            if d.get("av_disabled"):
                pdf.cell(0, 6, "  WARNING: Antivirus disabled!", ln=True)
            pdf.set_text_color(0, 0, 0)

    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return buf.read()


def build_single_device_pdf(d: dict) -> bytes:
    """Generate a PDF for a single device."""
    return build_inventory_pdf([d], title=f"Device Report: {d['name']}")


def _xlsx_safe(v) -> str:
    """Strip illegal control characters from a string for openpyxl cells."""
    import re as _re
    s = str(v) if v is not None else ""
    # openpyxl rejects chars outside printable XML range
    return _re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', s)


def build_inventory_xlsx(devices: list[dict]) -> bytes | None:
    """Generate an Excel report with Summary + per-group sheets."""
    if not HAS_OPENPYXL:
        return None

    wb = Workbook()
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    headers = [
        "Имя", "Группа", "Статус", "IP", "ОС", "Build",
        "CPU", "RAM", "GPU", "Диски", "Тома",
        "Антивирус", "Firewall", "TPM", "Агент",
    ]

    def write_device_sheet(ws, devs_list):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for row_idx, d in enumerate(sorted(devs_list, key=lambda x: x["name"]), 2):
            status = "Online" if d["online"] else "Offline"
            has_alerts = bool(d.get("vol_alerts") or d.get("av_disabled"))
            row_fill = green_fill if d["online"] else (yellow_fill if has_alerts else red_fill)

            vals = [
                d["name"], d["group"], status, d["ip"],
                d["os"], d["os_build"],
                d["cpu"], d["ram_total"], d["gpu"],
                " | ".join(d["drives"]), " | ".join(d["volumes"]),
                d["antivirus"], d["firewall"], d["tpm"], d["agent_ver"],
            ]
            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=_xlsx_safe(val))
                cell.border = thin_border
                if col_idx == 3:
                    cell.fill = row_fill

        # Auto column widths
        for col in range(1, len(headers) + 1):
            max_len = len(str(headers[col - 1]))
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val:
                    max_len = max(max_len, min(len(str(val)), 50))
            ws.column_dimensions[get_column_letter(col)].width = max_len + 2

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    write_device_sheet(ws, devices)

    # Per-group sheets
    by_group: dict[str, list] = {}
    for d in devices:
        by_group.setdefault(d["group"], []).append(d)
    for group_name in sorted(by_group.keys()):
        safe_name = re.sub(r'[\\/*?\[\]:]', '_', group_name)[:30]
        ws_g = wb.create_sheet(title=safe_name)
        write_device_sheet(ws_g, by_group[group_name])

    # ── Printers sheet ───────────────────────────────────────────────
    printers_db = _load_printers()
    if printers_db:
        ws_prn = wb.create_sheet(title="Принтеры")
        prn_headers = ["Устройство", "Группа", "Принтер", "Драйвер", "IP", "Статус",
                       "Умолч.", "Общий", "Чернила"]
        for col_idx, h in enumerate(prn_headers, 1):
            cell = ws_prn.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        prn_row = 2
        for dev_name in sorted(printers_db.keys()):
            pinfo = printers_db[dev_name]
            grp   = pinfo.get("group", "")
            for p in pinfo.get("printers", []):
                if p.get("is_virtual"):
                    continue
                pname = p.get("name", "") or ""
                if not pname.strip():
                    continue
                ink_parts = []
                for s in p.get("supplies", []):
                    pct = s.get("pct", -1)
                    desc = s.get("desc", "")
                    if pct >= 0:
                        ink_parts.append(f"{desc}: {pct}%")
                ink_str = " | ".join(ink_parts) if ink_parts else "—"
                vals = [
                    dev_name, grp, pname,
                    p.get("driver", ""), p.get("printer_ip", ""),
                    _printer_status_str(p.get("status", 0)).replace("✅ ", "").replace("❌ ", "").replace("⚠️ ", "").replace("🖨 ", "").replace("⏸ ", "").replace("🔌 ", "").replace("⏳ ", ""),
                    "Да" if p.get("default") else "", "Да" if p.get("shared") else "",
                    ink_str,
                ]
                for col_idx, val in enumerate(vals, 1):
                    cell = ws_prn.cell(row=prn_row, column=col_idx, value=_xlsx_safe(val))
                    cell.border = thin_border
                prn_row += 1
        # Auto column widths
        for col in range(1, len(prn_headers) + 1):
            max_len = len(str(prn_headers[col - 1]))
            for row in range(2, ws_prn.max_row + 1):
                val = ws_prn.cell(row=row, column=col).value
                if val:
                    max_len = max(max_len, min(len(str(val)), 50))
            ws_prn.column_dimensions[get_column_letter(col)].width = max_len + 2

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── Network Map Helpers ─────────────────────────────────────────────

def _get_local_ip(d: dict) -> str:
    """Best local IP: 192.168.x.x > 10.x.x.x > 172.x.x.x > other."""
    best, priority = "", 999
    for nic in d.get("nic_details", []):
        for ip in nic.get("ips", []):
            if ip.startswith("127.") or ip.startswith("169.254."):
                continue
            if ip.startswith("192.168."):
                p = 0
            elif ip.startswith("10."):
                p = 1
            elif ip.startswith("172."):
                p = 2
            else:
                p = 3
            if p < priority:
                priority, best = p, ip
    return best


def _os_icon(os_str: str) -> str:
    """Short OS label with emoji."""
    s = (os_str or "").lower()
    if "windows 11" in s:
        return "W11"
    if "windows 10" in s:
        return "W10"
    if "windows" in s:
        return "Win"
    if "linux" in s or "debian" in s or "ubuntu" in s:
        return "Lnx"
    if "mac" in s or "darwin" in s:
        return "Mac"
    return ""


def _os_emoji(os_str: str) -> str:
    """Emoji for OS (HTML map)."""
    s = (os_str or "").lower()
    if "windows" in s:
        return "🪟"
    if "linux" in s or "debian" in s or "ubuntu" in s:
        return "🐧"
    if "mac" in s or "darwin" in s:
        return "🍎"
    return "💻"


def _os_node_color(os_str: str) -> tuple[str, str]:
    """(border_color, bg_color) for HTML node by OS type."""
    s = (os_str or "").lower()
    if "windows" in s:
        return "#0078d4", "#e8f4fd"
    if "linux" in s or "debian" in s or "ubuntu" in s:
        return "#e67e22", "#fef9f0"
    if "mac" in s or "darwin" in s:
        return "#555", "#f5f5f5"
    return "#95a5a6", "#fafafa"


def _fmt_offline(hours: float) -> str:
    """Human-readable offline duration."""
    if hours <= 0:
        return ""
    if hours < 1:
        return f"{int(hours * 60)}м"
    if hours < 24:
        return f"{int(hours)}ч"
    return f"{int(hours / 24)}д"


def _read_vis_js() -> str:
    """Read local vis-network.min.js for embedding."""
    try:
        return Path("/opt/meshcentral-bot/vis-network.min.js").read_text(encoding="utf-8")
    except Exception:
        return ""


def build_network_map(devices: list[dict]) -> bytes | None:
    """Build a professional network map grouped by office/location (ext IP) and MC group."""
    from matplotlib.patches import FancyBboxPatch
    from matplotlib.lines import Line2D

    if not devices:
        return None

    # ── Group devices by external IP (= office/location) ──
    locations: dict[str, dict[str, list[dict]]] = {}
    for d in devices:
        ext_ip = d.get("ip", "") or "Unknown"
        group = d.get("group", "?")
        d["_local_ip"] = _get_local_ip(d)
        locations.setdefault(ext_ip, {}).setdefault(group, []).append(d)

    # ── Office colors ──
    office_colors = [
        ("#1a5276", "#d4e6f1"),  # dark blue / light blue
        ("#6c3483", "#e8daef"),  # purple / light purple
        ("#117a65", "#d1f2eb"),  # teal / light teal
        ("#935116", "#fae5d3"),  # brown / light brown
        ("#1b4f72", "#d6eaf8"),  # navy / light navy
    ]
    group_badge_colors = ["#2980b9", "#8e44ad", "#27ae60", "#e67e22", "#c0392b", "#16a085", "#f39c12"]

    # ── Sort locations: biggest first ──
    sorted_locs = sorted(locations.items(), key=lambda x: -sum(len(v) for v in x[1].values()))

    # ── Calculate layout ──
    CARD_W, CARD_H = 2.8, 0.8
    CARD_PAD = 0.3
    COLS_PER_ROW = 4
    GROUP_PAD = 0.6
    LOC_PAD = 1.0
    LEFT_MARGIN = 0.5
    TOP_START = 0.0  # will be computed from top

    # Pre-calculate total height
    total_height = 3.0  # server block + gap
    loc_layouts = []
    for loc_ip, groups in sorted_locs:
        loc_h = 1.0  # header
        for grp_name, grp_devs in sorted(groups.items()):
            n_devs = len(grp_devs)
            n_rows = (n_devs + COLS_PER_ROW - 1) // COLS_PER_ROW
            loc_h += 0.6 + n_rows * (CARD_H + CARD_PAD) + GROUP_PAD
        loc_layouts.append((loc_ip, groups, loc_h))
        total_height += loc_h + LOC_PAD

    fig_w = max(14, COLS_PER_ROW * (CARD_W + CARD_PAD) + 3)
    fig_h = max(8, total_height + 1)
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.set_xlim(-0.5, fig_w - 0.5)
    ax.set_ylim(-0.5, fig_h + 0.5)
    ax.axis("off")
    fig.patch.set_facecolor("#f8f9fa")
    ax.set_facecolor("#f8f9fa")

    # ── Draw MeshCentral Server block at top ──
    srv_y = fig_h - 1.5
    srv_w, srv_h = 5, 1.2
    srv_x = (fig_w - srv_w) / 2
    srv_box = FancyBboxPatch((srv_x, srv_y - srv_h / 2), srv_w, srv_h,
                              boxstyle="round,pad=0.15", facecolor="#2c3e50",
                              edgecolor="#1a252f", linewidth=2)
    ax.add_patch(srv_box)
    ax.text(srv_x + srv_w / 2, srv_y + 0.15, "MeshCentral Server",
            fontsize=11, fontweight="bold", color="white", ha="center", va="center")
    mc_host = MC_URL.replace("https://", "").replace("http://", "").rstrip("/")
    ax.text(srv_x + srv_w / 2, srv_y - 0.25, mc_host,
            fontsize=8, color="#bdc3c7", ha="center", va="center", fontfamily="monospace")

    # ── Draw each location ──
    cur_y = srv_y - srv_h / 2 - 1.5
    n_online = sum(1 for d in devices if d.get("online"))
    n_total = len(devices)

    for li, (loc_ip, groups, loc_h) in enumerate(loc_layouts):
        ci = li % len(office_colors)
        hdr_color, bg_color = office_colors[ci]

        # Count devices in this location
        loc_devs = sum(len(v) for v in groups.values())
        loc_online = sum(1 for grp in groups.values() for d in grp if d.get("online"))

        # Determine location subnet from local IPs
        local_subnets = set()
        for grp_devs in groups.values():
            for d in grp_devs:
                lip = d.get("_local_ip", "")
                if lip:
                    try:
                        local_subnets.add(str(ipaddress.ip_network(f"{lip}/24", strict=False)))
                    except (ValueError, TypeError):
                        pass
        subnet_str = ", ".join(sorted(local_subnets)) if local_subnets else "?"

        # Location background
        loc_box = FancyBboxPatch((LEFT_MARGIN - 0.3, cur_y - loc_h + 0.5), fig_w - 1.2, loc_h,
                                  boxstyle="round,pad=0.2", facecolor=bg_color,
                                  edgecolor=hdr_color, linewidth=1.5, alpha=0.6)
        ax.add_patch(loc_box)

        # Connection line from server to location
        line_x = fig_w / 2
        ax.annotate("", xy=(line_x, cur_y + 0.3), xytext=(line_x, cur_y + 1.0),
                     arrowprops=dict(arrowstyle="-|>", color=hdr_color, lw=1.5))

        # Location header
        ax.text(LEFT_MARGIN, cur_y, f"WAN: {loc_ip}", fontsize=10, fontweight="bold",
                color=hdr_color, va="center", fontfamily="monospace")
        ax.text(LEFT_MARGIN, cur_y - 0.35, f"LAN: {subnet_str}  |  {loc_online}/{loc_devs} online",
                fontsize=8, color="#555", va="center")

        # Draw groups inside location
        inner_y = cur_y - 0.8
        for gi, (grp_name, grp_devs) in enumerate(sorted(groups.items())):
            badge_color = group_badge_colors[gi % len(group_badge_colors)]

            # Group badge
            badge_w = len(grp_name) * 0.12 + 0.5
            badge = FancyBboxPatch((LEFT_MARGIN, inner_y - 0.15), badge_w, 0.35,
                                    boxstyle="round,pad=0.08", facecolor=badge_color,
                                    edgecolor="none", alpha=0.85)
            ax.add_patch(badge)
            ax.text(LEFT_MARGIN + badge_w / 2, inner_y + 0.03, grp_name,
                    fontsize=8, fontweight="bold", color="white", ha="center", va="center")

            inner_y -= 0.55

            # Draw device cards in grid
            sorted_devs = sorted(grp_devs, key=lambda x: x.get("_local_ip", ""))
            for di, d in enumerate(sorted_devs):
                col = di % COLS_PER_ROW
                row = di // COLS_PER_ROW
                cx = LEFT_MARGIN + col * (CARD_W + CARD_PAD)
                cy = inner_y - row * (CARD_H + CARD_PAD)

                # Card background
                is_on = d.get("online", False)
                off_h = d.get("offline_hours", 0)
                is_stale = (not is_on) and off_h > 7 * 24  # >7 days
                if is_on:
                    card_face, card_edge, card_lw = "#ffffff", "#2ecc71", 1.5
                elif is_stale:
                    card_face, card_edge, card_lw = "#f0f0f0", "#aaaaaa", 0.6
                else:
                    card_face, card_edge, card_lw = "#fff5f5", "#e74c3c", 0.9

                card = FancyBboxPatch((cx, cy - CARD_H), CARD_W, CARD_H,
                                       boxstyle="round,pad=0.08", facecolor=card_face,
                                       edgecolor=card_edge, linewidth=card_lw)
                ax.add_patch(card)

                # Status dot
                dot_color = "#2ecc71" if is_on else ("#aaaaaa" if is_stale else "#e74c3c")
                dot = plt.Circle((cx + 0.18, cy - 0.2), 0.08, color=dot_color, ec="none")
                ax.add_patch(dot)

                # Device name (truncate if needed)
                dev_name = d.get("name", "?")
                if len(dev_name) > 22:
                    dev_name = dev_name[:20] + ".."
                ax.text(cx + 0.35, cy - 0.22, dev_name,
                        fontsize=7, fontweight="bold", color="#2c3e50", va="center",
                        clip_on=True)

                # Local IP
                lip = d.get("_local_ip", "")
                if lip:
                    ax.text(cx + 0.18, cy - 0.5, lip,
                            fontsize=6.5, color="#7f8c8d", va="center", fontfamily="monospace")

                # OS label (right side)
                os_short = _os_icon(d.get("os", ""))
                if os_short:
                    os_colors = {"W11": "#0078d4", "W10": "#0078d4", "Win": "#0078d4",
                                 "Lnx": "#e67e22", "Mac": "#555"}
                    ax.text(cx + CARD_W - 0.12, cy - 0.22, os_short,
                            fontsize=6, fontweight="bold",
                            color=os_colors.get(os_short, "#95a5a6"),
                            va="center", ha="right")

                # Offline duration (bottom-right of card)
                if not is_on and off_h > 0:
                    off_str = _fmt_offline(off_h)
                    ax.text(cx + CARD_W - 0.12, cy - 0.52, off_str,
                            fontsize=5.5, color="#e74c3c" if not is_stale else "#aaaaaa",
                            va="center", ha="right")

            n_rows = (len(sorted_devs) + COLS_PER_ROW - 1) // COLS_PER_ROW
            inner_y -= n_rows * (CARD_H + CARD_PAD) + GROUP_PAD

        cur_y -= loc_h + LOC_PAD

    # ── Legend ──
    n_stale = sum(1 for d in devices if not d.get("online") and d.get("offline_hours", 0) > 7 * 24)
    legend_y = fig_h - 0.15
    legend_elements = [
        Line2D([0], [0], marker="o", color="w", markerfacecolor="#2ecc71",
               markersize=8, label=f"Online ({n_online})"),
        Line2D([0], [0], marker="o", color="w", markerfacecolor="#e74c3c",
               markersize=8, label=f"Offline ({n_total - n_online - n_stale})"),
        Line2D([0], [0], marker="o", color="w", markerfacecolor="#aaaaaa",
               markersize=8, label=f"Stale >7д ({n_stale})"),
        Line2D([0], [0], marker="s", color="w", markerfacecolor="#0078d4",
               markersize=8, label="Windows"),
        Line2D([0], [0], marker="s", color="w", markerfacecolor="#e67e22",
               markersize=8, label="Linux"),
    ]
    ax.legend(handles=legend_elements, loc="upper right", fontsize=8,
              framealpha=0.9, edgecolor="#ccc", fancybox=True,
              bbox_to_anchor=(1.0, (legend_y + 0.5) / fig_h))

    # Title + timestamp
    now_str = datetime.now().strftime("%d.%m.%Y %H:%M")
    ax.text(fig_w / 2, fig_h + 0.2, "Network Map — MeshCentral",
            fontsize=14, fontweight="bold", color="#2c3e50", ha="center", va="center")
    ax.text(fig_w / 2, fig_h - 0.15,
            f"Устройств: {n_total}  |  Online: {n_online}  |  Локации: {len(sorted_locs)}  |  {now_str}",
            fontsize=9, color="#7f8c8d", ha="center", va="center")

    fig.tight_layout(pad=0.5)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def build_network_map_html(devices: list[dict], web_mode: bool = False) -> str | None:  # noqa: C901
    """Build interactive SVG network topology map (pan/zoom/click). No external deps."""
    import json as _json
    import math as _math
    if not devices:
        return None

    def _xe(s: str) -> str:
        return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")

    # Group by MeshCentral group (= office / location name)
    locations: dict[str, list[dict]] = {}
    for d in devices:
        group = d.get("group", "?") or "?"
        d["_local_ip"] = _get_local_ip(d)
        locations.setdefault(group, []).append(d)

    n_online = sum(1 for d in devices if d.get("online"))
    n_offline = sum(1 for d in devices if not d.get("online") and d.get("offline_hours", 0) <= 7 * 24)
    n_stale = sum(1 for d in devices if not d.get("online") and d.get("offline_hours", 0) > 7 * 24)
    n_total = len(devices)
    now_str = datetime.now().strftime("%d.%m.%Y %H:%M")
    n_win = sum(1 for d in devices if "windows" in (d.get("os") or "").lower())
    n_lnx = sum(1 for d in devices if any(x in (d.get("os") or "").lower() for x in ("linux", "ubuntu", "debian")))
    mc_host = MC_URL.replace("https://", "").replace("http://", "").rstrip("/")

    loc_colors = ["#3498db", "#9b59b6", "#2ecc71", "#e67e22", "#e74c3c", "#1abc9c", "#f39c12", "#00b4d8"]
    sorted_locs = sorted(locations.items(), key=lambda x: -len(x[1]))
    n_locs = len(sorted_locs)

    # ── Layout constants ────────────────────────────────────────────────
    MARGIN    = 60
    SRV_W, SRV_H  = 240, 60
    LOC_W, LOC_H  = 200, 58
    DEV_W, DEV_H  = 172, 80
    DEV_GAP_X     = 14
    DEV_GAP_Y     = 10
    LOC_GAP       = 36
    DEVS_PER_ROW  = 4
    Y_SRV         = MARGIN + SRV_H // 2          # center y of server node
    Y_LOC         = Y_SRV + SRV_H // 2 + 110     # center y of location nodes
    Y_DEV_TOP     = Y_LOC + LOC_H // 2 + 80      # top y of first device row

    # Per-location geometry
    loc_meta: list[dict] = []
    for li, (grp_name, grp_devs) in enumerate(sorted_locs):
        lc = loc_colors[li % len(loc_colors)]
        all_devs: list[dict] = sorted(grp_devs, key=lambda d: d.get("_local_ip", ""))
        n = len(all_devs)
        n_rows  = max(1, (n + DEVS_PER_ROW - 1) // DEVS_PER_ROW)
        max_col = min(n, DEVS_PER_ROW)
        col_w   = max(LOC_W, max_col * DEV_W + (max_col - 1) * DEV_GAP_X)
        dev_h   = n_rows * DEV_H + (n_rows - 1) * DEV_GAP_Y

        subnets: set[str] = set()
        wan_ips: set[str] = set()
        for d in all_devs:
            lip = d.get("_local_ip", "")
            if lip:
                try:
                    subnets.add(str(ipaddress.ip_network(f"{lip}/24", strict=False)))
                except Exception:
                    pass
            w = d.get("ip", "")
            if w:
                wan_ips.add(w)
        loc_online = sum(1 for d in all_devs if d.get("online"))
        loc_meta.append({
            "name": grp_name, "wan_ips": sorted(wan_ips), "color": lc, "devs": all_devs,
            "n": n, "col_w": col_w, "dev_h": dev_h,
            "subnets": sorted(subnets), "online": loc_online,
        })

    # X positions (centered)
    total_w = sum(m["col_w"] for m in loc_meta) + (n_locs - 1) * LOC_GAP
    canvas_w = max(SRV_W + 100, total_w + 2 * MARGIN)
    x0 = (canvas_w - total_w) // 2
    xc = x0
    for m in loc_meta:
        m["cx"] = xc + m["col_w"] // 2
        xc += m["col_w"] + LOC_GAP

    srv_cx = canvas_w // 2
    max_dev_h = max((m["dev_h"] for m in loc_meta), default=0)
    canvas_h = Y_DEV_TOP + max_dev_h + MARGIN

    # ── SVG element builder ──────────────────────────────────────────────
    parts: list[str] = []

    def bez(x1: int, y1: int, x2: int, y2: int, color: str, w: float = 2, dash: str = "") -> str:
        cx, cy = (x1 + x2) // 2, (y1 + y2) // 2
        ctrl = abs(y2 - y1) // 2
        da = f' stroke-dasharray="{dash}"' if dash else ""
        return (f'<path d="M{x1},{y1} C{x1},{y1+ctrl} {x2},{y2-ctrl} {x2},{y2}"'
                f' stroke="{color}" stroke-width="{w}" fill="none" stroke-opacity="0.6"{da}/>')

    def rect_node(x: int, y: int, w: int, h: int, fill: str, stroke: str, sw: float = 2.0,
                  rx: int = 10) -> str:
        return (f'<rect x="{x - w//2}" y="{y - h//2}" width="{w}" height="{h}"'
                f' rx="{rx}" fill="{fill}" stroke="{stroke}" stroke-width="{sw}"/>')

    def txt(x: int, y: int, s: str, fill: str, sz: int, anchor: str = "middle",
            bold: bool = False, mono: bool = False) -> str:
        fw = ' font-weight="bold"' if bold else ""
        ff = ' font-family="monospace"' if mono else ""
        return (f'<text x="{x}" y="{y}" text-anchor="{anchor}" fill="{fill}"'
                f' font-size="{sz}"{fw}{ff}>{_xe(s)}</text>')

    # ── Edges (drawn first, behind nodes) ───────────────────────────────
    for m in loc_meta:
        lc = m["color"]
        parts.append(bez(srv_cx, Y_SRV + SRV_H // 2,
                         m["cx"], Y_LOC - LOC_H // 2, lc, 2.5, "8,5"))
        for di, d in enumerate(m["devs"]):
            row = di // DEVS_PER_ROW
            col = di % DEVS_PER_ROW
            n_in_row = min(DEVS_PER_ROW, m["n"] - row * DEVS_PER_ROW)
            row_w = n_in_row * DEV_W + (n_in_row - 1) * DEV_GAP_X
            dx = m["cx"] - row_w // 2 + col * (DEV_W + DEV_GAP_X) + DEV_W // 2
            dy = Y_DEV_TOP + row * (DEV_H + DEV_GAP_Y) + DEV_H // 2
            parts.append(bez(m["cx"], Y_LOC + LOC_H // 2,
                             dx, dy - DEV_H // 2, lc, 1.2))

    # ── Server node ──────────────────────────────────────────────────────
    srv_info = _json.dumps({
        "type": "server", "name": "MeshCentral Server",
        "host": mc_host, "url": MC_URL,
        "total": n_total, "online": n_online, "locs": n_locs,
    })
    parts.append(f'<g class="nd" data-i="{_xe(srv_info)}">')
    parts.append(rect_node(srv_cx, Y_SRV, SRV_W, SRV_H, "#0d1b2a", "#3498db", 2.5))
    parts.append(txt(srv_cx, Y_SRV - 12, "🖥 MeshCentral", "#85c1e9", 13, bold=True))
    parts.append(txt(srv_cx, Y_SRV + 5, mc_host, "#5dade2", 10, mono=True))
    parts.append(txt(srv_cx, Y_SRV + 20, f"{n_total} устройств  ·  {n_online} online", "#6b8fa8", 10))
    parts.append('</g>')

    # ── Location nodes ───────────────────────────────────────────────────
    for m in loc_meta:
        lc = m["color"]
        sn = " · ".join(m["subnets"]) if m["subnets"] else ""
        wan_str = ", ".join(m["wan_ips"]) if m["wan_ips"] else ""
        loc_info = _json.dumps({
            "type": "loc", "name": m["name"], "wan": wan_str,
            "subnets": m["subnets"], "total": m["n"], "online": m["online"],
        })
        parts.append(f'<g class="nd" data-i="{_xe(loc_info)}">')
        parts.append(rect_node(m["cx"], Y_LOC, LOC_W, LOC_H, "#0f1c29", lc, 2.5))
        parts.append(txt(m["cx"], Y_LOC - 12, f"📍 {m['name']}", lc, 12, bold=True))
        parts.append(txt(m["cx"], Y_LOC + 6,
                         f"{m['online']}/{m['n']} online", "#8fb3cc", 10))
        if sn:
            parts.append(txt(m["cx"], Y_LOC + 21, sn[:38], "#4a7a99", 9, mono=True))
        parts.append('</g>')

    # ── Device nodes ─────────────────────────────────────────────────────
    for m in loc_meta:
        lc = m["color"]
        for di, d in enumerate(m["devs"]):
            row = di // DEVS_PER_ROW
            col = di % DEVS_PER_ROW
            n_in_row = min(DEVS_PER_ROW, m["n"] - row * DEVS_PER_ROW)
            row_w = n_in_row * DEV_W + (n_in_row - 1) * DEV_GAP_X
            dx = m["cx"] - row_w // 2 + col * (DEV_W + DEV_GAP_X) + DEV_W // 2
            dy = Y_DEV_TOP + row * (DEV_H + DEV_GAP_Y) + DEV_H // 2

            is_on  = d.get("online", False)
            off_h  = d.get("offline_hours", 0)
            is_st  = not is_on and off_h > 7 * 24
            name   = str(d.get("name", "?") or "?")
            lip    = d.get("_local_ip", "")
            os_s   = str(d.get("os", "") or "")
            cpu    = str(d.get("cpu", "-") or "-")
            ram    = str(d.get("ram_total", "-") or "-")
            grp    = d.get("group", "?")
            drives_list = d.get("drives", [])
            drives_s = "; ".join(drives_list) if isinstance(drives_list, list) else str(drives_list)

            if is_on:
                os_border, _ = _os_node_color(os_s)
                bg, bc = "#081c2e", os_border
            elif is_st:
                bg, bc = "#111114", "#555"
            else:
                bg, bc = "#1c0909", "#c0392b"

            status = "🟢" if is_on else ("⚫" if is_st else "🔴")
            os_em  = _os_emoji(os_s)

            _d_macs = [nic.get("mac","") for nic in d.get("nic_details",[])
                       if nic.get("mac","") not in ("","00:00:00:00:00:00")]
            dev_info = _json.dumps({
                "type": "dev", "name": name, "group": grp,
                "wan": d.get("ip", "") or "", "lan": lip, "online": is_on,
                "stale": is_st, "off_h": round(off_h, 1),
                "os": os_s[:100], "cpu": cpu[:80], "ram": ram,
                "gpu": str(d.get("gpu", "-") or "-")[:80],
                "drives": drives_s[:200],
                "av": str(d.get("antivirus", "-") or "-")[:80],
                "boot": str(d.get("last_boot", "-") or "-"),
                "agent": str(d.get("agent_ver", "-") or "-"),
                "mac": _d_macs[0] if _d_macs else "",
                "mc_id": str(d.get("id", "") or ""),
                "mc_url": MC_URL,
            })

            parts.append(f'<g class="nd" data-i="{_xe(dev_info)}">')
            parts.append(rect_node(dx, dy, DEV_W, DEV_H, bg, bc, 1.8))
            lbl = f"{status}{os_em} {name[:20]}"
            parts.append(txt(dx, dy - 24, lbl, "#d0e4f7", 11, bold=True))
            if lip:
                parts.append(txt(dx, dy - 9, lip, "#5dade2", 9, mono=True))
            if os_s:
                parts.append(txt(dx, dy + 6, os_s[:30], "#7a9ab8", 9))
            if not is_on and off_h > 0:
                parts.append(txt(dx, dy + 21, f"⏱ {_fmt_offline(off_h)} назад", "#e74c3c", 9))
            elif cpu and cpu != "-":
                hw = f"{cpu[:22]}  ·  {ram}" if ram and ram != "-" else cpu[:30]
                parts.append(txt(dx, dy + 21, hw[:36], "#4a7a99", 9))
            parts.append('</g>')

    # ── Printer nodes: horizontal rows per location, below all devices ───
    PRN_W, PRN_H   = 150, 46
    PRN_GAP_X      = 10
    PRN_GAP_Y      = 8
    PRN_PER_ROW    = 4
    PRN_TOP_MARGIN = 28   # gap from bottom of device section to printer label

    _VIRT_KW = ("anydesk", "pdf", "xps", "microsoft", "onenote", "fax",
                "cutepdf", "adobe", "bullzip", "nitro", "biztalk")

    try:
        _pdb = _load_printers()
    except Exception:
        _pdb = {}

    # Store device center positions for edge drawing: name → (dx, dy)
    _dev_pos: dict[str, tuple[int, int]] = {}
    for m in loc_meta:
        for di, d in enumerate(m["devs"]):
            _ri = di // DEVS_PER_ROW
            _ci = di % DEVS_PER_ROW
            _nir = min(DEVS_PER_ROW, m["n"] - _ri * DEVS_PER_ROW)
            _rw = _nir * DEV_W + (_nir - 1) * DEV_GAP_X
            _dx = m["cx"] - _rw // 2 + _ci * (DEV_W + DEV_GAP_X) + DEV_W // 2
            _dy = Y_DEV_TOP + _ri * (DEV_H + DEV_GAP_Y) + DEV_H // 2
            _dev_pos[d.get("name", "")] = (_dx, _dy)

    # Per-location: collect real printers, deduplicate by (host, name)
    _loc_printers: dict[str, list[dict]] = {}  # loc_name → [{pp, host, name}]
    for _pc, _pinfo in sorted(_pdb.items()):
        _grp = _pinfo.get("group", "")
        _seen = {(item["host"], item["name"]) for item in _loc_printers.get(_grp, [])}
        for _pp in _pinfo.get("printers", []):
            _pn = (_pp.get("name") or "").strip()
            if not _pn or (_pc, _pn) in _seen:
                continue
            _drv = (_pp.get("driver") or "").lower()
            if _pp.get("is_virtual") or any(k in _drv for k in _VIRT_KW):
                continue
            _loc_printers.setdefault(_grp, []).append({"name": _pn, "pp": _pp, "host": _pc})
            _seen.add((_pc, _pn))

    _prn_section_h = 0   # extra canvas height for all printer sections
    for m in loc_meta:
        loc_prns = _loc_printers.get(m["name"], [])
        if not loc_prns:
            continue

        # Base Y: below the tallest device section (global baseline keeps all columns aligned)
        base_y = Y_DEV_TOP + max_dev_h + PRN_TOP_MARGIN
        _prn_section_h = max(_prn_section_h, PRN_TOP_MARGIN)

        # Separator label
        parts.append(txt(m["cx"], base_y, "── 🖨 Принтеры ──", "#7a3aaa", 9))

        n_prn = len(loc_prns)
        for pi, item in enumerate(loc_prns):
            row = pi // PRN_PER_ROW
            col = pi % PRN_PER_ROW
            n_in_row = min(PRN_PER_ROW, n_prn - row * PRN_PER_ROW)
            row_w = n_in_row * PRN_W + (n_in_row - 1) * PRN_GAP_X
            pnx = m["cx"] - row_w // 2 + col * (PRN_W + PRN_GAP_X) + PRN_W // 2
            pny = base_y + 14 + row * (PRN_H + PRN_GAP_Y) + PRN_H // 2

            pp  = item["pp"]
            prn_status = pp.get("status", 0)
            prn_ok     = (prn_status == 0)
            pstroke    = "#7a2aaa" if prn_ok else "#aa2a2a"
            pbg        = "#160820" if prn_ok else "#250d10"

            # Edge: host PC → printer node
            host_pos = _dev_pos.get(item["host"], None)
            if host_pos:
                hx, hy = host_pos
                parts.append(bez(hx, hy + DEV_H // 2, pnx, pny - PRN_H // 2,
                                  "#6a3a9a", 1.0, "2,4"))

            prn_info = _json.dumps({
                "type": "printer_node", "name": item["name"],
                "host": item["host"], "ip": pp.get("printer_ip", ""),
                "driver": pp.get("driver", ""), "status": prn_status,
                "default": pp.get("default", False),
            })
            parts.append(f'<g class="nd" data-i="{_xe(prn_info)}">')
            # Node body
            parts.append(f'<rect x="{pnx-PRN_W//2}" y="{pny-PRN_H//2}" '
                          f'width="{PRN_W}" height="{PRN_H}" rx="7" '
                          f'fill="{pbg}" stroke="{pstroke}" stroke-width="1.8"/>')
            # Top accent strip
            parts.append(f'<rect x="{pnx-PRN_W//2+2}" y="{pny-PRN_H//2}" '
                          f'width="{PRN_W-4}" height="5" rx="4" fill="{pstroke}" opacity="0.7"/>')
            # Printer SVG icon (left side)
            ix = pnx - PRN_W // 2 + 8
            iy = pny - 11
            parts.append(f'<rect x="{ix}" y="{iy+6}" width="18" height="13" rx="2" '
                          f'fill="#3a1060" stroke="#c060f8" stroke-width="1.2"/>')
            parts.append(f'<rect x="{ix+4}" y="{iy+1}" width="10" height="6" rx="1" fill="#c890e8"/>')
            parts.append(f'<rect x="{ix+4}" y="{iy+18}" width="8" height="2" rx="1" fill="#c890e8"/>')
            parts.append(f'<rect x="{ix+4}" y="{iy+22}" width="5" height="1.5" rx="0.5" fill="#9060c0"/>')
            led_c = "#00ff88" if prn_ok else "#ff4444"
            parts.append(f'<circle cx="{ix+15}" cy="{iy+12}" r="2.5" fill="{led_c}"/>')
            # Text: printer name (= model from agent)
            tx0 = pnx - PRN_W // 2 + 32
            parts.append(txt(pnx + 8, pny - 11, item["name"][:22], "#e0b8ff", 9, bold=True))
            # Host PC line
            parts.append(txt(pnx + 8, pny + 2, f"📌 {item['host'][:18]}", "#9a7ac0", 8))
            # IP or USB
            pip = pp.get("printer_ip", "")
            pdef = "⭐ " if pp.get("default") else ""
            if pip:
                parts.append(txt(pnx + 8, pny + 14, f"{pdef}{pip}", "#5dade2", 8, mono=True))
            else:
                parts.append(txt(pnx + 8, pny + 14, f"{pdef}USB/WSD", "#7a5aaa", 8))
            parts.append('</g>')

            # Track height
            _prn_section_h = max(_prn_section_h,
                                  PRN_TOP_MARGIN + (row + 1) * (PRN_H + PRN_GAP_Y) + 14)

    canvas_h += _prn_section_h

    # ── WiFi / LAN client nodes ───────────────────────────────────────────
    # Build lookup: device_name → group_name, also carry router IP
    _load_wifi_clients()
    wifi_by_loc: dict[str, dict] = {}  # group_name → {"router": str|None, "clients": [...]}

    # Build printer info lookup: printer_ip → {name, host_pc} (from printers.json)
    _printer_info_by_ip: dict[str, dict] = {}
    try:
        for _pc_name, _pinfo in _load_printers().items():
            for _pp in _pinfo.get("printers", []):
                _pmd_ip = _pp.get("printer_ip", "")
                _pmd_name = _pp.get("name", "")
                if _pmd_ip and _pmd_name and not _pp.get("is_virtual"):
                    _printer_info_by_ip[_pmd_ip] = {"name": _pmd_name, "host": _pc_name}
    except Exception:
        pass
    if _wifi_clients:
        dev_name_to_grp = {}
        for m in loc_meta:
            for d in m["devs"]:
                dev_name_to_grp[d.get("name", "")] = m["name"]
        for aname, wdata in _wifi_clients.items():
            if not wdata.get("ok"):
                continue
            grp = dev_name_to_grp.get(aname)
            if grp:
                entry = wifi_by_loc.setdefault(grp, {"router": None, "clients": []})
                entry["clients"].extend(wdata.get("clients") or [])
                if not entry["router"] and wdata.get("router"):
                    entry["router"] = wdata["router"]

    WIFI_W, WIFI_H   = 148, 58
    WIFI_GAP_X       = 10
    WIFI_GAP_Y       = 8
    WIFI_PER_ROW     = 4
    WIFI_TOP_MARGIN  = 24   # gap between device rows and separator label
    RTR_W, RTR_H     = 180, 50  # router node size
    RTR_MARGIN       = 16   # gap between router node bottom and wifi clients top

    for m in loc_meta:
        entry = wifi_by_loc.get(m["name"])
        if not entry:
            continue
        wclients_all = entry["clients"]
        rtr_ip       = entry.get("router")
        # Printer-type clients are already shown in the dedicated printer section
        wclients = [c for c in wclients_all if c.get("type") != "printer"]
        if not wclients and not rtr_ip:
            continue

        n_wifi = len(wclients)

        # Base Y: below devices + printer section (global baselines, no per-location overlap)
        base_y = Y_DEV_TOP + max_dev_h + _prn_section_h + WIFI_TOP_MARGIN

        # separator label
        parts.append(txt(m["cx"], base_y - 10, "── 🔌 Сеть офиса ──", "#5a7a9a", 9))

        # ── Router node ──────────────────────────────────────────────────
        rtr_cy = base_y + RTR_H // 2
        if rtr_ip:
            n_w = sum(1 for c in wclients_all if c.get("type") == "wifi")
            n_l = sum(1 for c in wclients_all if c.get("type") == "lan")
            n_p = sum(1 for c in wclients_all if c.get("type") == "printer")
            rtr_info = _json.dumps({
                "type": "router", "ip": rtr_ip, "location": m["name"],
                "clients_wifi": n_w, "clients_lan": n_l, "clients_printer": n_p,
            })
            # edge: location → router
            parts.append(bez(m["cx"], Y_LOC + LOC_H // 2,
                             m["cx"], rtr_cy - RTR_H // 2, m["color"], 2.0, "6,4"))
            parts.append(f'<g class="nd" data-i="{_xe(rtr_info)}">')
            parts.append(rect_node(m["cx"], rtr_cy, RTR_W, RTR_H, "#081828", "#1a8a9a", 2.0))
            parts.append(txt(m["cx"], rtr_cy - 13, "🌐 Роутер", "#7ecfda", 12, bold=True))
            parts.append(txt(m["cx"], rtr_cy + 4,  rtr_ip, "#5dade2", 10, mono=True))
            summary_parts = []
            if n_w: summary_parts.append(f"📶{n_w}")
            if n_l: summary_parts.append(f"🔌{n_l}")
            if n_p: summary_parts.append(f"🖨{n_p}")
            parts.append(txt(m["cx"], rtr_cy + 18,
                             "  ".join(summary_parts), "#4a8a9a", 9))
            parts.append('</g>')
        else:
            # No router IP: connect location directly to wifi block
            rtr_cy = base_y - RTR_MARGIN  # collapse router space

        # Y where wifi client rows begin
        wifi_y_top = rtr_cy + RTR_H // 2 + RTR_MARGIN

        # ── WiFi/LAN client nodes ─────────────────────────────────────────
        for wi, wc in enumerate(wclients):
            row = wi // WIFI_PER_ROW
            col = wi % WIFI_PER_ROW
            n_in_row = min(WIFI_PER_ROW, n_wifi - row * WIFI_PER_ROW)
            row_w = n_in_row * WIFI_W + (n_in_row - 1) * WIFI_GAP_X
            wx = m["cx"] - row_w // 2 + col * (WIFI_W + WIFI_GAP_X) + WIFI_W // 2
            wy = wifi_y_top + row * (WIFI_H + WIFI_GAP_Y) + WIFI_H // 2

            ctype   = wc.get("type", "lan")
            is_wifi    = (ctype == "wifi")
            is_printer = (ctype == "printer")
            if is_wifi:       cicon = "📶"
            elif is_printer:  cicon = "🖨"
            else:             cicon = "🔌"
            cname   = str(wc.get("name") or wc.get("mac") or "?")
            cip     = str(wc.get("ip") or "")
            cmac    = str(wc.get("mac") or "")
            crssi   = wc.get("rssi")
            cup     = wc.get("online_sec")
            cup_s   = f"⏱ {cup//3600}h{(cup%3600)//60}m" if cup else ""
            crssi_s = f"{crssi}dBm" if crssi is not None else ""

            # Visual style: WiFi=green, LAN=blue, Printer=purple
            if is_wifi:
                cbg, cstroke, clbl = "#0e1f0e", "#4a9a3a", "#90d878"
            elif is_printer:
                cbg, cstroke, clbl = "#1a0a28", "#8a4aaa", "#c890e8"
            else:
                cbg, cstroke, clbl = "#0a1525", "#2a6aaa", "#6ab4e8"

            # edge: router (or location) → client
            src_x = m["cx"]
            src_y = (rtr_cy + RTR_H // 2) if rtr_ip else (Y_LOC + LOC_H // 2)
            parts.append(bez(src_x, src_y, wx, wy - WIFI_H // 2, cstroke, 0.9, "3,4"))

            # Printer: look up model and host PC from printers.json by IP
            pinfo_match = _printer_info_by_ip.get(cip, {}) if is_printer else {}
            pmodel_name = pinfo_match.get("name", "")
            phost_pc    = pinfo_match.get("host", "")

            wifi_info = _json.dumps({
                "type": ctype, "name": cname, "mac": cmac,
                "ip": cip, "iface": wc.get("iface", ""), "kind": ctype,
                "rssi": crssi_s, "uptime": cup_s,
                "model": pmodel_name, "host_pc": phost_pc,
            })

            parts.append(f'<g class="nd" data-i="{_xe(wifi_info)}">')

            if is_printer:
                # ── Printer node: distinct shape + SVG printer icon ──────────
                PW, PH = WIFI_W + 10, WIFI_H + 10   # slightly larger
                px, py = wx - PW // 2, wy - PH // 2
                # Outer border with dashed effect (top accent stripe)
                parts.append(f'<rect x="{px}" y="{py}" width="{PW}" height="{PH}"'
                              f' rx="6" fill="#1a0828" stroke="#9a3acc" stroke-width="2"/>')
                # Top accent bar (paper output indicator)
                parts.append(f'<rect x="{px}" y="{py}" width="{PW}" height="7"'
                              f' rx="6" fill="#6a1a9a"/>')
                parts.append(f'<rect x="{px}" y="{py+4}" width="{PW}" height="3" fill="#6a1a9a"/>')
                # SVG printer icon (left side, 22x18px)
                ix, iy = px + 6, py + 10
                # Printer body
                parts.append(f'<rect x="{ix}" y="{iy+5}" width="20" height="13" rx="2"'
                              f' fill="#4a1a7a" stroke="#b060e8" stroke-width="1.2"/>')
                # Paper in slot (top)
                parts.append(f'<rect x="{ix+4}" y="{iy}" width="12" height="6" rx="1" fill="#c890e8"/>')
                # Paper out slot (bottom lines)
                parts.append(f'<rect x="{ix+4}" y="{iy+17}" width="12" height="2" rx="1" fill="#c890e8"/>')
                parts.append(f'<rect x="{ix+4}" y="{iy+21}" width="8" height="2" rx="1" fill="#a070c8"/>')
                # LED dot on printer body
                parts.append(f'<circle cx="{ix+17}" cy="{iy+10}" r="2" fill="#00ff88"/>')
                # Printer label: model name
                model_lbl = (pmodel_name[:19] if pmodel_name else cname[:19])
                parts.append(txt(wx + 14, wy - 13, model_lbl, "#d8a8f8", 9, bold=True))
                if cip:
                    parts.append(txt(wx + 14, wy + 1, cip, "#5dade2", 8, mono=True))
                if phost_pc:
                    parts.append(txt(wx + 14, wy + 13, f"via {phost_pc[:14]}", "#8a6aaa", 8))
                elif cmac:
                    parts.append(txt(wx + 14, wy + 13, cmac[:17], "#4a3a6a", 8, mono=True))
            else:
                # ── WiFi / LAN node (unchanged) ──────────────────────────────
                parts.append(f'<rect x="{wx - WIFI_W//2}" y="{wy - WIFI_H//2}" width="{WIFI_W}" height="{WIFI_H}"'
                             f' rx="8" fill="{cbg}" stroke="{cstroke}" stroke-width="1.4"/>')
                lbl = f"{cicon} {cname[:18]}"
                parts.append(txt(wx, wy - 18, lbl, clbl, 10, bold=True))
                if cip:
                    parts.append(txt(wx, wy - 4, cip, "#5dade2", 9, mono=True))
                if is_wifi and crssi_s:
                    parts.append(txt(wx, wy + 10, crssi_s, "#a0c090", 9))
                else:
                    parts.append(txt(wx, wy + 10, "LAN", "#4a7aaa", 9))
                if cup_s:
                    parts.append(txt(wx, wy + 22, cup_s, "#708060", 9))
                elif cmac:
                    parts.append(txt(wx, wy + 22, cmac[:17], "#4a5a6a", 8, mono=True))
            parts.append('</g>')

    # Recalculate canvas height to include router + wifi rows
    extra_h = 0
    for m in loc_meta:
        entry = wifi_by_loc.get(m["name"])
        if entry and (entry["clients"] or entry.get("router")):
            n_wifi = sum(1 for c in entry["clients"] if c.get("type") != "printer")
            n_rows_w = max(1, (n_wifi + WIFI_PER_ROW - 1) // WIFI_PER_ROW)
            block_h = (RTR_H + RTR_MARGIN +
                       n_rows_w * WIFI_H + (n_rows_w - 1) * WIFI_GAP_Y +
                       WIFI_TOP_MARGIN + 20)
            extra_h = max(extra_h, block_h)
    canvas_h = canvas_h + extra_h

    svg_body = "\n".join(parts)

    # ── Full HTML ────────────────────────────────────────────────────────
    html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Network Map — MeshCentral</title>
{'<meta http-equiv="cache-control" content="no-cache"><meta http-equiv="pragma" content="no-cache">' if web_mode else ''}
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
html,body{{width:100%;height:100%;overflow:hidden;background:#0a1120;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;color:#c8d6e5}}
#hdr{{background:linear-gradient(135deg,#0d1b2a,#172535);padding:8px 14px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:4px;border-bottom:1px solid #1e3a5f;position:relative;z-index:10}}
#hdr h1{{font-size:14px;font-weight:700;color:#85c1e9}}
.st{{font-size:11px;display:flex;gap:8px;flex-wrap:wrap;align-items:center;color:#8fb3cc}}
#srch{{background:#0a1628;border:1px solid #1e3a5f;border-radius:5px;color:#c8d6e5;padding:3px 8px;font-size:12px;width:160px;outline:none}}
#srch:focus{{border-color:#3498db}}
#srch-count{{font-size:10px;color:#4a7a99;min-width:60px}}
.nd.nd-dim rect,.nd.nd-dim text,.nd.nd-dim circle{{opacity:0.18}}
.nd.nd-dim{{pointer-events:none}}
.nd.nd-match rect{{filter:drop-shadow(0 0 6px #f39c12) brightness(1.3)}}
.nd.nd-match text{{fill:#fff!important}}
.dot{{display:inline-block;width:7px;height:7px;border-radius:50%;margin-right:2px;vertical-align:middle}}
.don{{background:#2ecc71}}.doff{{background:#e74c3c}}.dst{{background:#555}}
.osbg{{background:#ffffff12;border-radius:3px;padding:0 5px;font-size:10px}}
#wrap{{width:100%;height:calc(100vh - 44px);overflow:hidden;cursor:grab;user-select:none;-webkit-user-select:none;-webkit-touch-callout:none;touch-action:none}}
#wrap.dragging{{cursor:grabbing}}
.nd{{cursor:pointer}}
.nd rect{{transition:filter .15s}}
.nd:hover rect{{filter:brightness(1.35)}}
#panel{{position:fixed;top:50px;right:14px;width:280px;background:#0f1e2e;border:1px solid #1e3a5f;border-radius:10px;padding:14px;font-size:12px;display:none;z-index:20;box-shadow:0 4px 24px #000a;max-height:calc(100vh - 70px);overflow-y:auto}}
#panel h2{{font-size:13px;color:#aed6f1;margin-bottom:8px;padding-right:20px;word-break:break-all}}
.row{{display:flex;gap:6px;margin:4px 0;align-items:flex-start}}
.lbl{{color:#6b8fa8;min-width:70px;font-size:11px;flex-shrink:0;padding-top:1px}}
.val{{color:#c8d6e5;word-break:break-all;flex:1}}
.xbtn{{position:absolute;top:12px;right:12px;cursor:pointer;font-size:16px;color:#6b8fa8;line-height:1}}
.xbtn:hover{{color:#aed6f1}}
.son{{color:#2ecc71;font-weight:600}}.soff{{color:#e74c3c;font-weight:600}}.sst{{color:#777}}
#hint{{position:fixed;bottom:12px;left:50%;transform:translateX(-50%);background:#0f1e2e88;border:1px solid #1e3a5f;border-radius:6px;padding:4px 12px;font-size:10px;color:#4a7a99;pointer-events:none}}
#reload-badge{{position:fixed;bottom:12px;right:14px;background:#0f1e2e88;border:1px solid #1e3a5f;border-radius:6px;padding:4px 10px;font-size:10px;color:#4a7a99;pointer-events:none;display:none}}
@media(max-width:640px){{
  #hdr{{padding:5px 8px}}
  #hdr h1{{font-size:11px}}
  .st{{font-size:10px;gap:3px}}
  #srch{{width:90px;font-size:10px}}
  #wrap{{height:calc(100dvh - 48px)}}
  #panel{{width:calc(100vw - 16px)!important;right:8px!important;left:8px!important;top:auto!important;bottom:6px;max-height:52vh;border-radius:12px 12px 0 0}}
  #hint{{display:none}}
  #reload-badge{{bottom:auto;top:6px;right:8px;font-size:9px;padding:2px 6px}}
}}
</style>
</head>
<body>
<div id="hdr">
  <h1>🗺 Network Map — MeshCentral</h1>
  <div class="st">
    <span><span class="dot don"></span>{n_online} online</span>
    <span><span class="dot doff"></span>{n_offline} offline</span>
    <span><span class="dot dst"></span>{n_stale} stale</span>
    <span>|</span>
    <span>🪟<span class="osbg">{n_win}</span></span>
    <span>🐧<span class="osbg">{n_lnx}</span></span>
    <span>💻 {n_total}</span>
    <span>|</span>
    <span>📍 {n_locs} лок.</span>
    <span>🕐 {now_str}</span>
    <span>|</span>
    <input id="srch" type="search" placeholder="🔍 Поиск устройства…" autocomplete="off">
    <span id="srch-count"></span>
  </div>
</div>
<div id="wrap">
  <svg id="cvs" xmlns="http://www.w3.org/2000/svg"
       width="100%" height="100%"
       viewBox="0 0 {canvas_w} {canvas_h}"
       preserveAspectRatio="none"
       data-w="{canvas_w}" data-h="{canvas_h}">
    <rect width="{canvas_w}" height="{canvas_h}" fill="#0a1120"/>
{svg_body}
  </svg>
</div>
<div id="panel">
  <span class="xbtn" id="xbtn">✕</span>
  <h2 id="pname"></h2>
  <div id="pbody"></div>
</div>
<div id="hint">Колёсико: зум · Тащи фон: перемещение · Клик на узел: детали</div>
<div id="reload-badge"></div>
<script>
(function(){{
  var wrap=document.getElementById('wrap'),
      cvs=document.getElementById('cvs'),
      panel=document.getElementById('panel');
  var cw=parseInt(cvs.getAttribute('data-w')),
      ch=parseInt(cvs.getAttribute('data-h'));
  // viewBox state: vx/vy = top-left in SVG coords; sc = pixels per SVG unit
  var vx=0,vy=0,sc=1;
  function vw(){{return wrap.clientWidth/sc;}}
  function vh(){{return wrap.clientHeight/sc;}}
  function applyVB(){{
    cvs.setAttribute('viewBox',vx+' '+vy+' '+vw()+' '+vh());
  }}
  // ── Fit on load — restore from sessionStorage after auto-reload ───────────────
  function fit(){{
    var saved=sessionStorage.getItem('nm_sc');
    if(saved!==null){{
      sc=parseFloat(saved);
      vx=parseFloat(sessionStorage.getItem('nm_vx')||0);
      vy=parseFloat(sessionStorage.getItem('nm_vy')||0);
      sessionStorage.removeItem('nm_sc');sessionStorage.removeItem('nm_vx');sessionStorage.removeItem('nm_vy');
      applyVB();return;
    }}
    var fw=wrap.clientWidth,fh=wrap.clientHeight;
    var fitSc=Math.min(fw/cw,fh/ch)*0.95;
    sc=fw<640?Math.max(fitSc,fw/(cw*0.35)):fitSc;
    vx=(cw-fw/sc)/2; vy=(ch-fh/sc)/2;
    applyVB();
  }}
  // ── Wheel zoom ───────────────────────────────────────────────────────────────
  wrap.addEventListener('wheel',function(e){{
    e.preventDefault();
    var r=wrap.getBoundingClientRect();
    var mx=e.clientX-r.left,my=e.clientY-r.top;
    var factor=e.deltaY<0?1.15:0.87;
    var nsc=Math.max(0.1,Math.min(sc*factor,10));
    var svgX=vx+mx/sc,svgY=vy+my/sc;
    sc=nsc;vx=svgX-mx/sc;vy=svgY-my/sc;
    applyVB();
  }},{{passive:false}});
  // ── Mouse pan ────────────────────────────────────────────────────────────────
  var mDown=false,mDist=0,mLx=0,mLy=0,mSx=0,mSy=0;
  wrap.addEventListener('mousedown',function(e){{
    if(e.button!==0)return;
    mDown=true;mDist=0;
    mLx=mSx=e.clientX;mLy=mSy=e.clientY;
    wrap.classList.add('dragging');
    e.preventDefault();
  }});
  document.addEventListener('mousemove',function(e){{
    if(!mDown)return;
    var dx=e.clientX-mLx,dy=e.clientY-mLy;
    vx-=dx/sc;vy-=dy/sc;
    mDist+=Math.abs(dx)+Math.abs(dy);
    mLx=e.clientX;mLy=e.clientY;
    applyVB();
  }});
  document.addEventListener('mouseup',function(e){{
    if(!mDown)return;
    mDown=false;wrap.classList.remove('dragging');
    if(mDist<6){{
      var el=document.elementFromPoint(mSx,mSy);
      var nd=el&&el.closest('.nd');
      if(nd)openPanel(nd);else panel.style.display='none';
    }}
  }});
  // ── Touch pan & pinch zoom ───────────────────────────────────────────────────
  var tDown=false,tDist=0,tLx=0,tLy=0,tSx=0,tSy=0;
  var pinching=false,pinchD0=0,pinchSc0=0,pinchVx0=0,pinchVy0=0,pinchMx=0,pinchMy=0;
  wrap.addEventListener('touchstart',function(e){{
    e.preventDefault();
    if(e.touches.length===1){{
      tDown=true;pinching=false;tDist=0;
      tLx=tSx=e.touches[0].clientX;tLy=tSy=e.touches[0].clientY;
    }}else if(e.touches.length===2){{
      tDown=false;pinching=true;
      var a=e.touches[0],b=e.touches[1];
      pinchD0=Math.hypot(b.clientX-a.clientX,b.clientY-a.clientY);
      pinchSc0=sc;pinchVx0=vx;pinchVy0=vy;
      var r=wrap.getBoundingClientRect();
      pinchMx=(a.clientX+b.clientX)/2-r.left;
      pinchMy=(a.clientY+b.clientY)/2-r.top;
    }}
  }},{{passive:false}});
  wrap.addEventListener('touchmove',function(e){{
    e.preventDefault();
    if(tDown&&e.touches.length===1){{
      var dx=e.touches[0].clientX-tLx,dy=e.touches[0].clientY-tLy;
      vx-=dx/sc;vy-=dy/sc;
      tDist+=Math.abs(dx)+Math.abs(dy);
      tLx=e.touches[0].clientX;tLy=e.touches[0].clientY;
      applyVB();
    }}else if(pinching&&e.touches.length===2){{
      var a=e.touches[0],b=e.touches[1];
      var d=Math.hypot(b.clientX-a.clientX,b.clientY-a.clientY);
      var nsc=Math.max(0.1,Math.min(pinchSc0*(d/pinchD0),10));
      var svgX=pinchVx0+pinchMx/pinchSc0,svgY=pinchVy0+pinchMy/pinchSc0;
      sc=nsc;vx=svgX-pinchMx/sc;vy=svgY-pinchMy/sc;
      applyVB();
    }}
  }},{{passive:false}});
  wrap.addEventListener('touchend',function(e){{
    if(tDown&&tDist<8&&e.changedTouches.length){{
      var t=e.changedTouches[0];
      var el=document.elementFromPoint(t.clientX,t.clientY);
      var nd=el&&el.closest('.nd');
      if(nd)openPanel(nd);else panel.style.display='none';
    }}
    tDown=false;pinching=false;
  }});
  // Node panel
  function xe(s){{return s==null||s==='-'?'—':String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}}
  function row(l,v){{return '<div class="row"><span class="lbl">'+l+'</span><span class="val">'+v+'</span></div>';}}
  function fmtH(h){{return h<1?Math.round(h*60)+'м':h<24?Math.round(h)+'ч':Math.round(h/24)+'д';}}
  function openPanel(nd){{
    var i=JSON.parse(nd.getAttribute('data-i'));
    var pn=document.getElementById('pname'),pb=document.getElementById('pbody');
    if(i.type==='server'){{
      pn.textContent='🖥 '+i.name;
      pb.innerHTML=row('Хост','<a href="'+i.url+'" target="_blank" style="color:#3498db">'+xe(i.host)+'</a>')+
        row('Устройств',i.total)+row('Online',i.online)+row('Локаций',i.locs);
    }}else if(i.type==='loc'){{
      pn.textContent='📍 '+i.name;
      pb.innerHTML=row('Online',i.online+'/'+i.total)+
        (i.wan?row('WAN IP',xe(i.wan)):'')+(i.subnets.length?row('LAN',i.subnets.join(', ')):'');
    }}else if(i.type==='router'){{
      pn.textContent='🌐 Роутер — '+i.location;
      pb.innerHTML=row('IP (шлюз)',xe(i.ip))+(i.clients_wifi?row('📶 WiFi',i.clients_wifi):'')+
        (i.clients_lan?row('🔌 LAN',i.clients_lan):'')+
        (i.clients_printer?row('🖨 Принтеры',i.clients_printer):'');
    }}else if(i.type==='wifi'||i.type==='lan'||i.type==='printer'){{
      var icons={{'wifi':'📶','lan':'🔌','printer':'🖨'}};
      pn.textContent=(icons[i.type]||'🔌')+' '+i.name;
      pb.innerHTML=row('IP',xe(i.ip))+(i.mac?row('MAC','<code>'+xe(i.mac)+'</code>'):'')+
        row('Тип',xe(i.kind||i.type))+(i.iface?row('Интерфейс',xe(i.iface)):'')+
        (i.rssi?row('Сигнал',xe(i.rssi)):'')+
        (i.uptime?row('Онлайн',xe(i.uptime)):'');
    }}else{{
      var sc2=i.online?'son':(i.stale?'sst':'soff');
      var st=i.online?'🟢 Online':(i.stale?'⚫ Stale (>7 дней)':'🔴 Offline'+(i.off_h>0?' — '+fmtH(i.off_h)+' назад':''));
      pn.textContent=i.name;
      pb.innerHTML=row('Статус','<span class="'+sc2+'">'+st+'</span>')+
        row('Группа',xe(i.group))+row('WAN',xe(i.wan))+(i.lan?row('LAN',xe(i.lan)):'')+
        (i.mac?row('MAC','<code>'+xe(i.mac)+'</code>'):'')+
        row('OS',xe(i.os))+row('CPU',xe(i.cpu))+row('RAM',xe(i.ram))+
        (i.gpu&&i.gpu!=='-'?row('GPU',xe(i.gpu)):'')+
        row('Диски',xe(i.drives))+row('Антивирус',xe(i.av))+
        row('Загрузка',xe(i.boot))+row('Агент',xe(i.agent))+
        (i.mc_id?row('MC Link','<a href="'+i.mc_url+'" target="_blank" style="color:#3498db">Открыть ↗</a>'):'');
    }}
    panel.style.display='block';
  }}
  document.getElementById('xbtn').onclick=function(){{panel.style.display='none';}};
  window.addEventListener('resize',fit);
  fit();
  // ── Search ──
  var srch=document.getElementById('srch');
  var srchCount=document.getElementById('srch-count');
  function doSearch(){{
    var q=(srch.value||'').trim().toLowerCase();
    var nodes=cvs.querySelectorAll('.nd');
    if(!q){{
      nodes.forEach(function(n){{n.classList.remove('nd-dim','nd-match');}});
      srchCount.textContent='';
      return;
    }}
    var matches=0;
    nodes.forEach(function(n){{
      var info={{}};
      try{{info=JSON.parse(n.getAttribute('data-i'));}}catch(e){{}}
      var name=(info.name||'').toLowerCase();
      var grp=(info.group||'').toLowerCase();
      var ip=(info.wan||info.lan||info.ip||'').toLowerCase();
      var os2=(info.os||'').toLowerCase();
      var hit=name.includes(q)||grp.includes(q)||ip.includes(q)||os2.includes(q);
      if(hit){{n.classList.remove('nd-dim');n.classList.add('nd-match');matches++;}}
      else{{n.classList.remove('nd-match');n.classList.add('nd-dim');}}
    }});
    srchCount.textContent=matches>0?matches+' найд.':'не найдено';
    var first=cvs.querySelector('.nd-match');
    if(first){{
      var r=first.getBoundingClientRect(),wr=wrap.getBoundingClientRect();
      vx-=(wr.left+wr.width/2-(r.left+r.width/2))/sc;
      vy-=(wr.top+wr.height/2-(r.top+r.height/2))/sc;
      applyVB();
    }}
  }}
  srch.addEventListener('input',doSearch);
  srch.addEventListener('keydown',function(e){{
    if(e.key==='Escape'){{srch.value='';doSearch();srch.blur();}}
    if(e.key==='Enter'){{
      var all=Array.from(cvs.querySelectorAll('.nd-match'));
      if(all.length>0){{
        var cur=cvs.querySelector('.nd-match.nd-current');
        var idx=cur?all.indexOf(cur):-1;
        if(cur)cur.classList.remove('nd-current');
        var next=all[(idx+1)%all.length];
        next.classList.add('nd-current');
        var r=next.getBoundingClientRect(),wr=wrap.getBoundingClientRect();
        vx-=(wr.left+wr.width/2-(r.left+r.width/2))/sc;
        vy-=(wr.top+wr.height/2-(r.top+r.height/2))/sc;
        applyVB();
      }}
    }}
  }});
  {'// auto-refresh countdown' if web_mode else '// static mode'}
  {'''var badge=document.getElementById('reload-badge');
  badge.style.display='block';
  var secs=300;
  function tick(){{
    var m=Math.floor(secs/60),s=secs%60;
    badge.textContent='🔄 '+m+'м'+s+'с';
    secs--;
    if(secs<0){{
      sessionStorage.setItem('nm_vx',vx);
      sessionStorage.setItem('nm_vy',vy);
      sessionStorage.setItem('nm_sc',sc);
      location.reload(true);
    }}
  }}
  tick(); setInterval(tick,1000);''' if web_mode else ''}
}})();
</script>
</body>
</html>"""
    return html


def _svg_server_icon() -> str:
    """Inline SVG for server node icon."""
    return (
        "%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 64 64'%3E"
        "%3Crect x='8' y='6' width='48' height='16' rx='3' fill='%232c3e50' stroke='%231a252f' stroke-width='1.5'/%3E"
        "%3Ccircle cx='16' cy='14' r='2' fill='%232ecc71'/%3E"
        "%3Crect x='22' y='12' width='28' height='3' rx='1' fill='%2334495e'/%3E"
        "%3Crect x='8' y='24' width='48' height='16' rx='3' fill='%232c3e50' stroke='%231a252f' stroke-width='1.5'/%3E"
        "%3Ccircle cx='16' cy='32' r='2' fill='%232ecc71'/%3E"
        "%3Crect x='22' y='30' width='28' height='3' rx='1' fill='%2334495e'/%3E"
        "%3Crect x='8' y='42' width='48' height='16' rx='3' fill='%232c3e50' stroke='%231a252f' stroke-width='1.5'/%3E"
        "%3Ccircle cx='16' cy='50' r='2' fill='%23e74c3c'/%3E"
        "%3Crect x='22' y='48' width='28' height='3' rx='1' fill='%2334495e'/%3E"
        "%3C/svg%3E"
    )


def _svg_router_icon(color: str = "#2980b9") -> str:
    """Inline SVG for router/office node icon."""
    c = color.replace("#", "%23")
    return (
        "%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 64 64'%3E"
        f"%3Crect x='4' y='20' width='56' height='28' rx='5' fill='{c}' stroke='%231a252f' stroke-width='1'/%3E"
        "%3Ccircle cx='16' cy='34' r='3' fill='%23ffffff80'/%3E"
        "%3Ccircle cx='28' cy='34' r='3' fill='%23ffffff80'/%3E"
        "%3Crect x='38' y='30' width='16' height='2' rx='1' fill='%23ffffff60'/%3E"
        "%3Crect x='38' y='36' width='12' height='2' rx='1' fill='%23ffffff40'/%3E"
        "%3Cline x1='20' y1='48' x2='20' y2='56' stroke='{c}' stroke-width='2'/%3E"
        "%3Cline x1='32' y1='48' x2='32' y2='56' stroke='{c}' stroke-width='2'/%3E"
        "%3Cline x1='44' y1='48' x2='44' y2='56' stroke='{c}' stroke-width='2'/%3E"
        "%3Cline x1='32' y1='14' x2='32' y2='20' stroke='{c}' stroke-width='2'/%3E"
        "%3Ccircle cx='32' cy='12' r='4' fill='none' stroke='{c}' stroke-width='1.5'/%3E"
        "%3C/svg%3E"
    )


# ─── Utility ─────────────────────────────────────────────────────────

async def mc_is_alive() -> bool:
    ssl_ctx = ssl.create_default_context()
    ssl_ctx.check_hostname = False
    ssl_ctx.verify_mode = ssl.CERT_NONE
    try:
        async with aiohttp.ClientSession() as sess:
            async with sess.get(MC_URL, timeout=aiohttp.ClientTimeout(total=10), ssl=ssl_ctx) as resp:
                return resp.status == 200
    except Exception:
        return False


async def check_http_service(url: str) -> tuple[bool, int | None]:
    """Return (ok, status_code). ok=True if response is 2xx or 3xx."""
    ssl_ctx = ssl.create_default_context()
    ssl_ctx.check_hostname = False
    ssl_ctx.verify_mode = ssl.CERT_NONE
    try:
        async with aiohttp.ClientSession() as sess:
            async with sess.get(url, timeout=aiohttp.ClientTimeout(total=10),
                                ssl=ssl_ctx, allow_redirects=False) as resp:
                return resp.status < 500, resp.status
    except Exception:
        return False, None


async def check_all_http_services() -> list[dict]:
    """Check all HTTP_SERVICES, return list of {name, url, ok, status}."""
    results = []
    for name, url in HTTP_SERVICES:
        ok, status = await check_http_service(url)
        results.append({"name": name, "url": url, "ok": ok, "status": status})
    return results

async def mc_restart():
    proc = await asyncio.create_subprocess_exec(
        "systemctl", "restart", "meshcentral",
        stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
    )
    await proc.wait()

async def check_mc_update() -> dict:
    """Check if a newer MeshCentral version is available on npm.
    Returns {"current": "x.y.z", "latest": "x.y.z", "has_update": bool}.
    """
    current = "unknown"
    latest = "unknown"
    try:
        proc = await asyncio.create_subprocess_exec(
            "node", "-e",
            "console.log(require('/opt/meshcentral/node_modules/meshcentral/package.json').version)",
            stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
        )
        stdout, _ = await asyncio.wait_for(proc.communicate(), timeout=10)
        current = stdout.decode().strip()
    except Exception as e:
        log.error(f"Update check (current): {e}")

    try:
        proc = await asyncio.create_subprocess_exec(
            "npm", "view", "meshcentral", "version",
            stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
        )
        stdout, _ = await asyncio.wait_for(proc.communicate(), timeout=30)
        latest = stdout.decode().strip()
    except Exception as e:
        log.error(f"Update check (npm): {e}")

    has_update = False
    if current != "unknown" and latest != "unknown":
        try:
            from packaging.version import Version
            has_update = Version(latest) > Version(current)
        except Exception:
            has_update = latest != current
    return {"current": current, "latest": latest, "has_update": has_update}


async def perform_mc_update(aid: int):
    """Perform MeshCentral update: backup config, npm update, restart, verify."""
    try:
        await bot.send_message(aid, "🔄 <b>Обновление MeshCentral</b>\n\n1/4 Бэкап конфига...", parse_mode="HTML")
        config_path = f"{MC_DATA}/config.json"
        if os.path.exists(config_path):
            ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            backup_path = f"{MC_DATA}/config_backup_{ts}.json"
            shutil.copy2(config_path, backup_path)

        await bot.send_message(aid, "2/4 npm update meshcentral...", parse_mode="HTML")
        proc = await asyncio.create_subprocess_exec(
            "npm", "update", "meshcentral",
            stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
            cwd=MC_DIR,
        )
        stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=120)
        npm_out = stdout.decode(errors="replace").strip()

        await bot.send_message(aid, "3/4 Перезапуск MeshCentral...", parse_mode="HTML")
        await mc_restart()
        await asyncio.sleep(15)

        alive = await mc_is_alive()
        info = await check_mc_update()

        if alive:
            await bot.send_message(
                aid,
                f"✅ <b>Обновление завершено!</b>\n\n"
                f"📦 Версия: <b>{info['current']}</b>\n"
                f"🟢 MeshCentral работает\n\n"
                f"<pre>{npm_out[:500]}</pre>",
                parse_mode="HTML",
            )
        else:
            await bot.send_message(
                aid,
                f"⚠️ <b>Обновление выполнено, но MC не отвечает</b>\n\n"
                f"Подождите 30 секунд или проверьте вручную.\n"
                f"<pre>{npm_out[:500]}</pre>",
                parse_mode="HTML",
            )
    except asyncio.TimeoutError:
        await bot.send_message(aid, "❌ npm update timed out (120s)", parse_mode="HTML")
    except Exception as e:
        await bot.send_message(aid, f"❌ Ошибка обновления: {e}", parse_mode="HTML")


async def mc_service_status() -> str:
    proc = await asyncio.create_subprocess_exec(
        "systemctl", "is-active", "meshcentral",
        stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
    )
    stdout, _ = await proc.communicate()
    return stdout.decode().strip()

def fmt_bytes(b) -> str:
    b = float(b)
    for u in ["B", "KB", "MB", "GB", "TB"]:
        if b < 1024:
            return f"{b:.1f} {u}"
        b /= 1024
    return f"{b:.1f} PB"

def fmt_uptime(s: float) -> str:
    d, s = divmod(int(s), 86400)
    h, s = divmod(s, 3600)
    m = s // 60
    return f"{d}д {h}ч {m}м" if d else (f"{h}ч {m}м" if h else f"{m}м")

def fmt_offline(hours: float) -> str:
    """Format offline duration from hours: '3ч 20м' or '2д 5ч'."""
    if hours <= 0:
        return "только что"
    total_min = int(hours * 60)
    d, rem = divmod(total_min, 1440)
    h, m = divmod(rem, 60)
    if d:
        return f"{d}д {h}ч" if h else f"{d}д"
    return f"{h}ч {m}м" if h else f"{m}м"

def pbar(pct: float, w=10) -> str:
    f = int(pct / 100 * w)
    return "█" * f + "░" * (w - f)


# ─── SSL certificate check ─────────────────────────────────────────

async def check_ssl_cert(hostname: str) -> dict:
    """Check SSL cert expiry for hostname. Returns dict with days_left, expires, ok, error."""
    loop = asyncio.get_event_loop()
    def _check():
        ctx = ssl.create_default_context()
        try:
            with ctx.wrap_socket(socket.create_connection((hostname, 443), timeout=8), server_hostname=hostname) as s:
                cert = s.getpeercert()
                not_after = cert.get("notAfter", "")
                exp = datetime.strptime(not_after, "%b %d %H:%M:%S %Y %Z").replace(tzinfo=timezone.utc)
                days = (exp - datetime.now(timezone.utc)).days
                return {"domain": hostname, "days_left": days, "expires": exp.strftime("%d.%m.%Y"), "ok": True, "error": ""}
        except ssl.SSLCertVerificationError as e:
            return {"domain": hostname, "days_left": -1, "expires": "—", "ok": False, "error": f"Ошибка верификации: {e}"}
        except Exception as e:
            return {"domain": hostname, "days_left": -1, "expires": "—", "ok": False, "error": str(e)}
    try:
        return await asyncio.wait_for(loop.run_in_executor(None, _check), timeout=12)
    except asyncio.TimeoutError:
        return {"domain": hostname, "days_left": -1, "expires": "—", "ok": False, "error": "Timeout"}


async def check_all_ssl() -> list[dict]:
    """Check all configured SSL domains concurrently."""
    results = await asyncio.gather(*[check_ssl_cert(d) for d in SSL_DOMAINS], return_exceptions=False)
    return list(results)


def ssl_status_text(results: list[dict]) -> str:
    lines = []
    for r in results:
        if not r["ok"]:
            icon = "❌"
            info = r["error"][:60]
        elif r["days_left"] <= SSL_CRIT_DAYS:
            icon = "🔴"
            info = f"истекает {r['expires']} (осталось {r['days_left']}д!)"
        elif r["days_left"] <= SSL_WARN_DAYS:
            icon = "🟡"
            info = f"истекает {r['expires']} ({r['days_left']}д)"
        else:
            icon = "🟢"
            info = f"до {r['expires']} ({r['days_left']}д)"
        lines.append(f"{icon} <code>{r['domain']}</code>: {info}")
    return "\n".join(lines)


# ─── Uptime tracking ────────────────────────────────────────────────

def record_uptime(devices: list[dict]):
    data = _load_json(UPTIME_FILE, {})
    now = datetime.now(timezone.utc).isoformat()
    for d in devices:
        name = d["name"]
        if name not in data:
            data[name] = []
        data[name].append({"t": now, "on": d["online"]})
        # keep last 7 days = ~13440 entries at 45s interval
        if len(data[name]) > 14000:
            data[name] = data[name][-13000:]
    _save_json(UPTIME_FILE, data)


def build_uptime_graph(device_name: str) -> bytes | None:
    data = _load_json(UPTIME_FILE, {})
    records = data.get(device_name, [])
    if len(records) < 2:
        return None

    times = []
    values = []
    for r in records[-2000:]:
        try:
            t = datetime.fromisoformat(r["t"])
            times.append(t)
            values.append(1 if r["on"] else 0)
        except Exception:
            continue

    if not times:
        return None

    fig, ax = plt.subplots(figsize=(10, 3))
    ax.fill_between(times, values, alpha=0.4, color="#2ecc71", step="post")
    ax.step(times, values, where="post", color="#27ae60", linewidth=1)
    ax.set_yticks([0, 1])
    ax.set_yticklabels(["Offline", "Online"])
    ax.set_title(f"Uptime: {device_name}", fontsize=12)
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%d.%m %H:%M"))
    ax.tick_params(axis="x", rotation=30)
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=100)
    plt.close(fig)
    buf.seek(0)
    return buf.read()


# ─── Snapshots / change tracking ────────────────────────────────────

def save_snapshot(devices: list[dict]):
    snaps = _load_json(SNAPSHOTS_FILE, {})
    for d in devices:
        snap = {
            "os": d["os"],
            "cpu": d["cpu"],
            "ram_total": d["ram_total"],
            "gpu": d["gpu"],
            "drives": d["drives"],
            "antivirus": d["antivirus"],
            "software_count": len(d.get("software", [])),
            "ip": d["ip"],
            "agent_ver": d["agent_ver"],
        }
        snaps[d["name"]] = snap
    _save_json(SNAPSHOTS_FILE, snaps)


def detect_changes(devices: list[dict]) -> list[str]:
    snaps = _load_json(SNAPSHOTS_FILE, {})
    changes = []
    for d in devices:
        name = d["name"]
        old = snaps.get(name)
        if not old:
            continue
        fields = {
            "os": "ОС", "cpu": "CPU", "ram_total": "RAM",
            "gpu": "GPU", "antivirus": "Антивирус",
            "ip": "IP", "agent_ver": "Агент",
        }
        for key, label in fields.items():
            old_val = old.get(key, "")
            new_val = d.get(key, "")
            if old_val and new_val and str(old_val) != str(new_val):
                changes.append(f"📱 <b>{name}</b>: {label} <code>{old_val}</code> → <code>{new_val}</code>")
        old_drives = old.get("drives", [])
        new_drives = d.get("drives", [])
        if old_drives and new_drives and set(map(str, old_drives)) != set(map(str, new_drives)):
            changes.append(f"📱 <b>{name}</b>: Диски изменились")
    return changes


# ─── Weekly digest ───────────────────────────────────────────────────

async def _send_weekly_digest(admin_id: int, devs: list[dict]):
    """Send weekly summary report (called every Sunday)."""
    try:
        from datetime import timedelta
        now = datetime.now(timezone.utc)
        week_start = (now - timedelta(days=7)).strftime("%d.%m")
        week_end   = now.strftime("%d.%m.%Y")

        total   = len(devs)
        online  = sum(1 for d in devs if d["online"])
        offline = total - online

        # Long-offline devices (>24h)
        cfg = load_alerts_cfg()
        offline_thresh = cfg.get("offline_hours", 24)
        long_offline = [d for d in devs if not d["online"] and d.get("offline_hours", 0) >= offline_thresh]
        long_offline.sort(key=lambda d: d.get("offline_hours", 0), reverse=True)

        # New devices (in known_devices but not in previous snapshot)
        snaps = _load_json(SNAPSHOTS_FILE, {})
        new_devices = [d for d in devs if d["name"] not in snaps]

        # Disk warnings
        disk_warn = []
        for d in devs:
            for alert in d.get("vol_alerts", []):
                disk_warn.append(f"  💿 {d['name']}: {alert}")

        # Security issues
        sec_issues = []
        for d in devs:
            if d.get("av_disabled"):
                sec_issues.append(f"  🛡 {d['name']}: AV выключен")

        # Disk trends: any critical
        trends = get_disk_trends()
        crit_trends = [t for t in trends if t["days_to_full"] is not None and t["days_to_full"] <= 30]

        # Average uptime ratio
        uptime_pct = (online / total * 100) if total > 0 else 0

        lines = [
            f"━━━━━━━━━━━━━━━━━━━━━━",
            f"📅 <b>Еженедельный дайджест</b>",
            f"{week_start} – {week_end}",
            f"━━━━━━━━━━━━━━━━━━━━━━",
            f"",
            f"📱 Устройств: {total}  🟢 {online} онлайн  ⚪ {offline} офлайн",
            f"📈 Доступность: {uptime_pct:.0f}%",
        ]

        if new_devices:
            lines += ["", f"✨ <b>Новые устройства ({len(new_devices)}):</b>"]
            for d in new_devices[:5]:
                lines.append(f"  • {d['name']} ({d['group']})")
            if len(new_devices) > 5:
                lines.append(f"  <i>... и ещё {len(new_devices)-5}</i>")

        if long_offline:
            lines += ["", f"⚪ <b>Долго офлайн ({len(long_offline)}):</b>"]
            for d in long_offline[:5]:
                lines.append(f"  • {d['name']}: {fmt_offline(d['offline_hours'])}")
            if len(long_offline) > 5:
                lines.append(f"  <i>... и ещё {len(long_offline)-5}</i>")

        if disk_warn:
            lines += ["", f"💿 <b>Диски (&gt;90%):</b>"]
            lines += disk_warn[:5]
            if len(disk_warn) > 5:
                lines.append(f"  <i>... и ещё {len(disk_warn)-5}</i>")

        if crit_trends:
            lines += ["", f"📈 <b>Диски заполнятся &lt;30 дней:</b>"]
            for t in crit_trends[:3]:
                lines.append(f"  • {t['device']} {t['letter']}: ~{int(t['days_to_full'])} д. ({t['used_pct']:.0f}%)")

        if sec_issues:
            lines += ["", f"🛡 <b>Безопасность:</b>"]
            lines += sec_issues[:5]

        if not (new_devices or long_offline or disk_warn or sec_issues or crit_trends):
            lines += ["", "✅ Неделя прошла без проблем!"]

        lines += ["", "━━━━━━━━━━━━━━━━━━━━━━"]

        await bot.send_message(admin_id, "\n".join(lines), parse_mode="HTML")
    except Exception as e:
        log.error(f"weekly_digest: {e}")


# ─── Disk fill-rate trend ────────────────────────────────────────────

def save_disk_snapshot(devices: list[dict]) -> None:
    """Append today's disk snapshot for each online device (one entry per day)."""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    hist = _load_json(DISK_HISTORY_FILE, {})
    for d in devices:
        if not d.get("online"):
            continue
        vols = d.get("volumes_raw", {})
        if not vols:
            continue
        name = d["name"]
        entries = hist.setdefault(name, [])
        # Replace today's entry or append
        existing = next((e for e in entries if e["date"] == today), None)
        snap = {"date": today, "volumes": vols}
        if existing:
            existing["volumes"] = vols
        else:
            entries.append(snap)
        # Keep last 60 days
        hist[name] = sorted(entries, key=lambda e: e["date"])[-60:]
    _save_json(DISK_HISTORY_FILE, hist)


def save_snap_history(devices: list[dict]) -> None:
    """Save daily snapshot of device info to SNAP_HISTORY_FILE. Keeps last 30 days."""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    hist = _load_json(SNAP_HISTORY_FILE, {})
    day_snap = {}
    for d in devices:
        day_snap[d["name"]] = {
            "os": d.get("os", ""), "cpu": d.get("cpu", ""),
            "ram_total": d.get("ram_total", ""), "gpu": d.get("gpu", ""),
            "ip": d.get("ip", ""), "agent_ver": d.get("agent_ver", ""),
            "online": d.get("online", False),
        }
    hist[today] = day_snap
    # Keep last 30 days
    keys = sorted(hist.keys())[-30:]
    _save_json(SNAP_HISTORY_FILE, {k: hist[k] for k in keys})


def compare_snap_history(devices: list[dict], days: int = 7) -> str:
    """Compare current device state with snapshot from N days ago."""
    hist = _load_json(SNAP_HISTORY_FILE, {})
    if not hist:
        return "⚠️ История снапшотов пуста. Данные накапливаются постепенно."
    now = datetime.now(timezone.utc)
    target_date = (now - timedelta(days=days)).strftime("%Y-%m-%d")
    # Find closest snapshot on or before target_date
    past_key = None
    for k in sorted(hist.keys()):
        if k <= target_date:
            past_key = k
    if not past_key:
        oldest = min(hist.keys())
        return f"⚠️ Нет снапшота за {days} дней назад. Самый ранний: {oldest}."
    past = hist[past_key]
    curr_map = {d["name"]: d for d in devices}
    lines = [f"📊 <b>Сравнение с {past_key}</b>", ""]
    fields = [("ОС", "os"), ("CPU", "cpu"), ("RAM", "ram_total"), ("IP", "ip"), ("Агент", "agent_ver")]
    appeared, disappeared, changed = [], [], []
    for name, old in past.items():
        if name not in curr_map:
            disappeared.append(name)
    for name, d in curr_map.items():
        if name not in past:
            appeared.append(name)
        else:
            old = past[name]
            diffs = []
            for label, key in fields:
                ov = str(old.get(key, "") or "")
                nv = str(d.get(key, "") or "")
                if ov and nv and ov != nv:
                    diffs.append(f"  {label}: <code>{ov[:30]}</code> → <code>{nv[:30]}</code>")
            if diffs:
                changed.append(f"📱 <b>{name}</b>\n" + "\n".join(diffs))
    if appeared:
        lines += [f"✨ <b>Новые ({len(appeared)}):</b>"] + [f"  + {n}" for n in appeared[:10]] + [""]
    if disappeared:
        lines += [f"🗑 <b>Исчезли ({len(disappeared)}):</b>"] + [f"  — {n}" for n in disappeared[:10]] + [""]
    if changed:
        lines += [f"🔄 <b>Изменились ({len(changed)}):</b>", ""] + changed[:20]
    if not appeared and not disappeared and not changed:
        lines.append("✅ Изменений не обнаружено.")
    return "\n".join(lines)


def get_disk_trends() -> list[dict]:
    """
    Returns list of dicts with fill-rate info per device per volume.
    Only includes volumes where trend is calculable (≥2 data points).
    Result sorted by days_to_full ascending (most critical first).
    """
    hist = _load_json(DISK_HISTORY_FILE, {})
    trends = []
    for device_name, entries in hist.items():
        if len(entries) < 2:
            continue
        # Group by volume letter
        vol_series: dict[str, list[tuple[float, float]]] = {}  # letter -> [(day_num, used_gb)]
        dates = [e["date"] for e in entries]
        # Use day offset from first entry
        from datetime import datetime as dt
        d0 = dt.strptime(dates[0], "%Y-%m-%d")
        for entry in entries:
            day_num = (dt.strptime(entry["date"], "%Y-%m-%d") - d0).days
            for letter, v in entry.get("volumes", {}).items():
                total = v.get("total", 0)
                free  = v.get("free", 0)
                if total <= 0:
                    continue
                used_gb = (total - free) / 1_073_741_824
                vol_series.setdefault(letter, []).append((day_num, used_gb))

        for letter, pts in vol_series.items():
            if len(pts) < 2:
                continue
            # Linear regression: used = a*day + b
            n = len(pts)
            sx  = sum(p[0] for p in pts)
            sy  = sum(p[1] for p in pts)
            sxx = sum(p[0] ** 2 for p in pts)
            sxy = sum(p[0] * p[1] for p in pts)
            denom = n * sxx - sx * sx
            if denom == 0:
                continue
            a = (n * sxy - sx * sy) / denom  # GB per day fill rate
            b = (sy - a * sx) / n

            # Current values from latest entry
            latest = sorted(pts, key=lambda p: p[0])[-1]
            day_latest = latest[0]
            used_latest = latest[1]

            # Get total_gb from last entry
            last_entry = entries[-1]
            total_bytes = last_entry.get("volumes", {}).get(letter, {}).get("total", 0)
            if total_bytes <= 0:
                continue
            total_gb = total_bytes / 1_073_741_824
            free_gb  = total_gb - used_latest
            used_pct = (used_latest / total_gb * 100) if total_gb > 0 else 0

            if a > 0.001:  # growing
                days_to_full = (total_gb - used_latest) / a if a > 0 else None
            else:
                days_to_full = None  # stable or shrinking

            trends.append({
                "device": device_name,
                "letter": letter,
                "used_gb": round(used_latest, 1),
                "total_gb": round(total_gb, 1),
                "free_gb": round(free_gb, 1),
                "used_pct": round(used_pct, 1),
                "fill_rate_gb_day": round(a, 3),
                "days_to_full": round(days_to_full, 0) if days_to_full is not None else None,
                "points": n,
            })

    # Sort: disks filling fastest first
    trends.sort(key=lambda t: (t["days_to_full"] is None, t["days_to_full"] or 99999))
    return trends


# ─── Wake-on-LAN ────────────────────────────────────────────────────

def send_wol(mac_address: str) -> bool:
    """Send a magic WoL packet to the given MAC address."""
    mac = mac_address.replace(":", "").replace("-", "").upper()
    if len(mac) != 12:
        return False
    try:
        mac_bytes = bytes.fromhex(mac)
        magic = b'\xff' * 6 + mac_bytes * 16
        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sock.setsockopt(socket.SOL_SOCKET, socket.SO_BROADCAST, 1)
        sock.sendto(magic, ("255.255.255.255", 9))
        sock.close()
        return True
    except Exception:
        return False


# ─── Agent Installer Generator ───────────────────────────────────────

async def _get_mesh_groups() -> dict[str, str]:
    """Return {group_name: mesh_id} mapping (raw MC IDs)."""
    raw = await _export_db_async()
    return {r.get("name", "?"): r["_id"].replace("mesh//", "") for r in raw if r.get("type") == "mesh"}


async def _download_configured_agent(mesh_id: str, agent_type: int = 4) -> bytes | None:
    """Download pre-configured agent binary from MC server (with MeshID+ServerID embedded).
    agent_type: 4=win64 exe, 6=win64 msi
    """
    import urllib.parse
    encoded = urllib.parse.quote(mesh_id, safe='')
    url = f"{MC_URL}/meshagents?id={agent_type}&meshid={encoded}&installflags=0"

    ssl_ctx = ssl.create_default_context()
    ssl_ctx.check_hostname = False
    ssl_ctx.verify_mode = ssl.CERT_NONE

    try:
        async with aiohttp.ClientSession() as sess:
            async with sess.get(url, timeout=aiohttp.ClientTimeout(total=30), ssl=ssl_ctx) as resp:
                if resp.status == 200:
                    data = await resp.read()
                    if len(data) > 100_000:  # agent should be >1MB
                        return data
    except Exception as e:
        log.error(f"Agent download failed: {e}")
    return None


def generate_local_installer_bat(group_name: str) -> str:
    """BAT installer for MeshAgent.exe — supports Windows 10/11 including 24H2/25H2.
    Workaround for KB5074105/KB5077181 breaking ShellExecuteW in the agent.
    """
    return f'''@echo off
chcp 65001 >nul 2>&1
title MeshCentral Agent — {group_name}
color 0A

:: ─── Save original directory before elevation ───
set "ORIG_DIR=%~dp0"

:: ─── Check administrator rights ───
net session >nul 2>&1
if %errorlevel% neq 0 (
    echo [*] Requesting administrator rights...
    powershell -Command "Start-Process cmd -ArgumentList '/c cd /d \"%ORIG_DIR%\" && \"%~f0\"' -Verb RunAs"
    exit /b
)

:: Ensure we are in the correct directory (where EXE is)
cd /d "%ORIG_DIR%"

echo.
echo  =============================================
echo   MeshCentral Agent Installer
echo   Group: {group_name}
echo   Windows 10 / 11 (including 24H2)
echo  =============================================
echo.

:: ─── Check that MeshAgent.exe is next to this BAT ───
if not exist "%~dp0MeshAgent.exe" (
    echo [ERROR] MeshAgent.exe not found!
    echo         Place this BAT in same folder as MeshAgent.exe
    pause
    exit /b 1
)

:: ─── Fix DLL search order (KB5074105/KB5077181 workaround) ───
:: Ensure system DLLs are found first, not from user folders
set "PATH=%SystemRoot%\\System32;%SystemRoot%;%PATH%"

:: ─── Unblock downloaded files (MOTW / Zone.Identifier) ───
echo [*] Configuring Windows security...
powershell -Command "Unblock-File -Path '%~dp0MeshAgent.exe' -ErrorAction SilentlyContinue; Remove-Item -Path '%~dp0MeshAgent.exe:Zone.Identifier' -ErrorAction SilentlyContinue; Unblock-File -Path '%~f0' -ErrorAction SilentlyContinue" >nul 2>&1

:: ─── Defender exclusions (paths + process) ───
echo [*] Adding Defender exclusions...
powershell -Command "$p='%ProgramFiles%\\Mesh Agent'; Add-MpPreference -ExclusionPath $p -ErrorAction SilentlyContinue; Add-MpPreference -ExclusionPath '%~dp0' -ErrorAction SilentlyContinue; Add-MpPreference -ExclusionProcess 'MeshAgent.exe' -ErrorAction SilentlyContinue" >nul 2>&1
timeout /t 2 /nobreak >nul

:: ─── Remove old agent if exists ───
sc query "Mesh Agent" >nul 2>&1
if %errorlevel% equ 0 (
    echo [*] Removing previous agent...
    net stop "Mesh Agent" >nul 2>&1
    timeout /t 3 /nobreak >nul
    "%ProgramFiles%\\Mesh Agent\\MeshAgent.exe" -uninstall >nul 2>&1
    timeout /t 3 /nobreak >nul
    sc delete "Mesh Agent" >nul 2>&1
    timeout /t 2 /nobreak >nul
    rd /s /q "%ProgramFiles%\\Mesh Agent" >nul 2>&1
)

:: ─── Copy agent to Program Files and install ───
echo [*] Installing agent...
if not exist "%ProgramFiles%\\Mesh Agent" mkdir "%ProgramFiles%\\Mesh Agent"
copy /y "%~dp0MeshAgent.exe" "%ProgramFiles%\\Mesh Agent\\MeshAgent.exe" >nul 2>&1

:: Unblock the copy in Program Files too
powershell -Command "Unblock-File -Path '%ProgramFiles%\\Mesh Agent\\MeshAgent.exe' -ErrorAction SilentlyContinue" >nul 2>&1

:: Run -fullinstall from Program Files (with system PATH)
:: This bypasses the ShellExecuteW bug in the agent
cd /d "%ProgramFiles%\\Mesh Agent"
echo [*] Running: MeshAgent.exe -fullinstall
"%ProgramFiles%\\Mesh Agent\\MeshAgent.exe" -fullinstall

echo [*] Waiting for service to start...
timeout /t 5 /nobreak >nul

:: ─── Firewall rules ───
echo [*] Configuring firewall...
netsh advfirewall firewall delete rule name="MeshCentral Agent" >nul 2>&1
netsh advfirewall firewall delete rule name="MeshCentral Agent In" >nul 2>&1
netsh advfirewall firewall add rule name="MeshCentral Agent" dir=out action=allow program="%ProgramFiles%\\Mesh Agent\\MeshAgent.exe" enable=yes >nul 2>&1
netsh advfirewall firewall add rule name="MeshCentral Agent In" dir=in action=allow program="%ProgramFiles%\\Mesh Agent\\MeshAgent.exe" enable=yes >nul 2>&1

:: ─── Verify service status ───
echo.
sc query "Mesh Agent" | findstr "RUNNING" >nul 2>&1
if %errorlevel% equ 0 (
    echo  =============================================
    echo   [OK] Agent installed and running!
    echo  =============================================
    echo.
    echo  Service: Mesh Agent
    echo  Location: %ProgramFiles%\\Mesh Agent
    echo  Group: {group_name}
) else (
    echo  [!] Service is not running yet.
    echo  [*] Retrying start...
    net start "Mesh Agent" >nul 2>&1
    timeout /t 3 /nobreak >nul
    sc query "Mesh Agent" | findstr "RUNNING" >nul 2>&1
    if %errorlevel% equ 0 (
        echo  [OK] Agent started on second attempt!
    ) else (
        echo  [!] Could not start. Try manually:
        echo      cd "%ProgramFiles%\\Mesh Agent"
        echo      MeshAgent.exe -fullinstall
    )
)

echo.
pause
'''


# ─── Remote commands via MeshCentral API ─────────────────────────────

async def _get_login_key() -> str:
    """Generate a fresh meshctrl login key."""
    proc = await asyncio.create_subprocess_exec(
        "node", f"{MC_DIR}/node_modules/meshcentral/meshcentral.js", "--logintokenkey",
        stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE, cwd=MC_DIR,
    )
    stdout, _ = await proc.communicate()
    return stdout.decode().strip()


async def mc_run_command(device_id: str, command: str, powershell: bool = False,
                         run_as_user: bool = False, timeout: int = 30) -> str:
    """Execute a command on a remote device via meshctrl RunCommand."""
    login_key = await _get_login_key()
    if not login_key:
        return "Error: failed to generate login key"

    args = [
        "node", MESHCTRL, "RunCommand",
        "--url", MC_WSS,
        "--loginkey", login_key,
        "--id", device_id,
        "--run", command,
        "--reply",
    ]
    if powershell:
        args.append("--powershell")
    if run_as_user:
        args.append("--runasuser")

    try:
        proc = await asyncio.create_subprocess_exec(
            *args,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
            cwd=MC_DIR,
        )
        stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=timeout)
        output = stdout.decode(errors="replace").strip()
        if not output and stderr:
            output = stderr.decode(errors="replace").strip()
        return output[:4000] if output else "(пустой ответ)"
    except asyncio.TimeoutError:
        return "Error: command timed out (30s)"
    except Exception as e:
        return f"Error: {e}"


async def mc_device_power(device_id: str, action: str) -> str:
    """Send power action (wake/sleep/reset/off) to a device."""
    login_key = await _get_login_key()
    if not login_key:
        return "Error: failed to generate login key"

    args = [
        "node", MESHCTRL, "DevicePower",
        "--url", MC_WSS,
        "--loginkey", login_key,
        "--id", device_id,
        f"--{action}",
    ]
    try:
        proc = await asyncio.create_subprocess_exec(
            *args, stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE, cwd=MC_DIR,
        )
        stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=15)
        return stdout.decode(errors="replace").strip() or "OK"
    except asyncio.TimeoutError:
        return "Error: timed out"
    except Exception as e:
        return f"Error: {e}"


# ─── Keenetic WiFi probe ───────────────────────────────────────────────

def _load_wifi_clients() -> dict:
    """Load wifi_clients.json into _wifi_clients."""
    global _wifi_clients
    try:
        if WIFI_FILE.exists():
            _wifi_clients = json.loads(WIFI_FILE.read_text())
    except Exception:
        _wifi_clients = {}
    return _wifi_clients


def _save_wifi_clients() -> None:
    try:
        WIFI_FILE.write_text(json.dumps(_wifi_clients, ensure_ascii=False, indent=2))
    except Exception as e:
        log.error(f"wifi save: {e}")


def _load_keenetic_probes() -> list[dict]:
    try:
        if KEENETIC_PROBES_FILE.exists():
            return json.loads(KEENETIC_PROBES_FILE.read_text())
    except Exception:
        pass
    return []


async def run_keenetic_probe(device_id: str, probe: dict) -> dict | None:
    """Run keenetic_probe.ps1 on the remote device; return parsed JSON or None."""
    try:
        script = KEENETIC_PROBE_SCRIPT.read_text(encoding="utf-8")
    except Exception as e:
        log.error(f"keenetic probe: cannot read script: {e}")
        return None

    script = script.replace("ROUTER_LOGIN", probe.get("router_login", "admin"))
    script = script.replace("ROUTER_PASSWORD", probe.get("router_password", ""))

    login_key = await _get_login_key()
    if not login_key:
        return None

    args = [
        "node", MESHCTRL, "RunCommand",
        "--url", MC_WSS,
        "--loginkey", login_key,
        "--id", device_id,
        "--run", script,
        "--reply",
        "--powershell",
    ]
    try:
        proc = await asyncio.create_subprocess_exec(
            *args,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
            cwd=MC_DIR,
        )
        stdout, _ = await asyncio.wait_for(proc.communicate(), timeout=60)
        raw = stdout.decode(errors="replace").strip()
        # meshctrl may emit log lines before JSON and after (PS warnings).
        # Find first '{', then use raw_decode to ignore trailing garbage.
        brace = raw.find("{")
        if brace == -1:
            return None
        obj, _ = json.JSONDecoder().raw_decode(raw, brace)
        return obj
    except asyncio.TimeoutError:
        log.warning(f"keenetic probe timeout for device {device_id}")
    except Exception as e:
        log.error(f"keenetic probe error: {e}")
    return None


async def wifi_poll_loop():
    """Background loop: poll all keenetic probes every WIFI_POLL_INTERVAL seconds."""
    global _wifi_clients
    _load_wifi_clients()
    await asyncio.sleep(10)  # short delay on startup
    while not _shutdown_event.is_set():
        try:
            probes = _load_keenetic_probes()
            if probes:
                devs = await get_full_devices()
                name_to_id = {d["name"]: d["id"] for d in devs}
                for probe in probes:
                    aname = probe.get("agent_name", "")
                    dev_id = name_to_id.get(aname)
                    if not dev_id:
                        log.info(f"wifi_poll: agent '{aname}' not found in devices")
                        continue
                    # only poll if device is online
                    dev = next((d for d in devs if d["id"] == dev_id), None)
                    if not dev or not dev.get("online"):
                        log.info(f"wifi_poll: agent '{aname}' is offline, skipping")
                        continue
                    log.info(f"wifi_poll: polling keenetic via {aname} ({dev_id})")
                    result = await run_keenetic_probe(dev_id, probe)
                    if result:
                        _wifi_clients[aname] = result
                        _save_wifi_clients()
                        log.info(f"wifi_poll: {aname} → {result.get('count', '?')} clients, ok={result.get('ok')}")
        except Exception as e:
            log.error(f"wifi_poll_loop: {e}")
        try:
            await asyncio.wait_for(_shutdown_event.wait(), timeout=WIFI_POLL_INTERVAL)
            break
        except asyncio.TimeoutError:
            pass


# ─── Status page builder ─────────────────────────────────────────────

def build_status_html(devices: list[dict]) -> str:
    """Generate a public status page grouped by location."""
    from collections import defaultdict
    now_str = datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M UTC")
    groups: dict[str, list[dict]] = defaultdict(list)
    for d in devices:
        loc = d.get("group", "Без группы") or "Без группы"
        groups[loc].append(d)

    rows = ""
    for loc in sorted(groups.keys()):
        devs = groups[loc]
        n_on  = sum(1 for d in devs if d.get("online"))
        n_off = len(devs) - n_on
        if n_on == len(devs):
            cls, icon, label = "ok",   "●", "Все онлайн"
        elif n_on == 0:
            cls, icon, label = "down", "●", "Все офлайн"
        else:
            cls, icon, label = "warn", "●", f"{n_on}/{len(devs)} онлайн"
        rows += (
            f'<div class="card">'
            f'<span class="dot {cls}">{icon}</span>'
            f'<div class="info"><div class="loc">{loc}</div>'
            f'<div class="sub">{label} &nbsp;·&nbsp; '
            f'<span class="on">{n_on} ↑</span> '
            f'<span class="off">{n_off} ↓</span></div></div>'
            f'</div>\n'
        )

    n_total = len(devices)
    n_online = sum(1 for d in devices if d.get("online"))
    overall = "ok" if n_online == n_total else ("down" if n_online == 0 else "warn")
    overall_label = (
        "Все системы работают" if n_online == n_total
        else f"Частичный сбой: {n_online}/{n_total} онлайн"
        if n_online > 0 else "Нет связи с устройствами"
    )

    return f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<meta http-equiv="refresh" content="60">
<title>Статус сети — MeshCentral</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
      background:#0a1120;color:#c8d6e5;min-height:100vh;padding:20px}}
.hdr{{text-align:center;padding:24px 0 8px}}
.hdr h1{{font-size:22px;color:#85c1e9;margin-bottom:6px}}
.hdr .ts{{font-size:12px;color:#4a7a99}}
.overall{{display:flex;align-items:center;justify-content:center;gap:10px;
          margin:18px auto;padding:12px 24px;border-radius:8px;max-width:500px;
          font-size:15px;font-weight:600}}
.overall.ok  {{background:#0f2a1a;border:1px solid #2ecc71;color:#2ecc71}}
.overall.warn{{background:#2a1f0a;border:1px solid #f39c12;color:#f39c12}}
.overall.down{{background:#2a0a0a;border:1px solid #e74c3c;color:#e74c3c}}
.cards{{max-width:600px;margin:0 auto;display:flex;flex-direction:column;gap:10px}}
.card{{display:flex;align-items:center;gap:14px;padding:14px 18px;
       border-radius:8px;background:#0f1e2e;border:1px solid #1e3a5f}}
.dot{{font-size:22px;flex-shrink:0}}
.dot.ok  {{color:#2ecc71}}
.dot.warn{{color:#f39c12}}
.dot.down{{color:#e74c3c}}
.loc{{font-size:14px;font-weight:600;color:#aed6f1}}
.sub{{font-size:12px;color:#6b8fa8;margin-top:3px}}
.on {{color:#2ecc71}}.off{{color:#e74c3c}}
.footer{{text-align:center;margin-top:30px;font-size:11px;color:#2a4a6a}}
.footer a{{color:#3498db;text-decoration:none}}
</style>
</head>
<body>
<div class="hdr">
  <h1>🗺 Статус сети</h1>
  <div class="ts">Обновлено: {now_str} &nbsp;·&nbsp; Обновляется каждые 60 сек</div>
</div>
<div class="overall {overall}">{overall_label}</div>
<div class="cards">
{rows}</div>
<div class="footer">
  MeshCentral &nbsp;·&nbsp; <a href="/netmap">Карта сети</a> &nbsp;·&nbsp; <a href="/rack">RackViz</a>
</div>
</body>
</html>"""


# ─── Availability heatmap (text) ─────────────────────────────────────

def build_availability_heatmap(device_name: str) -> str:
    """Build a 7-day per-hour text heatmap from uptime data."""
    data = _load_json(UPTIME_FILE, {})
    records = data.get(device_name, [])
    if not records:
        return f"Нет данных о доступности для «{device_name}»"

    # bucket records by (date, hour)
    from collections import defaultdict
    buckets: dict[tuple, list[int]] = defaultdict(list)
    for r in records:
        try:
            t = datetime.fromisoformat(r["t"])
            buckets[(t.date(), t.hour)].append(1 if r["on"] else 0)
        except Exception:
            continue

    now = datetime.now(timezone.utc)
    lines = ["<pre>"]
    lines.append(f"  Доступность: <b>{device_name}</b> (7 дней × 24ч)\n")
    lines.append("  Чч: " + " ".join(f"{h:02d}" for h in range(0, 24, 2)) + "\n")

    total_on = 0
    total_buckets = 0
    for day_offset in range(6, -1, -1):
        day = (now - timedelta(days=day_offset)).date()
        day_str = day.strftime("%d.%m")
        cells = []
        for hour in range(24):
            vals = buckets.get((day, hour), [])
            if not vals:
                cells.append("·")
            else:
                pct = sum(vals) / len(vals)
                total_on += sum(vals)
                total_buckets += len(vals)
                cells.append("█" if pct >= 0.8 else ("▒" if pct >= 0.4 else "░"))
        lines.append(f"  {day_str}: " + " ".join(cells[h] for h in range(0, 24, 2)) + "\n")

    uptime_pct = round(total_on / total_buckets * 100, 1) if total_buckets else 0
    lines.append(f"\n  █ онлайн  ▒ частично  ░ офлайн  · нет данных")
    lines.append(f"\n  Доступность за 7 дней: <b>{uptime_pct}%</b>")
    lines.append("</pre>")
    return "".join(lines)


# ─── HW Inventory helpers ────────────────────────────────────────────

def _hw_inventory_text(inv: dict, device_name: str) -> str:
    if not inv:
        return f"Нет данных HW для <b>{device_name}</b>"
    lines = [f"💻 <b>{device_name}</b> — аппаратный инвентарь\n"]
    lines.append(f"🏭 {inv.get('manufacturer','')} {inv.get('model','')}")
    if inv.get('serial'): lines.append(f"🔢 Серийный: <code>{inv['serial']}</code>")
    lines.append(f"\n🖥 ОС: {inv.get('os_name','')} {inv.get('os_arch','')}")
    if inv.get('os_install'): lines.append(f"   Установлена: {inv['os_install']}")
    if inv.get('last_boot'):  lines.append(f"   Последний старт: {inv['last_boot']}")
    lines.append(f"\n⚡ CPU: {inv.get('cpu_name','?')}")
    lines.append(f"   Ядра/Потоки: {inv.get('cpu_cores','?')}/{inv.get('cpu_threads','?')}"
                 f" @ {inv.get('cpu_mhz','?')} МГц")
    lines.append(f"\n🧠 RAM: {inv.get('ram_total_gb','?')} GB"
                 f"  ({inv.get('ram_slots','?')} модулей)")
    disks = inv.get('disks', [])
    if disks:
        lines.append("\n💾 Диски:")
        for d in disks:
            bar = "▓" * int(d.get('used_pct', 0) / 10) + "░" * (10 - int(d.get('used_pct', 0) / 10))
            lines.append(f"   {d['letter']} [{d.get('dtype','?')}] "
                         f"{d.get('size_gb','?')}GB — "
                         f"свободно {d.get('free_gb','?')}GB [{bar}] {d.get('used_pct','?')}%")
    if inv.get('gpu'): lines.append(f"\n🎮 GPU: {inv['gpu']}")
    if inv.get('updated'): lines.append(f"\n🕐 Обновлено: {inv['updated']}")
    return "\n".join(lines)


async def _collect_hw_for_device(device_id: str, device_name: str):
    """Run hw_inventory.ps1 on one device via meshctrl."""
    global _hw_inventory
    if not HW_INVENTORY_PS1.exists():
        return
    try:
        script = HW_INVENTORY_PS1.read_text(encoding="utf-8")
        raw = await mc_run_command(device_id, script, powershell=True, timeout=60)
        if not raw:
            return
        # extract JSON from output (clean BOM/control chars from PowerShell output)
        for line in reversed(raw.strip().splitlines()):
            line = line.strip().lstrip("\ufeff").replace("\r", "").replace("\x00", "")
            if line.startswith("{"):
                inv = json.loads(line)
                inv["updated"] = datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M")
                _hw_inventory[device_name] = inv
                _save_json(HW_INVENTORY_FILE, _hw_inventory)
                log.info(f"hw_inventory: collected {device_name}")
                break
    except Exception as e:
        log.warning(f"hw_inventory: {device_name}: {e}")


async def _collect_temp_for_device(device_id: str, device_name: str) -> dict | None:
    """Run temp_probe.ps1 on one device."""
    if not TEMP_PROBE_PS1.exists():
        return None
    try:
        script = TEMP_PROBE_PS1.read_text(encoding="utf-8")
        raw = await mc_run_command(device_id, script, powershell=True, timeout=30)
        if not raw:
            return None
        for line in reversed(raw.strip().splitlines()):
            line = line.strip().lstrip("\ufeff").replace("\r", "").replace("\x00", "")
            if line.startswith("{"):
                return json.loads(line)
    except Exception as e:
        log.warning(f"temp_probe: {device_name}: {e}")
    return None


# ─── HW Inventory loop ───────────────────────────────────────────────

async def hw_inventory_loop():
    """Collect hardware inventory from online agents every HW_POLL_INTERVAL seconds."""
    global _hw_inventory
    _hw_inventory = _load_json(HW_INVENTORY_FILE, {})
    await asyncio.sleep(120)  # delay on startup
    while not _shutdown_event.is_set():
        try:
            devs = await get_full_devices()
            online = [d for d in devs if d.get("online")]
            log.info(f"hw_inventory_loop: polling {len(online)} online devices")
            for d in online:
                if _shutdown_event.is_set():
                    break
                await _collect_hw_for_device(d["id"], d["name"])
                await asyncio.sleep(5)  # throttle
        except Exception as e:
            log.error(f"hw_inventory_loop: {e}")
        try:
            await asyncio.wait_for(_shutdown_event.wait(), timeout=HW_POLL_INTERVAL)
            break
        except asyncio.TimeoutError:
            pass


# ─── Temperature loop ────────────────────────────────────────────────

async def temp_loop():
    """Collect CPU temperature from online agents every TEMP_POLL_INTERVAL."""
    global _temp_data
    _temp_data = _load_json(TEMP_DATA_FILE, {})
    await asyncio.sleep(90)
    while not _shutdown_event.is_set():
        try:
            aid = get_admin_id()
            devs = await get_full_devices()
            online = [d for d in devs if d.get("online")]
            for d in online:
                if _shutdown_event.is_set():
                    break
                result = await _collect_temp_for_device(d["id"], d["name"])
                if not result:
                    continue
                result["updated"] = datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M")
                _temp_data[d["name"]] = result
                _save_json(TEMP_DATA_FILE, _temp_data)
                # Alert if any sensor is critical
                if aid:
                    for sensor in result.get("temps", []):
                        if sensor.get("temp_c", 0) >= TEMP_WARN_C:
                            await bot.send_message(
                                aid,
                                f"🌡 <b>Высокая температура!</b>\n"
                                f"💻 {d['name']}\n"
                                f"🌡 {sensor['zone']}: <b>{sensor['temp_c']}°C</b>\n"
                                f"⚠️ Порог: {TEMP_WARN_C}°C",
                                parse_mode="HTML",
                            )
                            break  # one alert per device per cycle
                await asyncio.sleep(3)
        except Exception as e:
            log.error(f"temp_loop: {e}")
        try:
            await asyncio.wait_for(_shutdown_event.wait(), timeout=TEMP_POLL_INTERVAL)
            break
        except asyncio.TimeoutError:
            pass


async def netmap_loop():
    """Background loop: regenerate netmap.html every NETMAP_INTERVAL seconds."""
    NETMAP_FILE.parent.mkdir(parents=True, exist_ok=True)
    STATUS_HTML_FILE.parent.mkdir(parents=True, exist_ok=True)
    await asyncio.sleep(5)  # short delay on startup
    while not _shutdown_event.is_set():
        try:
            devs = await get_full_devices()
            if devs:
                html = build_network_map_html(devs, web_mode=True)
                if html:
                    NETMAP_FILE.write_text(html, encoding="utf-8")
                # Status page
                status_html = build_status_html(devs)
                STATUS_HTML_FILE.write_text(status_html, encoding="utf-8")
                log.info(f"netmap: updated ({len(devs)} devices)")
        except Exception as e:
            log.error(f"netmap_loop: {e}")
        try:
            await asyncio.wait_for(_shutdown_event.wait(), timeout=NETMAP_INTERVAL)
            break
        except asyncio.TimeoutError:
            pass


# Quick command presets for devices
QUICK_COMMANDS = {
    "info": ("systeminfo | findstr /B /C:\"OS Name\" /C:\"OS Version\" /C:\"System Model\" /C:\"Total Physical\"", False),
    "ipconfig": ("ipconfig /all", False),
    "disk": ("wmic logicaldisk get caption,freespace,size", False),
    "uptime": ("powershell (Get-Date) - (gcim Win32_OperatingSystem).LastBootUpTime | Select Days,Hours,Minutes | fl", True),
    "users": ("query user", False),
    "procs": ("powershell Get-Process | Sort-Object CPU -Descending | Select-Object -First 15 Name,CPU,WorkingSet | Format-Table -AutoSize", True),
    "services": ("powershell Get-Service | Where-Object {$_.Status -eq 'Running'} | Select-Object -First 20 Name,DisplayName | Format-Table -AutoSize", True),
    "netstat": ("netstat -an | findstr LISTEN", False),
}


# ─── Pagination helpers ──────────────────────────────────────────────

def paginated_buttons(items: list[dict], page: int, prefix: str, name_key: str = "name",
                      icon_fn=None, extra_buttons: list = None) -> InlineKeyboardMarkup:
    total_pages = max(1, (len(items) + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    start = page * PAGE_SIZE
    end = start + PAGE_SIZE
    buttons = []
    for item in items[start:end]:
        icon = icon_fn(item) if icon_fn else ""
        label = f"{icon} {item[name_key]}" if icon else item[name_key]
        buttons.append([InlineKeyboardButton(text=label, callback_data=f"{prefix}:{item[name_key][:40]}")])

    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton(text="◀️", callback_data=f"page:{prefix}:{page - 1}"))
    nav.append(InlineKeyboardButton(text=f"{page + 1}/{total_pages}", callback_data="noop"))
    if page < total_pages - 1:
        nav.append(InlineKeyboardButton(text="▶️", callback_data=f"page:{prefix}:{page + 1}"))
    if nav:
        buttons.append(nav)

    if extra_buttons:
        buttons.extend(extra_buttons)

    return InlineKeyboardMarkup(inline_keyboard=buttons)


# ─── Handlers ────────────────────────────────────────────────────────

@router.message(Command("start"))
async def cmd_start(msg: Message):
    uid = msg.from_user.id
    uname = msg.from_user.username or msg.from_user.first_name or str(uid)
    locked = lock_admin(uid, uname)
    if not is_admin(uid):
        await msg.answer("🔒 Доступ запрещён.")
        return
    t = (
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "🖥  <b>MeshCentral Monitor v4</b>\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"🌐 {MC_URL}\n\n"
        "• Уведомления о подключениях\n"
        "• Полный инвентарь железа\n"
        "• Поиск устройств\n"
        "• Алерты (диск, AV, офлайн)\n"
        "• История изменений\n"
        "• Графики аптайма\n"
        "• Сравнение устройств\n"
        "• Wake-on-LAN\n"
        "• Инвентаризация ПО\n"
        "• 🆕 Удалённые команды\n"
        "• 🆕 PDF-отчёты\n"
        "• 🆕 Установщик агента (BAT/PS1)\n"
        "• Бэкап конфига MC\n"
        "• Мониторинг + автоперезапуск\n\n"
    )
    if locked:
        t += "✅ Вы — администратор.\n\n"
    t += "Клавиатура 👇"
    await msg.answer(t, reply_markup=MAIN_KB, parse_mode="HTML")


# ─── Status ──────────────────────────────────────────────────────────

@router.message(F.text == BTN_STATUS)
async def msg_status(msg: Message):
    if not is_admin(msg.from_user.id):
        return
    alive = await mc_is_alive()
    svc = await mc_service_status()
    devs = await get_full_devices()
    online = sum(1 for d in devs if d["online"])
    cpu = psutil.cpu_percent(interval=0.5)
    mem = psutil.virtual_memory()
    disk = shutil.disk_usage("/")
    up = time.time() - psutil.boot_time()
    # SSL quick summary from cache
    ssl_summary = ""
    if _ssl_cache:
        crit = [r for r in _ssl_cache if not r["ok"] or r["days_left"] <= SSL_CRIT_DAYS]
        warn = [r for r in _ssl_cache if r["ok"] and SSL_CRIT_DAYS < r["days_left"] <= SSL_WARN_DAYS]
        ok_n = len(_ssl_cache) - len(crit) - len(warn)
        if crit:
            ssl_summary = f"\n🔐 SSL: 🔴 {len(crit)} крит  🟡 {len(warn)} предупр  🟢 {ok_n} ок"
        elif warn:
            ssl_summary = f"\n🔐 SSL: 🟡 {len(warn)} предупр  🟢 {ok_n} ок"
        else:
            ssl_summary = f"\n🔐 SSL: 🟢 все {ok_n} ок"
    # HTTP services status from cache (updated by health_loop every 60s)
    http_lines = ""
    if HTTP_SERVICES:
        svc_parts = []
        for svc_name, svc_url in HTTP_SERVICES:
            is_down = _http_down.get(svc_name, False)
            svc_parts.append(f"{'🔴' if is_down else '🟢'} {svc_name}")
        http_lines = "\n" + "  ".join(svc_parts)

    t = (
        "━━━━━━━━━━━━━━━━━━━━━━\n🖥  <b>Status</b>\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"{'🟢' if alive else '🔴'} Web: <b>{'OK' if alive else 'DOWN'}</b>  •  ⚙️ <code>{svc}</code>\n"
        f"📱 Устройств: <b>{len(devs)}</b> (🟢 {online} online)\n"
        f"⏱ Uptime: {fmt_uptime(up)}\n\n"
        f"🧠 CPU: {pbar(cpu)} {cpu:.0f}%\n"
        f"💾 RAM: {pbar(mem.percent)} {mem.percent:.0f}%\n"
        f"💿 Disk: {pbar(disk.used / disk.total * 100)} {disk.used / disk.total * 100:.0f}%"
        f"{ssl_summary}"
        f"{http_lines}"
    )
    await msg.answer(t, parse_mode="HTML", reply_markup=MAIN_KB)
    if _ssl_cache:
        crit_items = [r for r in _ssl_cache if not r["ok"] or r["days_left"] <= SSL_WARN_DAYS]
        if crit_items:
            await msg.answer(
                "🔐 <b>SSL детали:</b>\n" + ssl_status_text(crit_items),
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="🔄 Проверить SSL", callback_data="tool:certs")],
                ]),
            )


# ─── Devices (paginated) ────────────────────────────────────────────

@router.message(F.text == BTN_DEVICES)
async def msg_devices(msg: Message):
    if not is_admin(msg.from_user.id):
        return
    devs = await get_full_devices()
    if not devs:
        await msg.answer("📭 Нет устройств.", reply_markup=MAIN_KB)
        return

    by_group: dict[str, list] = {}
    for d in devs:
        by_group.setdefault(d["group"], []).append(d)

    t = "━━━━━━━━━━━━━━━━━━━━━━\n📋  <b>Устройства</b>\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
    for group, gdevs in sorted(by_group.items()):
        t += f"📁 <b>{group}</b>\n"
        for d in sorted(gdevs, key=lambda x: x["name"]):
            icon = "🟢" if d["online"] else "⚪"
            t += f"  {icon} {d['name']} — <code>{d['ip']}</code>\n"
        t += "\n"

    sorted_devs = sorted(devs, key=lambda x: x["name"])
    extra = []
    groups = sorted(by_group.keys())
    if groups:
        grp_btns = []
        for g in groups:
            grp_btns.append(InlineKeyboardButton(text=f"📁 {g}", callback_data=f"grp:{g[:40]}"))
        extra.append(grp_btns)

    kb = paginated_buttons(
        sorted_devs, 0, "dev",
        icon_fn=lambda d: "🟢" if d["online"] else "⚪",
        extra_buttons=extra,
    )
    await msg.answer(t, parse_mode="HTML", reply_markup=kb)


@router.callback_query(F.data.startswith("page:"))
async def cb_page(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    parts = cb.data.split(":")
    prefix = parts[1]
    page = int(parts[2])

    if prefix == "dev":
        devs = await get_full_devices()
        sorted_devs = sorted(devs, key=lambda x: x["name"])
        by_group = {}
        for d in devs:
            by_group.setdefault(d["group"], [])
        extra = []
        groups = sorted(by_group.keys())
        if groups:
            grp_btns = [InlineKeyboardButton(text=f"📁 {g}", callback_data=f"grp:{g[:40]}") for g in groups]
            extra.append(grp_btns)
        kb = paginated_buttons(sorted_devs, page, "dev",
                               icon_fn=lambda d: "🟢" if d["online"] else "⚪",
                               extra_buttons=extra)
        try:
            await cb.message.edit_reply_markup(reply_markup=kb)
        except Exception:
            pass
    elif prefix == "sw":
        # software pagination - data in callback
        dev_name = parts[3] if len(parts) > 3 else ""
        devs = await get_full_devices()
        d = next((x for x in devs if x["name"] == dev_name), None)
        if d and d.get("software"):
            sw = d["software"]
            total_pages = max(1, (len(sw) + 20 - 1) // 20)
            page = max(0, min(page, total_pages - 1))
            start = page * 20
            end = start + 20
            lines = [f"📦 <b>ПО: {dev_name}</b> (стр {page + 1}/{total_pages})\n"]
            for s in sw[start:end]:
                lines.append(f"  • {s['name']} {s['version']}")
            nav = []
            if page > 0:
                nav.append(InlineKeyboardButton(text="◀️", callback_data=f"page:sw:{page - 1}:{dev_name[:30]}"))
            nav.append(InlineKeyboardButton(text=f"{page + 1}/{total_pages}", callback_data="noop"))
            if page < total_pages - 1:
                nav.append(InlineKeyboardButton(text="▶️", callback_data=f"page:sw:{page + 1}:{dev_name[:30]}"))
            try:
                await cb.message.edit_text("\n".join(lines), parse_mode="HTML",
                                           reply_markup=InlineKeyboardMarkup(inline_keyboard=[nav]))
            except Exception:
                pass
    elif prefix == "search":
        query = parts[3] if len(parts) > 3 else ""
        devs = await get_full_devices()
        q = query.lower()
        results = [d for d in devs if q in d["name"].lower() or q in d["ip"].lower()
                   or q in d["os"].lower() or q in d["cpu"].lower()
                   or q in d["board_sn"].lower() or q in d["os_sn"].lower()]
        results.sort(key=lambda x: x["name"])
        kb = paginated_buttons(results, page, "dev",
                               icon_fn=lambda d: "🟢" if d["online"] else "⚪")
        try:
            await cb.message.edit_reply_markup(reply_markup=kb)
        except Exception:
            pass

    await cb.answer()


@router.callback_query(F.data == "noop")
async def cb_noop(cb: CallbackQuery):
    await cb.answer()


# ─── Device detail ───────────────────────────────────────────────────

@router.callback_query(F.data.startswith("dev:"))
async def cb_device(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    name = cb.data.split(":", 1)[1]
    devs = await get_full_devices()
    d = next((x for x in devs if x["name"] == name), None)
    if not d:
        await cb.answer("Устройство не найдено", show_alert=True)
        return

    card = build_device_card(d)
    buttons_list = [
        [InlineKeyboardButton(text="📦 CSV", callback_data=f"csv1:{name[:40]}"),
         InlineKeyboardButton(text="📊 Аптайм", callback_data=f"upt:{name[:40]}")],
        [InlineKeyboardButton(text="💿 ПО", callback_data=f"soft:{name[:40]}"),
         InlineKeyboardButton(text="📡 WoL", callback_data=f"wol:{name[:40]}")],
        [InlineKeyboardButton(text="🖥 Команды", callback_data=f"rcmd:{name[:40]}"),
         InlineKeyboardButton(text="📄 PDF", callback_data=f"pdf1:{name[:40]}")],
    ]
    kb = InlineKeyboardMarkup(inline_keyboard=buttons_list)

    if len(card) > 4000:
        parts = [card[i:i + 4000] for i in range(0, len(card), 4000)]
        for i, part in enumerate(parts):
            if i == len(parts) - 1:
                await cb.message.answer(part, parse_mode="HTML", reply_markup=kb)
            else:
                await cb.message.answer(part, parse_mode="HTML")
    else:
        await cb.message.answer(card, parse_mode="HTML", reply_markup=kb)
    await cb.answer()


# ─── Group detail ────────────────────────────────────────────────────

@router.callback_query(F.data.startswith("grp:"))
async def cb_group(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    group = cb.data.split(":", 1)[1]
    devs = [d for d in await get_full_devices() if d["group"] == group]
    if not devs:
        await cb.answer("Группа пуста", show_alert=True)
        return

    # Group summary with WoL button for offline devices
    offline_with_mac = []
    for d in devs:
        if not d["online"]:
            for nic in d.get("nic_details", []):
                mac = nic.get("mac", "")
                if mac and mac != "00:00:00:00:00:00":
                    offline_with_mac.append(d["name"])
                    break

    online_c  = sum(1 for d in devs if d["online"])
    offline_c = len(devs) - online_c
    summary = (f"📁 <b>Группа: {group}</b>\n"
               f"🟢 {online_c} онлайн  🔴 {offline_c} офлайн  |  Всего: {len(devs)}")
    wol_kb = None
    if offline_with_mac:
        wol_kb = InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text=f"💤 Разбудить всех офлайн ({len(offline_with_mac)})",
                                 callback_data=f"wol_grp:{group[:50]}")
        ]])
    await cb.message.answer(summary, parse_mode="HTML", reply_markup=wol_kb)

    for d in sorted(devs, key=lambda x: x["name"]):
        card = build_device_card(d)
        if len(card) > 4000:
            card = card[:4000] + "\n..."
        await cb.message.answer(card, parse_mode="HTML")
        await asyncio.sleep(0.3)

    csv_data = build_inventory_csv(devs)
    await cb.message.answer_document(
        BufferedInputFile(csv_data, filename=f"inventory_{group}.csv"),
        caption=f"📦 <b>Инвентарь группы {group}</b> — {len(devs)} устройств",
        parse_mode="HTML",
    )
    await cb.answer()


# ─── Single device CSV ───────────────────────────────────────────────

@router.callback_query(F.data.startswith("csv1:"))
async def cb_csv_single(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    name = cb.data.split(":", 1)[1]
    devs = await get_full_devices()
    d = next((x for x in devs if x["name"] == name), None)
    if not d:
        await cb.answer("Не найдено", show_alert=True)
        return
    csv_data = build_inventory_csv([d])
    await cb.message.answer_document(
        BufferedInputFile(csv_data, filename=f"inventory_{name}.csv"),
        caption=f"📦 <b>Инвентарь: {name}</b>", parse_mode="HTML",
    )
    await cb.answer()


# ─── Single device PDF ────────────────────────────────────────────────

@router.callback_query(F.data.startswith("pdf1:"))
async def cb_pdf_single(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    name = cb.data.split(":", 1)[1]
    devs = await get_full_devices()
    d = next((x for x in devs if x["name"] == name), None)
    if not d:
        await cb.answer("Не найдено", show_alert=True)
        return
    pdf_data = build_single_device_pdf(d)
    await cb.message.answer_document(
        BufferedInputFile(pdf_data, filename=f"report_{name}.pdf"),
        caption=f"📄 <b>PDF-отчёт: {name}</b>", parse_mode="HTML",
    )
    await cb.answer()


# ─── Uptime graph ───────────────────────────────────────────────────

@router.callback_query(F.data.startswith("upt:"))
async def cb_uptime(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    name = cb.data.split(":", 1)[1]
    img = build_uptime_graph(name)
    if not img:
        await cb.answer("Недостаточно данных для графика", show_alert=True)
        return
    await cb.message.answer_photo(
        BufferedInputFile(img, filename=f"uptime_{name}.png"),
        caption=f"📊 <b>Аптайм: {name}</b>", parse_mode="HTML",
    )
    await cb.answer()


# ─── Software inventory ─────────────────────────────────────────────

@router.callback_query(F.data.startswith("soft:"))
async def cb_software(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    name = cb.data.split(":", 1)[1]
    devs = await get_full_devices()
    d = next((x for x in devs if x["name"] == name), None)
    if not d:
        await cb.answer("Не найдено", show_alert=True)
        return

    sw = d.get("software", [])
    if not sw:
        await cb.answer("Список ПО недоступен для этого устройства", show_alert=True)
        return

    sw.sort(key=lambda s: s["name"].lower())
    total_pages = max(1, (len(sw) + 20 - 1) // 20)
    lines = [f"📦 <b>ПО: {name}</b> ({len(sw)} программ, стр 1/{total_pages})\n"]
    for s in sw[:20]:
        lines.append(f"  • {s['name']} {s['version']}")

    nav = []
    nav.append(InlineKeyboardButton(text=f"1/{total_pages}", callback_data="noop"))
    if total_pages > 1:
        nav.append(InlineKeyboardButton(text="▶️", callback_data=f"page:sw:1:{name[:30]}"))

    await cb.message.answer("\n".join(lines), parse_mode="HTML",
                            reply_markup=InlineKeyboardMarkup(inline_keyboard=[nav]))
    await cb.answer()


# ─── Wake-on-LAN ────────────────────────────────────────────────────

@router.callback_query(F.data.startswith("wol:"))
async def cb_wol(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    name = cb.data.split(":", 1)[1]
    devs = await get_full_devices()
    d = next((x for x in devs if x["name"] == name), None)
    if not d:
        await cb.answer("Не найдено", show_alert=True)
        return

    macs = []
    for nic in d.get("nic_details", []):
        mac = nic.get("mac", "")
        if mac and mac != "00:00:00:00:00:00":
            macs.append(mac)

    if not macs:
        await cb.answer("MAC-адрес не найден", show_alert=True)
        return

    results = []
    for mac in macs[:2]:
        ok = send_wol(mac)
        results.append(f"{'✅' if ok else '❌'} {mac}")

    await cb.message.answer(
        f"📡 <b>Wake-on-LAN: {name}</b>\n\n" + "\n".join(results) +
        "\n\n<i>Magic-пакет отправлен. Устройство должно быть в одной сети.</i>",
        parse_mode="HTML",
    )
    await cb.answer()


# ─── Remote Commands ─────────────────────────────────────────────────

@router.callback_query(F.data.startswith("rcmd:"))
async def cb_remote_cmd_menu(cb: CallbackQuery):
    """Show quick-command menu for a device."""
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    name = cb.data.split(":", 1)[1]
    devs = await get_full_devices()
    d = next((x for x in devs if x["name"] == name), None)
    if not d:
        await cb.answer("Не найдено", show_alert=True)
        return
    if not d["online"]:
        await cb.answer("Устройство офлайн", show_alert=True)
        return

    buttons = []
    row = []
    for i, (key, (_, _)) in enumerate(QUICK_COMMANDS.items()):
        labels = {
            "info": "ℹ️ Sys Info", "ipconfig": "🌐 IP Config",
            "disk": "💿 Диски", "uptime": "⏱ Аптайм",
            "users": "👤 Юзеры", "procs": "📊 Процессы",
            "services": "⚙️ Службы", "netstat": "🔌 Порты",
        }
        label = labels.get(key, key)
        row.append(InlineKeyboardButton(text=label, callback_data=f"qcmd:{key}:{name[:30]}"))
        if len(row) == 2:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)

    buttons.append([InlineKeyboardButton(text="⌨️ Своя команда", callback_data=f"ccmd:{name[:30]}")])
    buttons.append([
        InlineKeyboardButton(text="🔄 Перезагрузка", callback_data=f"pwr:reset:{name[:30]}"),
        InlineKeyboardButton(text="⏻ Выключение", callback_data=f"pwr:off:{name[:30]}"),
    ])

    await cb.message.answer(
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🖥 <b>Команды: {name}</b>\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"Выберите быструю команду или введите свою:",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
    )
    await cb.answer()


@router.callback_query(F.data.startswith("qcmd:"))
async def cb_quick_cmd(cb: CallbackQuery):
    """Execute a quick command preset."""
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    parts = cb.data.split(":", 2)
    cmd_key = parts[1]
    dev_name = parts[2]

    if cmd_key not in QUICK_COMMANDS:
        await cb.answer("Неизвестная команда", show_alert=True)
        return

    command, is_ps = QUICK_COMMANDS[cmd_key]
    devs = await get_full_devices()
    d = next((x for x in devs if x["name"] == dev_name), None)
    if not d:
        await cb.answer("Не найдено", show_alert=True)
        return
    if not d["online"]:
        await cb.answer("Устройство офлайн", show_alert=True)
        return

    wait_msg = await cb.message.answer(f"⏳ Выполняю на <b>{dev_name}</b>...", parse_mode="HTML")
    result = await mc_run_command(d["id"], command, powershell=is_ps)

    # Clean up meshctrl output noise
    lines = result.split("\n")
    clean = []
    skip_prefixes = ("Microsoft Windows", "(c) ", "C:\\Program Files\\Mesh Agent>")
    for line in lines:
        stripped = line.strip()
        if stripped and not any(stripped.startswith(p) for p in skip_prefixes):
            if stripped != command and stripped != "exit":
                clean.append(line)
    output = "\n".join(clean).strip() or "(пустой ответ)"
    if len(output) > 3800:
        output = output[:3800] + "\n..."

    labels = {"info": "System Info", "ipconfig": "IP Config", "disk": "Диски",
              "uptime": "Аптайм", "users": "Пользователи", "procs": "Процессы",
              "services": "Службы", "netstat": "Порты"}

    await wait_msg.edit_text(
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🖥 <b>{dev_name}</b> → {labels.get(cmd_key, cmd_key)}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"<pre>{output}</pre>",
        parse_mode="HTML",
    )
    await cb.answer()


@router.callback_query(F.data.startswith("ccmd:"))
async def cb_custom_cmd_prompt(cb: CallbackQuery):
    """Prompt user to send a custom command."""
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    dev_name = cb.data.split(":", 1)[1]
    await cb.message.answer(
        f"⌨️ Отправьте команду для <b>{dev_name}</b>:\n\n"
        f"<code>/run {dev_name} hostname</code>\n"
        f"<code>/run {dev_name} -ps Get-Process | Select -First 10</code>\n\n"
        f"Флаги: <code>-ps</code> (PowerShell), <code>-u</code> (от имени пользователя)",
        parse_mode="HTML",
    )
    await cb.answer()


@router.callback_query(F.data.startswith("pwr:"))
async def cb_power_action(cb: CallbackQuery):
    """Send power action to a device."""
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    parts = cb.data.split(":", 2)
    action = parts[1]  # reset or off
    dev_name = parts[2]

    if action not in ("reset", "off"):
        await cb.answer("Неизвестное действие", show_alert=True)
        return

    devs = await get_full_devices()
    d = next((x for x in devs if x["name"] == dev_name), None)
    if not d:
        await cb.answer("Не найдено", show_alert=True)
        return

    # Confirm action
    action_label = "🔄 Перезагрузка" if action == "reset" else "⏻ Выключение"
    buttons = [[
        InlineKeyboardButton(text=f"✅ Да, {action_label.lower()}", callback_data=f"pwrdo:{action}:{dev_name[:30]}"),
        InlineKeyboardButton(text="❌ Отмена", callback_data="noop"),
    ]]
    await cb.message.answer(
        f"⚠️ <b>{action_label}: {dev_name}</b>\n\nВы уверены?",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
    )
    await cb.answer()


@router.callback_query(F.data.startswith("pwrdo:"))
async def cb_power_confirm(cb: CallbackQuery):
    """Execute confirmed power action."""
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    parts = cb.data.split(":", 2)
    action = parts[1]
    dev_name = parts[2]

    devs = await get_full_devices()
    d = next((x for x in devs if x["name"] == dev_name), None)
    if not d:
        await cb.answer("Не найдено", show_alert=True)
        return

    result = await mc_device_power(d["id"], action)
    action_label = "Перезагрузка" if action == "reset" else "Выключение"
    await cb.message.edit_text(
        f"⚡ <b>{action_label}: {dev_name}</b>\n\n{result}",
        parse_mode="HTML",
    )
    await cb.answer()


@router.message(Command("run"))
async def cmd_run(msg: Message):
    """Execute a command on a remote device: /run <device> [-ps] [-u] <command>"""
    if not is_admin(msg.from_user.id):
        return
    text = msg.text.split(maxsplit=1)
    if len(text) < 2:
        await msg.answer(
            "Использование:\n"
            "<code>/run ИмяПК hostname</code>\n"
            "<code>/run ИмяПК -ps Get-Process</code>\n"
            "<code>/run ИмяПК -u whoami</code>\n\n"
            "Флаги: <code>-ps</code> PowerShell, <code>-u</code> от пользователя",
            parse_mode="HTML", reply_markup=MAIN_KB,
        )
        return

    rest = text[1].strip()
    # Parse device name (first word) and flags
    tokens = rest.split()
    dev_name = tokens[0]
    powershell = False
    run_as_user = False
    cmd_start_idx = 1

    for i in range(1, len(tokens)):
        if tokens[i] == "-ps":
            powershell = True
            cmd_start_idx = i + 1
        elif tokens[i] == "-u":
            run_as_user = True
            cmd_start_idx = i + 1
        else:
            cmd_start_idx = i
            break

    command = " ".join(tokens[cmd_start_idx:])
    if not command:
        await msg.answer("❌ Не указана команда.", reply_markup=MAIN_KB)
        return

    devs = await get_full_devices()
    d = next((x for x in devs if x["name"].lower() == dev_name.lower()), None)
    if not d:
        await msg.answer(f"❌ Устройство «{dev_name}» не найдено.", reply_markup=MAIN_KB)
        return
    if not d["online"]:
        await msg.answer(f"⚪ <b>{d['name']}</b> офлайн.", parse_mode="HTML", reply_markup=MAIN_KB)
        return

    wait_msg = await msg.answer(f"⏳ Выполняю на <b>{d['name']}</b>:\n<code>{command}</code>", parse_mode="HTML")
    result = await mc_run_command(d["id"], command, powershell=powershell, run_as_user=run_as_user)

    # Clean output
    lines = result.split("\n")
    clean = []
    skip_prefixes = ("Microsoft Windows", "(c) ", "C:\\Program Files\\Mesh Agent>")
    for line in lines:
        stripped = line.strip()
        if stripped and not any(stripped.startswith(p) for p in skip_prefixes):
            if stripped != command and stripped != "exit":
                clean.append(line)
    output = "\n".join(clean).strip() or "(пустой ответ)"
    if len(output) > 3800:
        output = output[:3800] + "\n..."

    flags = []
    if powershell:
        flags.append("PowerShell")
    if run_as_user:
        flags.append("as user")
    flags_str = f" ({', '.join(flags)})" if flags else ""

    await wait_msg.edit_text(
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🖥 <b>{d['name']}</b>{flags_str}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"<code>$ {command}</code>\n\n"
        f"<pre>{output}</pre>",
        parse_mode="HTML", reply_markup=MAIN_KB,
    )


# ─── Agent Installer ──────────────────────────────────────────────────

@router.message(Command("install"))
async def cmd_install(msg: Message):
    """Show device group selection for installer generation."""
    if not is_admin(msg.from_user.id):
        return
    groups = await _get_mesh_groups()
    if not groups:
        await msg.answer("❌ Не удалось получить список групп.", reply_markup=MAIN_KB)
        return

    buttons = []
    for name in sorted(groups.keys()):
        buttons.append([InlineKeyboardButton(text=f"📁 {name}", callback_data=f"inst:{name[:40]}")])
    buttons.append([InlineKeyboardButton(text="📁 Все группы", callback_data="inst:__all__")])

    await msg.answer(
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "📥 <b>Установщик агента</b>\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "Выберите группу для установки агента.\n"
        "Бот скачает готовый <b>MeshAgent.exe</b>\n"
        "с сервера (MeshID/ServerID встроены).\n\n"
        "Поддержка: Windows 10 / 11 (включая 24H2/25H2)\n"
        "Обходит SmartScreen, Defender, Smart App Control.\n\n"
        "⚠️ <b>Win 11 25H2:</b> не запускайте EXE двойным кликом!\n"
        "Используйте install.bat или CMD от администратора:\n"
        "<code>MeshAgent.exe -fullinstall</code>",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
    )


@router.callback_query(F.data.startswith("inst:"))
async def cb_install_group(cb: CallbackQuery):
    """Download pre-configured agent EXE from MC and send directly via Telegram."""
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return

    group_name = cb.data.split(":", 1)[1]
    groups = await _get_mesh_groups()

    if group_name == "__all__":
        target_groups = list(groups.keys())
    else:
        if group_name not in groups:
            await cb.answer("Группа не найдена", show_alert=True)
            return
        target_groups = [group_name]

    wait_msg = await cb.message.answer("⏳ Скачиваю агент с сервера MeshCentral...")

    for gname in target_groups:
        mesh_id = groups[gname]
        safe_name = re.sub(r'[^\w\-]', '_', gname)

        # Download the actual 3.4MB agent binary with config embedded
        agent_data = await _download_configured_agent(mesh_id, agent_type=4)
        if not agent_data:
            await cb.message.answer(f"❌ Не удалось скачать агент для {gname}.")
            continue

        size_mb = len(agent_data) / 1024 / 1024

        # Send the EXE directly
        await cb.message.answer_document(
            BufferedInputFile(agent_data, filename=f"MeshAgent.exe"),
            caption=(
                f"📥 <b>Агент MeshCentral: {gname}</b> ({size_mb:.1f} MB)\n\n"
                f"⚠️ <b>НЕ запускайте EXE двойным кликом!</b>\n"
                f"Windows 11 25H2 (KB5074105/KB5077181)\n"
                f"ломает авто-элевацию агента.\n\n"
                f"<b>Установка через CMD (администратор):</b>\n"
                f"1. Скачайте файл в папку, например <code>C:\\Temp</code>\n"
                f"2. Откройте <b>CMD от имени администратора</b>\n"
                f"3. Выполните:\n"
                f"<code>cd C:\\Temp\n"
                f"MeshAgent.exe -fullinstall</code>\n\n"
                f"Или используйте <b>install.bat</b> (файл ниже) —\n"
                f"он сделает всё автоматически."
            ),
            parse_mode="HTML",
        )

        # Send install BAT (the recommended way)
        bat_content = generate_local_installer_bat(gname)
        await cb.message.answer_document(
            BufferedInputFile(bat_content.encode("utf-8"), filename=f"install.bat"),
            caption=(
                f"📋 <b>Установщик — install.bat</b>\n\n"
                f"1. Скачайте оба файла в одну папку\n"
                f"2. Запустите <code>install.bat</code>\n"
                f"3. Подтвердите UAC (права админа)\n\n"
                f"Обходит баг KB5074105, добавляет\n"
                f"исключения Defender, ставит службу.\n"
                f"Работает на Win 10/11 включая 24H2/25H2."
            ),
            parse_mode="HTML",
        )
        await asyncio.sleep(0.5)

    await wait_msg.delete()
    await cb.answer()


# ─── Search ──────────────────────────────────────────────────────────

@router.message(Command("search"))
async def cmd_search(msg: Message):
    if not is_admin(msg.from_user.id):
        return
    parts = msg.text.split(maxsplit=1)
    if len(parts) < 2:
        await msg.answer("Использование: /search <запрос>\nПоиск по имени, IP, ОС, CPU, серийникам", reply_markup=MAIN_KB)
        return
    query = parts[1].strip()
    q = query.lower()
    devs = await get_full_devices()
    results = [d for d in devs if q in d["name"].lower() or q in d["ip"].lower()
               or q in d["os"].lower() or q in d["cpu"].lower()
               or q in d["board_sn"].lower() or q in d["os_sn"].lower()
               or q in d["group"].lower()
               or any(q in nic.lower() for nic in d["nics"])]

    if not results:
        await msg.answer(f"🔍 По запросу «{query}» ничего не найдено.", reply_markup=MAIN_KB)
        return

    results.sort(key=lambda x: x["name"])
    t = f"🔍 <b>Результаты: «{query}»</b> — {len(results)} устройств\n\n"
    kb = paginated_buttons(results, 0, "dev",
                           icon_fn=lambda d: "🟢" if d["online"] else "⚪")
    await msg.answer(t, parse_mode="HTML", reply_markup=kb)


# ─── Compare devices ────────────────────────────────────────────────

@router.message(Command("compare"))
async def cmd_compare(msg: Message):
    if not is_admin(msg.from_user.id):
        return
    parts = msg.text.split(maxsplit=2)
    if len(parts) < 3:
        await msg.answer("Использование: /compare <устройство1> <устройство2>", reply_markup=MAIN_KB)
        return

    name1 = parts[1].strip()
    name2 = parts[2].strip()
    devs = await get_full_devices()
    d1 = next((d for d in devs if d["name"].lower() == name1.lower()), None)
    d2 = next((d for d in devs if d["name"].lower() == name2.lower()), None)

    if not d1:
        await msg.answer(f"❌ Устройство «{name1}» не найдено.", reply_markup=MAIN_KB)
        return
    if not d2:
        await msg.answer(f"❌ Устройство «{name2}» не найдено.", reply_markup=MAIN_KB)
        return

    fields = [
        ("ОС", "os"), ("Архитектура", "os_arch"), ("Build", "os_build"),
        ("CPU", "cpu"), ("RAM", "ram_total"), ("GPU", "gpu"),
        ("Плата", "board"), ("BIOS", "bios"), ("TPM", "tpm"),
        ("Антивирус", "antivirus"), ("Firewall", "firewall"),
        ("IP", "ip"), ("Разрешение", "resolution"),
        ("Загрузка", "last_boot"), ("Агент", "agent_ver"),
    ]

    lines = [
        f"━━━━━━━━━━━━━━━━━━━━━━",
        f"⚖️ <b>Сравнение</b>",
        f"━━━━━━━━━━━━━━━━━━━━━━",
        f"",
        f"{'':20s} │ <b>{d1['name'][:18]}</b> │ <b>{d2['name'][:18]}</b>",
        f"{'─' * 20}┼{'─' * 20}┼{'─' * 20}",
    ]
    for label, key in fields:
        v1 = str(d1.get(key, "-"))[:18]
        v2 = str(d2.get(key, "-"))[:18]
        diff = " ≠" if v1 != v2 else ""
        lines.append(f"{label:20s} │ {v1:20s} │ {v2}{diff}")

    await msg.answer("<pre>" + "\n".join(lines) + "</pre>", parse_mode="HTML", reply_markup=MAIN_KB)


# ─── Quick Scripts ──────────────────────────────────────────────────

@router.message(Command("save_script"))
async def cmd_save_script(msg: Message):
    """Save a quick script: /save_script <name> [-ps] <command>"""
    if not is_admin(msg.from_user.id):
        return
    parts = msg.text.split(maxsplit=2)
    if len(parts) < 3:
        await msg.answer(
            "Использование:\n"
            "<code>/save_script flush_dns ipconfig /flushdns</code>\n"
            "<code>/save_script check_disk -ps Get-PSDrive C</code>\n\n"
            "Флаг <code>-ps</code> = PowerShell",
            parse_mode="HTML", reply_markup=MAIN_KB,
        )
        return

    name = parts[1].strip().lower()
    rest = parts[2].strip()
    ps = False
    if rest.startswith("-ps "):
        ps = True
        rest = rest[4:].strip()

    if not rest:
        await msg.answer("❌ Не указана команда.", reply_markup=MAIN_KB)
        return

    scripts = load_scripts()
    scripts[name] = {"cmd": rest, "ps": ps, "desc": f"Custom: {rest[:40]}"}
    save_scripts(scripts)
    await msg.answer(
        f"✅ Скрипт <b>{name}</b> сохранён\n"
        f"{'⚡ PowerShell' if ps else '🖥 CMD'}: <code>{rest[:100]}</code>",
        parse_mode="HTML", reply_markup=MAIN_KB,
    )


@router.message(Command("scripts"))
async def cmd_scripts(msg: Message):
    """List all saved scripts."""
    if not is_admin(msg.from_user.id):
        return
    scripts = load_scripts()
    if not scripts:
        await msg.answer("📝 Нет сохранённых скриптов.", reply_markup=MAIN_KB)
        return
    text, buttons = _scripts_message(scripts)
    await msg.answer(text, parse_mode="HTML",
                     reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))


@router.message(Command("script"))
async def cmd_script(msg: Message):
    """Run a saved script on a device: /script <name> <device>"""
    if not is_admin(msg.from_user.id):
        return
    parts = msg.text.split(maxsplit=2)
    if len(parts) < 3:
        await msg.answer(
            "Использование: <code>/script имя_скрипта имя_устройства</code>",
            parse_mode="HTML", reply_markup=MAIN_KB,
        )
        return

    script_name = parts[1].strip().lower()
    dev_name = parts[2].strip()

    scripts = load_scripts()
    if script_name not in scripts:
        await msg.answer(f"❌ Скрипт «{script_name}» не найден.\n/scripts — список", reply_markup=MAIN_KB)
        return

    devs = await get_full_devices()
    d = next((x for x in devs if x["name"].lower() == dev_name.lower()), None)
    if not d:
        await msg.answer(f"❌ Устройство «{dev_name}» не найдено.", reply_markup=MAIN_KB)
        return
    if not d["online"]:
        await msg.answer(f"⚪ <b>{d['name']}</b> офлайн.", parse_mode="HTML", reply_markup=MAIN_KB)
        return

    s = scripts[script_name]
    wait_msg = await msg.answer(
        f"⏳ Запускаю <b>{script_name}</b> на <b>{d['name']}</b>...", parse_mode="HTML",
    )
    result = await mc_run_command(d["id"], s["cmd"], powershell=s.get("ps", False))

    lines = result.split("\n")
    clean = []
    skip_prefixes = ("Microsoft Windows", "(c) ", "C:\\Program Files\\Mesh Agent>")
    for line in lines:
        stripped = line.strip()
        if stripped and not any(stripped.startswith(p) for p in skip_prefixes):
            if stripped != s["cmd"] and stripped != "exit":
                clean.append(line)
    output = "\n".join(clean).strip() or "(пустой ответ)"
    if len(output) > 3800:
        output = output[:3800] + "\n..."

    await wait_msg.edit_text(
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📝 <b>{script_name}</b> → {d['name']}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"<pre>{output}</pre>",
        parse_mode="HTML",
    )


@router.callback_query(F.data.startswith("sdel:"))
async def cb_script_delete(cb: CallbackQuery):
    """Delete a saved script."""
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    name = cb.data.split(":", 1)[1]
    scripts = load_scripts()
    if name in scripts:
        del scripts[name]
        save_scripts(scripts)
        await cb.answer(f"Скрипт {name} удалён")
        await cb.message.delete()
    else:
        await cb.answer("Не найден", show_alert=True)


@router.callback_query(F.data.startswith("srun:"))
async def cb_script_run_select(cb: CallbackQuery):
    """Prompt to select device for script."""
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    name = cb.data.split(":", 1)[1]
    scripts = load_scripts()
    if name not in scripts:
        await cb.answer("Скрипт не найден", show_alert=True)
        return
    devs = await get_full_devices()
    online_devs = [d for d in devs if d["online"]]
    if not online_devs:
        await cb.answer("Нет онлайн устройств", show_alert=True)
        return
    buttons = []
    for d in sorted(online_devs, key=lambda x: x["name"]):
        buttons.append([InlineKeyboardButton(
            text=f"🟢 {d['name']}",
            callback_data=f"sexec:{name[:15]}:{d['name'][:25]}",
        )])
    await cb.message.answer(
        f"📝 Выберите устройство для <b>{name}</b>:",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
    )
    await cb.answer()


@router.callback_query(F.data.startswith("sexec:"))
async def cb_script_execute(cb: CallbackQuery):
    """Execute script on selected device."""
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    parts = cb.data.split(":", 2)
    script_name = parts[1]
    dev_name = parts[2]

    scripts = load_scripts()
    if script_name not in scripts:
        await cb.answer("Скрипт не найден", show_alert=True)
        return

    devs = await get_full_devices()
    d = next((x for x in devs if x["name"] == dev_name), None)
    if not d or not d["online"]:
        await cb.answer("Устройство недоступно", show_alert=True)
        return

    s = scripts[script_name]
    wait_msg = await cb.message.answer(
        f"⏳ Запускаю <b>{script_name}</b> на <b>{d['name']}</b>...", parse_mode="HTML",
    )
    result = await mc_run_command(d["id"], s["cmd"], powershell=s.get("ps", False))

    lines = result.split("\n")
    clean = []
    skip_prefixes = ("Microsoft Windows", "(c) ", "C:\\Program Files\\Mesh Agent>")
    for line in lines:
        stripped = line.strip()
        if stripped and not any(stripped.startswith(p) for p in skip_prefixes):
            if stripped != s["cmd"] and stripped != "exit":
                clean.append(line)
    output = "\n".join(clean).strip() or "(пустой ответ)"
    if len(output) > 3800:
        output = output[:3800] + "\n..."

    await wait_msg.edit_text(
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📝 <b>{script_name}</b> → {d['name']}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"<pre>{output}</pre>",
        parse_mode="HTML",
    )
    await cb.answer()


# ─── Maintenance Mode Commands ──────────────────────────────────────

@router.message(Command("mute"))
async def cmd_mute(msg: Message):
    """/mute <target> <duration> [reason] — mute alerts for device/group."""
    if not is_admin(msg.from_user.id):
        return
    parts = msg.text.split(maxsplit=3)
    if len(parts) < 3:
        await msg.answer(
            "Использование:\n"
            "<code>/mute ИмяПК 2h Обслуживание</code>\n"
            "<code>/mute ГруппаОфис 30m</code>\n"
            "<code>/mute __all__ 1d Плановые работы</code>\n\n"
            "Длительность: 30m, 2h, 1d",
            parse_mode="HTML", reply_markup=MAIN_KB,
        )
        return

    target = parts[1].strip()
    dur_str = parts[2].strip()
    reason = parts[3].strip() if len(parts) > 3 else ""

    seconds = parse_duration(dur_str)
    if seconds is None:
        await msg.answer("❌ Неверный формат длительности. Примеры: 30m, 2h, 1d", reply_markup=MAIN_KB)
        return

    mutes = load_mutes()
    until = time.time() + seconds
    mutes[target] = {
        "until": until,
        "reason": reason,
        "muted_at": datetime.now(timezone.utc).isoformat(),
    }
    save_mutes(mutes)

    until_str = datetime.fromtimestamp(until, tz=timezone.utc).strftime("%d.%m %H:%M UTC")
    await msg.answer(
        f"🔇 <b>Алерты отключены:</b> {target}\n"
        f"⏰ До: {until_str}\n"
        f"{'💬 ' + reason if reason else ''}",
        parse_mode="HTML", reply_markup=MAIN_KB,
    )


@router.message(Command("unmute"))
async def cmd_unmute(msg: Message):
    """/unmute <target> — remove mute."""
    if not is_admin(msg.from_user.id):
        return
    parts = msg.text.split(maxsplit=1)
    if len(parts) < 2:
        await msg.answer("Использование: <code>/unmute ИмяПК</code>", parse_mode="HTML", reply_markup=MAIN_KB)
        return

    target = parts[1].strip()
    mutes = load_mutes()
    if target in mutes:
        del mutes[target]
        save_mutes(mutes)
        await msg.answer(f"🔊 <b>{target}</b> — алерты включены.", parse_mode="HTML", reply_markup=MAIN_KB)
    else:
        await msg.answer(f"❌ {target} не найден в списке мутов.", reply_markup=MAIN_KB)


# ─── Keenetic WiFi clients command ──────────────────────────────────

@router.message(Command("wifi"))
async def cmd_wifi(msg: Message):
    """/wifi — show current WiFi clients from all Keenetic probes."""
    if not is_admin(msg.from_user.id):
        return
    _load_wifi_clients()
    probes = _load_keenetic_probes()
    if not probes:
        await msg.answer("⚙️ Нет настроенных зондов Keenetic.\nФайл: keenetic_probes.json", reply_markup=MAIN_KB)
        return
    if not _wifi_clients:
        await msg.answer(
            "📡 <b>WiFi клиенты</b>\n\nДанных ещё нет — зонд опрашивается каждые 5 минут.\n"
            "Убедитесь, что агент <b>RootkinG</b> онлайн.",
            parse_mode="HTML", reply_markup=MAIN_KB,
        )
        return

    lines = ["━━━━━━━━━━━━━━━━━━━━━━\n📡 <b>WiFi клиенты (Keenetic)</b>\n━━━━━━━━━━━━━━━━━━━━━━\n"]
    for agent_name, data in _wifi_clients.items():
        if not data.get("ok"):
            lines.append(f"❌ <b>{agent_name}</b>: {data.get('error', 'ошибка')}\n")
            continue
        updated = data.get("updated", "?")
        router  = data.get("router", "?")
        clients = data.get("clients") or []
        lines.append(f"🖥 <b>{agent_name}</b> → роутер {router}  <i>({updated})</i>")
        lines.append(f"Всего клиентов: <b>{len(clients)}</b>\n")
        for c in clients:
            ctype_val = c.get("type", "lan")
            if ctype_val == "wifi":    ctype = "📶"
            elif ctype_val == "printer": ctype = "🖨"
            else:                       ctype = "🔌"
            name   = c.get("name") or c.get("mac", "?")
            ip     = c.get("ip", "")
            rssi   = c.get("rssi")
            rssi_s = f" {rssi}dBm" if rssi else ""
            up     = c.get("online_sec")
            up_s   = f" ⏱{up//3600}h{(up%3600)//60}m" if up else ""
            lines.append(f"{ctype} <code>{name}</code>  {ip}{rssi_s}{up_s}")
        lines.append("")

    text = "\n".join(lines)
    # send in chunks if too long
    for i in range(0, len(text), 4000):
        await msg.answer(text[i:i+4000], parse_mode="HTML", reply_markup=MAIN_KB)


@router.message(Command("wifi_poll"))
async def cmd_wifi_poll(msg: Message):
    """/wifi_poll — force immediate Keenetic probe run."""
    if not is_admin(msg.from_user.id):
        return
    probes = _load_keenetic_probes()
    if not probes:
        await msg.answer("⚙️ keenetic_probes.json пуст.", reply_markup=MAIN_KB)
        return
    wait = await msg.answer("🔄 Запускаю зонд на агенте...", reply_markup=MAIN_KB)
    devs = await get_full_devices()
    name_to_id = {d["name"]: d["id"] for d in devs}
    results = []
    for probe in probes:
        aname  = probe.get("agent_name", "")
        dev_id = name_to_id.get(aname)
        if not dev_id:
            results.append(f"❌ Агент <b>{aname}</b> не найден")
            continue
        dev = next((d for d in devs if d["id"] == dev_id), None)
        if not dev or not dev.get("online"):
            results.append(f"⏸ Агент <b>{aname}</b> офлайн")
            continue
        result = await run_keenetic_probe(dev_id, probe)
        global _wifi_clients
        if result:
            _wifi_clients[aname] = result
            _save_wifi_clients()
            ok = result.get("ok", False)
            if ok:
                cnt = result.get("count", 0)
                results.append(f"✅ <b>{aname}</b>: {cnt} клиентов")
            else:
                results.append(f"❌ <b>{aname}</b>: {result.get('error', 'ошибка')}")
        else:
            results.append(f"❌ <b>{aname}</b>: нет ответа от зонда")
    await wait.delete()
    await msg.answer("\n".join(results) or "Нет результатов", parse_mode="HTML", reply_markup=MAIN_KB)


# ─── Remote Command FSM ─────────────────────────────────────────────

class RemoteCmdState(StatesGroup):
    entering_cmd  = State()   # device stored in state, waiting for command text
    entering_gcmd = State()   # group stored in state, waiting for command text


class MuteSetup(StatesGroup):
    picking_target   = State()   # picked device or group (buttons)
    picking_duration = State()   # duration buttons shown, or free text


# ─── WiFi FSM States ────────────────────────────────────────────────

class WifiOfficeSetup(StatesGroup):
    picking_group    = State()   # select MC group as office
    router_login     = State()
    router_password  = State()
    picking_agent    = State()
    editing_password = State()   # change password for existing probe


class PingState(StatesGroup):
    waiting_ip = State()   # user is typing a custom IP/hostname


class NotesState(StatesGroup):
    picking_device = State()   # user picks device from list
    writing_note   = State()   # user types note text


class SchedulerFSM(StatesGroup):
    picking_group   = State()  # step 1: pick group
    picking_devices = State()  # step 2: pick devices (stored in state["selected"])
    entering_cmd    = State()  # step 3: type command
    entering_time   = State()  # step 4: type time ("через 30" or "14:30")


# ─── WiFi UI helpers ────────────────────────────────────────────────

def _wifi_main_kb(probes: list[dict]) -> InlineKeyboardMarkup:
    """Keyboard for main WiFi view: one button per office + controls."""
    rows = []
    for p in probes:
        loc = p.get("location", p.get("agent_name", "?"))
        data = _wifi_clients.get(p.get("agent_name", ""))
        cnt  = data.get("count", 0) if data and data.get("ok") else "—"
        rows.append([InlineKeyboardButton(
            text=f"🏢 {loc}  ({cnt} устр.)",
            callback_data=f"wifi:office:{loc[:40]}",
        )])
    rows.append([
        InlineKeyboardButton(text="🔄 Обновить всё", callback_data="wifi:refresh_all"),
        InlineKeyboardButton(text="⚙️ Зонды",        callback_data="wifi:probes"),
    ])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _wifi_office_text(location: str, probe: dict) -> str:
    """Build formatted text for one office's clients."""
    aname   = probe.get("agent_name", "")
    data    = _wifi_clients.get(aname)
    if not data:
        return f"📡 <b>{location}</b>\n\nДанных нет — нажми 🔄 Обновить."
    if not data.get("ok"):
        return f"📡 <b>{location}</b>\n\n❌ Ошибка: {data.get('error', '?')}"

    clients  = data.get("clients") or []
    updated  = data.get("updated", "?")
    router   = data.get("router", "?")
    method   = data.get("method", "")
    wifi_c     = [c for c in clients if c.get("type") == "wifi"]
    printer_c  = [c for c in clients if c.get("type") == "printer"]
    lan_c      = [c for c in clients if c.get("type") not in ("wifi", "printer")]

    method_icon = {"keenetic-api": "🔑", "keenetic-api-basic": "🔑", "keenetic-api-noauth": "🔓",
                   "keenetic-api-md5": "🔑", "neighbor+nbtstat": "🔍"}.get(method, "")
    lines = [f"━━━━━━━━━━━━━━━━━━━━━━",
             f"📡 <b>{location}</b>  —  🌐 {router}",
             f"🕐 {updated}  ·  {method_icon} {method}",
             f"Всего устройств: <b>{len(clients)}</b>",
             "━━━━━━━━━━━━━━━━━━━━━━"]

    if wifi_c:
        lines.append(f"\n📶 <b>WiFi ({len(wifi_c)})</b>")
        for c in wifi_c:
            name = c.get("name") or c.get("ip", "?")
            ip   = c.get("ip", "")
            mac  = c.get("mac", "")
            rssi = c.get("rssi")
            link = c.get("link_mbps")
            rssi_s = f"  {rssi}dBm" if rssi is not None else ""
            link_s = f"  {link}Мбит/с" if link is not None else ""
            lines.append(f"  📶 <b>{name}</b>{rssi_s}{link_s}")
            if ip or mac:
                lines.append(f"     <code>{ip}</code>" + (f"  <code>{mac}</code>" if mac else ""))

    if printer_c:
        lines.append(f"\n🖨 <b>Принтеры ({len(printer_c)})</b>")
        for c in printer_c:
            name = c.get("name") or c.get("ip", "?")
            ip   = c.get("ip", "")
            mac  = c.get("mac", "")
            port = c.get("printer_port", "")
            port_s = f"  порт {port}" if port else ""
            lines.append(f"  🖨 <b>{name}</b>{port_s}")
            if ip or mac:
                lines.append(f"     <code>{ip}</code>" + (f"  <code>{mac}</code>" if mac else ""))

    if lan_c:
        lines.append(f"\n🔌 <b>LAN ({len(lan_c)})</b>")
        for c in lan_c:
            name = c.get("name") or c.get("ip", "?")
            ip   = c.get("ip", "")
            mac  = c.get("mac", "")
            link = c.get("link_mbps")
            link_s = f"  {link}Мбит/с" if link is not None else ""
            lines.append(f"  🔌 <b>{name}</b>{link_s}")
            if ip or mac:
                lines.append(f"     <code>{ip}</code>" + (f"  <code>{mac}</code>" if mac else ""))

    return "\n".join(lines)


async def _run_probe_for(probe: dict, devs: list[dict]) -> str:
    """Run probe and update cache. Returns status string."""
    global _wifi_clients
    aname  = probe.get("agent_name", "")
    loc    = probe.get("location", aname)
    dev    = next((d for d in devs if d["name"] == aname), None)
    if not dev:
        return f"❌ <b>{loc}</b>: агент не найден"
    if not dev.get("online"):
        return f"⏸ <b>{loc}</b>: агент офлайн"
    result = await run_keenetic_probe(dev["id"], probe)
    if result:
        _wifi_clients[aname] = result
        _save_wifi_clients()
        return f"✅ <b>{loc}</b>: {result.get('count', 0)} устр."
    return f"❌ <b>{loc}</b>: нет ответа"


# ─── WiFi button handler ─────────────────────────────────────────────

@router.message(F.text == BTN_WIFI)
async def btn_wifi_main(msg: Message):
    if not is_admin(msg.from_user.id):
        return
    _load_wifi_clients()
    probes = _load_keenetic_probes()
    if not probes:
        await msg.answer(
            "📡 <b>WiFi сети</b>\n\nЗонды не настроены.\n"
            "Нажми ⚙️ Зонды чтобы назначить агента для офиса.",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="⚙️ Зонды", callback_data="wifi:probes"),
            ]]),
        )
        return
    await msg.answer(
        "📡 <b>WiFi сети по офисам</b>\n\nВыбери офис:",
        parse_mode="HTML",
        reply_markup=_wifi_main_kb(probes),
    )


# ─── WiFi callback: main view ────────────────────────────────────────

@router.callback_query(F.data == "wifi:main")
async def cb_wifi_main(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        return
    _load_wifi_clients()
    probes = _load_keenetic_probes()
    await cb.message.edit_text(
        "📡 <b>WiFi сети по офисам</b>\n\nВыбери офис:",
        parse_mode="HTML",
        reply_markup=_wifi_main_kb(probes),
    )
    await cb.answer()


# ─── WiFi callback: office view ──────────────────────────────────────

@router.callback_query(F.data.startswith("wifi:office:"))
async def cb_wifi_office(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        return
    location = cb.data[len("wifi:office:"):]
    probes   = _load_keenetic_probes()
    probe    = next((p for p in probes if p.get("location", p.get("agent_name")) == location), None)
    if not probe:
        await cb.answer("Офис не найден", show_alert=True)
        return
    text = _wifi_office_text(location, probe)
    kb   = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="🔄 Обновить", callback_data=f"wifi:refresh_one:{location[:40]}"),
        InlineKeyboardButton(text="◀️ Офисы",   callback_data="wifi:main"),
    ]])
    try:
        await cb.message.edit_text(text[:4096], parse_mode="HTML", reply_markup=kb)
    except Exception:
        await cb.message.answer(text[:4096], parse_mode="HTML", reply_markup=kb)
    await cb.answer()


# ─── WiFi callback: refresh all ──────────────────────────────────────

@router.callback_query(F.data == "wifi:refresh_all")
async def cb_wifi_refresh_all(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        return
    await cb.answer("🔄 Обновляю все офисы...", show_alert=False)
    probes = _load_keenetic_probes()
    devs   = await get_full_devices()
    for probe in probes:
        await _run_probe_for(probe, devs)
    probes = _load_keenetic_probes()
    await cb.message.edit_text(
        "📡 <b>WiFi сети по офисам</b>\n\nВыбери офис:",
        parse_mode="HTML",
        reply_markup=_wifi_main_kb(probes),
    )


# ─── WiFi callback: refresh one office ───────────────────────────────

@router.callback_query(F.data.startswith("wifi:refresh_one:"))
async def cb_wifi_refresh_one(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        return
    location = cb.data[len("wifi:refresh_one:"):]
    probes   = _load_keenetic_probes()
    probe    = next((p for p in probes if p.get("location", p.get("agent_name")) == location), None)
    if not probe:
        await cb.answer("Офис не найден", show_alert=True)
        return
    await cb.answer("🔄 Обновляю...")
    devs = await get_full_devices()
    await _run_probe_for(probe, devs)
    text = _wifi_office_text(location, probe)
    kb   = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="🔄 Обновить", callback_data=f"wifi:refresh_one:{location[:40]}"),
        InlineKeyboardButton(text="◀️ Офисы",   callback_data="wifi:main"),
    ]])
    try:
        await cb.message.edit_text(text[:4096], parse_mode="HTML", reply_markup=kb)
    except Exception:
        pass


# ─── WiFi callback: probe management ─────────────────────────────────

@router.callback_query(F.data == "wifi:probes")
async def cb_wifi_probes(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        return
    probes = _load_keenetic_probes()
    lines  = ["⚙️ <b>Зонды по офисам</b>\n"]
    if probes:
        for p in probes:
            loc   = p.get("location", p.get("agent_name", "?"))
            agent = p.get("agent_name", "?")
            data  = _wifi_clients.get(agent)
            status = f"✅ {data.get('count',0)} устр." if data and data.get("ok") else "⏳ нет данных"
            lines.append(f"🏢 <b>{loc}</b> — агент: {agent}  {status}")
    else:
        lines.append("Нет назначенных зондов")

    rows = []
    for idx, p in enumerate(probes):
        loc = p.get("location", p.get("agent_name", "?"))[:30]
        rows.append([
            InlineKeyboardButton(text=f"✏️ {loc}", callback_data=f"wifi:edit:{idx}"),
            InlineKeyboardButton(text="❌",          callback_data=f"wifi:del:{loc[:40]}"),
        ])
    rows.append([InlineKeyboardButton(text="➕ Новый офис", callback_data="wifi:new_office")])
    rows.append([InlineKeyboardButton(text="◀️ Офисы", callback_data="wifi:main")])

    await cb.message.edit_text(
        "\n".join(lines), parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=rows),
    )
    await cb.answer()


# ─── WiFi FSM: new office setup ──────────────────────────────────────

_CANCEL_KB = InlineKeyboardMarkup(inline_keyboard=[[
    InlineKeyboardButton(text="❌ Отмена", callback_data="wifi:cancel_setup"),
]])


@router.callback_query(F.data == "wifi:new_office")
async def cb_wifi_new_office(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        return
    await cb.answer()
    wait = await cb.message.answer("⏳ Загружаю группы...")
    devs = await _list_agents_quick()
    groups = sorted(set(d.get("group", "") for d in devs if d.get("group")))
    await wait.delete()
    if not groups:
        await cb.message.answer(
            "⚠️ Не удалось загрузить группы MeshCentral.\nПопробуй позже.",
            reply_markup=_CANCEL_KB,
        )
        return
    await state.set_state(WifiOfficeSetup.picking_group)
    rows = [[InlineKeyboardButton(
        text=f"🏢 {g}",
        callback_data=f"wifi:group:{g[:40]}",
    )] for g in groups]
    rows.append([InlineKeyboardButton(text="❌ Отмена", callback_data="wifi:cancel_setup")])
    await cb.message.answer(
        "🏢 <b>Новый офис — шаг 1/4</b>\n\n"
        "Выбери группу MeshCentral (= офис):",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=rows),
    )


@router.callback_query(F.data.startswith("wifi:group:"), WifiOfficeSetup.picking_group)
async def cb_wifi_pick_group(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        return
    group = cb.data[len("wifi:group:"):]
    await state.update_data(office_name=group, selected_group=group)
    await state.set_state(WifiOfficeSetup.router_login)
    await cb.message.answer(
        f"✅ Офис: <b>{group}</b>\n\n"
        "🔑 <b>Шаг 2/4</b> — Логин роутера\n"
        "Введи логин или нажми кнопку:",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="admin (по умолчанию)", callback_data="wifi:fsm_login:admin")],
            [InlineKeyboardButton(text="❌ Отмена", callback_data="wifi:cancel_setup")],
        ]),
    )
    await cb.answer()


@router.callback_query(F.data.startswith("wifi:fsm_login:"))
async def cb_wifi_fsm_login(cb: CallbackQuery, state: FSMContext):
    login = cb.data[len("wifi:fsm_login:"):]
    await state.update_data(router_login=login)
    await state.set_state(WifiOfficeSetup.router_password)
    data = await state.get_data()
    await cb.message.answer(
        f"✅ Офис: <b>{data['office_name']}</b>  Логин: <b>{login}</b>\n\n"
        "🔒 <b>Шаг 3/4</b> — Пароль роутера\n"
        "Введи пароль (или <code>-</code> если без пароля):",
        parse_mode="HTML", reply_markup=_CANCEL_KB,
    )
    await cb.answer()


@router.message(WifiOfficeSetup.router_login)
async def wifi_fsm_router_login(msg: Message, state: FSMContext):
    if not is_admin(msg.from_user.id):
        return
    login = "" if msg.text.strip() == "-" else msg.text.strip()
    await state.update_data(router_login=login)
    await state.set_state(WifiOfficeSetup.router_password)
    data = await state.get_data()
    await msg.answer(
        f"✅ Офис: <b>{data['office_name']}</b>  Логин: <b>{login or '—'}</b>\n\n"
        "🔒 <b>Шаг 3/4</b> — Пароль роутера\n"
        "Введи пароль (или <code>-</code> если без пароля):",
        parse_mode="HTML", reply_markup=_CANCEL_KB,
    )


@router.message(WifiOfficeSetup.router_password)
async def wifi_fsm_router_password(msg: Message, state: FSMContext):
    if not is_admin(msg.from_user.id):
        return
    pwd = "" if msg.text.strip() == "-" else msg.text.strip()
    await state.update_data(router_password=pwd)
    await state.set_state(WifiOfficeSetup.picking_agent)
    # Delete password message for security
    try:
        await msg.delete()
    except Exception:
        pass
    data  = await state.get_data()
    selected_group = data.get("selected_group", "")
    wait = await msg.answer("⏳ Загружаю список агентов...")
    try:
        devs = await _list_agents_quick()
    except Exception:
        devs = []
    try:
        await wait.delete()
    except Exception:
        pass
    current = {p["agent_name"] for p in _load_keenetic_probes()}
    # Filter by selected group if set
    if selected_group:
        group_devs = [d for d in devs if d.get("group", "") == selected_group]
        if not group_devs:  # fallback if group is empty in MC
            group_devs = devs
    else:
        group_devs = devs
    group_devs = sorted(group_devs, key=lambda x: (not x.get("online"), x.get("name", "")))
    # Store agent list in state (to look up by index — avoids 64-byte callback_data limit)
    await state.update_data(agent_list=[d["name"] for d in group_devs])
    rows = []
    for idx, d in enumerate(group_devs):
        agent  = d["name"]
        online = "🟢" if d.get("online") else "🔴"
        mark   = " ✓" if agent in current else ""
        # Use index as callback_data — safe, always short
        rows.append([InlineKeyboardButton(
            text=f"{online} {agent}{mark}",
            callback_data=f"wifi:finalize:{idx}",
        )])
    rows.append([InlineKeyboardButton(text="❌ Отмена", callback_data="wifi:cancel_setup")])
    group_label = f" <b>{selected_group}</b>" if selected_group else f" <b>{data.get('office_name', '?')}</b>"
    if not group_devs:
        text = (f"⚠️ <b>Шаг 4/4</b> — Нет агентов в группе{group_label}\n\n"
                "MeshCentral не вернул устройства. Попробуй позже.")
    else:
        text = (f"🖥 <b>Шаг 4/4</b> — Выбери агента для офиса{group_label}\n\n"
                "Это ПК который всегда включён в этом офисе.\n"
                "🟢 онлайн  🔴 офлайн")
    await msg.answer(text, parse_mode="HTML",
                     reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))


@router.callback_query(F.data.startswith("wifi:finalize:"))
async def cb_wifi_finalize(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        return
    idx_str = cb.data[len("wifi:finalize:"):]
    data = await state.get_data()
    # Resolve agent name from stored list (index-based)
    agent_list = data.get("agent_list", [])
    try:
        agent_name = agent_list[int(idx_str)]
    except (ValueError, IndexError):
        await cb.answer("Ошибка: список агентов устарел. Начни заново.", show_alert=True)
        await state.clear()
        return
    office_name     = data.get("office_name", agent_name)
    router_login    = data.get("router_login", "admin")
    router_password = data.get("router_password", "")
    await state.clear()

    probes = [p for p in _load_keenetic_probes() if p.get("agent_name") != agent_name]
    probes.append({
        "agent_name":     agent_name,
        "location":       office_name,
        "router_login":   router_login,
        "router_password": router_password,
    })
    KEENETIC_PROBES_FILE.write_text(json.dumps(probes, ensure_ascii=False, indent=2))

    await cb.answer(f"✅ Сохранено!")
    await cb.message.edit_text(
        f"✅ <b>Офис настроен</b>\n\n"
        f"🏢 Название: <b>{office_name}</b>\n"
        f"🖥 Агент: <b>{agent_name}</b>\n"
        f"🔑 Логин: <code>{router_login or '—'}</code>\n\n"
        "Данные соберутся в течение 5 минут автоматически,\n"
        "или нажми 🔄 Обновить в меню офисов.",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="◀️ К офисам", callback_data="wifi:main"),
        ]]),
    )


@router.callback_query(F.data == "wifi:cancel_setup")
async def cb_wifi_cancel_setup(cb: CallbackQuery, state: FSMContext):
    await state.clear()
    await cb.answer("Отменено")
    try:
        await cb.message.edit_text("❌ Настройка отменена.",
                                   reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                                       InlineKeyboardButton(text="◀️ Зонды", callback_data="wifi:probes"),
                                   ]]))
    except Exception:
        pass


# ─── WiFi callback: edit probe ───────────────────────────────────────

@router.callback_query(F.data.startswith("wifi:edit:"))
async def cb_wifi_edit(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        return
    try:
        idx = int(cb.data[len("wifi:edit:"):])
    except ValueError:
        await cb.answer("Ошибка", show_alert=True)
        return
    probes = _load_keenetic_probes()
    if idx >= len(probes):
        await cb.answer("Офис не найден", show_alert=True)
        return
    p   = probes[idx]
    loc = p.get("location", p.get("agent_name", "?"))
    agt = p.get("agent_name", "?")
    login = p.get("router_login", "admin")
    has_pwd = bool(p.get("router_password"))
    await cb.message.answer(
        f"✏️ <b>Редактирование: {loc}</b>\n\n"
        f"🖥 Агент: <b>{agt}</b>\n"
        f"🔑 Логин: <code>{login or '—'}</code>\n"
        f"🔒 Пароль: {'***' if has_pwd else '—'}\n\n"
        "Что изменить?",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔒 Изменить пароль",
                                  callback_data=f"wifi:chpwd:{idx}")],
            [InlineKeyboardButton(text="🔑 Изменить логин",
                                  callback_data=f"wifi:chlogin:{idx}")],
            [InlineKeyboardButton(text="◀️ Зонды", callback_data="wifi:probes")],
        ]),
    )
    await cb.answer()


@router.callback_query(F.data.startswith("wifi:chpwd:"))
async def cb_wifi_chpwd(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        return
    idx_str = cb.data[len("wifi:chpwd:"):]
    probes  = _load_keenetic_probes()
    try:
        idx = int(idx_str)
        p   = probes[idx]
    except (ValueError, IndexError):
        await cb.answer("Офис не найден", show_alert=True)
        return
    loc = p.get("location", p.get("agent_name", "?"))
    await state.update_data(edit_probe_idx=idx)
    await state.set_state(WifiOfficeSetup.editing_password)
    await cb.message.answer(
        f"🔒 Новый пароль для <b>{loc}</b>\n\n"
        "Введи новый пароль роутера\n"
        "(или <code>-</code> если без пароля):",
        parse_mode="HTML", reply_markup=_CANCEL_KB,
    )
    await cb.answer()


@router.callback_query(F.data.startswith("wifi:chlogin:"))
async def cb_wifi_chlogin(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        return
    idx_str = cb.data[len("wifi:chlogin:"):]
    probes  = _load_keenetic_probes()
    try:
        idx = int(idx_str)
        p   = probes[idx]
    except (ValueError, IndexError):
        await cb.answer("Офис не найден", show_alert=True)
        return
    loc   = p.get("location", p.get("agent_name", "?"))
    login = p.get("router_login", "admin")
    await state.update_data(edit_probe_idx=idx, edit_field="login")
    await state.set_state(WifiOfficeSetup.editing_password)
    await cb.message.answer(
        f"🔑 Новый логин для <b>{loc}</b>\n\n"
        f"Текущий: <code>{login or '—'}</code>\n\n"
        "Введи новый логин роутера\n"
        "(или <code>-</code> если пустой):",
        parse_mode="HTML", reply_markup=_CANCEL_KB,
    )
    await cb.answer()


@router.message(WifiOfficeSetup.editing_password)
async def wifi_fsm_edit_field(msg: Message, state: FSMContext):
    if not is_admin(msg.from_user.id):
        return
    value = "" if msg.text.strip() == "-" else msg.text.strip()
    data  = await state.get_data()
    idx   = data.get("edit_probe_idx")
    field = data.get("edit_field", "password")  # "password" or "login"
    await state.clear()
    # Delete message if it was a password
    if field == "password":
        try:
            await msg.delete()
        except Exception:
            pass
    probes = _load_keenetic_probes()
    if idx is None or idx >= len(probes):
        await msg.answer("❌ Офис не найден.")
        return
    p   = probes[idx]
    loc = p.get("location", p.get("agent_name", "?"))
    if field == "login":
        p["router_login"] = value
        field_name = "Логин"
    else:
        p["router_password"] = value
        field_name = "Пароль"
    probes[idx] = p
    KEENETIC_PROBES_FILE.write_text(json.dumps(probes, ensure_ascii=False, indent=2))
    display = value if field == "login" else ("***" if value else "—")
    await msg.answer(
        f"✅ {field_name} для <b>{loc}</b> обновлён: <code>{display}</code>\n\n"
        "Данные обновятся при следующем опросе.",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="◀️ Зонды", callback_data="wifi:probes"),
        ]]),
    )


# ─── WiFi callback: delete probe ─────────────────────────────────────

@router.callback_query(F.data.startswith("wifi:del:"))
async def cb_wifi_del(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        return
    location = cb.data[len("wifi:del:"):]
    probes   = _load_keenetic_probes()
    to_del   = next((p for p in probes if p.get("location", p.get("agent_name")) == location), None)
    if to_del:
        aname = to_del.get("agent_name", "")
        probes = [p for p in probes if p.get("agent_name") != aname]
        KEENETIC_PROBES_FILE.write_text(json.dumps(probes, ensure_ascii=False, indent=2))
        global _wifi_clients
        _wifi_clients.pop(aname, None)
        _save_wifi_clients()
        await cb.answer(f"Удалён: {location}")
    else:
        await cb.answer("Не найден", show_alert=True)
    await cb_wifi_probes(cb)


# ─── Top Resources ──────────────────────────────────────────────────

@router.message(Command("top"))
async def cmd_top(msg: Message):
    """Show top resource usage across devices."""
    if not is_admin(msg.from_user.id):
        return
    devs = await get_full_devices()
    if not devs:
        await msg.answer("📭 Нет устройств.", reply_markup=MAIN_KB)
        return

    lines = ["━━━━━━━━━━━━━━━━━━━━━━\n📊 <b>Top Resources</b>\n━━━━━━━━━━━━━━━━━━━━━━\n"]

    # Top disk usage
    disk_devs = []
    for d in devs:
        for va in d.get("vol_alerts", []):
            disk_devs.append((d["name"], va))
    if disk_devs:
        lines.append("<b>💿 Диски заполнены:</b>")
        for name, alert in disk_devs[:5]:
            lines.append(f"  • {name}: {alert}")
        lines.append("")

    # Longest offline
    offline_devs = sorted(
        [d for d in devs if not d["online"] and d.get("offline_hours", 0) > 0],
        key=lambda x: x["offline_hours"], reverse=True,
    )
    if offline_devs:
        lines.append("<b>⏰ Долго офлайн:</b>")
        for d in offline_devs[:5]:
            lines.append(f"  • {d['name']}: {fmt_offline(d['offline_hours'])}")
        lines.append("")

    # Smallest RAM
    def parse_ram_gb(ram_str: str) -> float:
        ram_str = ram_str.strip()
        if "TB" in ram_str:
            return float(ram_str.replace("TB", "").strip()) * 1024
        if "GB" in ram_str:
            return float(ram_str.replace("GB", "").strip())
        if "MB" in ram_str:
            return float(ram_str.replace("MB", "").strip()) / 1024
        return 0

    ram_devs = sorted(
        [d for d in devs if parse_ram_gb(d["ram_total"]) > 0],
        key=lambda x: parse_ram_gb(x["ram_total"]),
    )
    if ram_devs:
        lines.append("<b>💾 Наименьшая RAM:</b>")
        for d in ram_devs[:5]:
            lines.append(f"  • {d['name']}: {d['ram_total']}")
        lines.append("")

    # Outdated agents
    agent_versions = {}
    for d in devs:
        v = d.get("agent_ver", "")
        if v:
            agent_versions.setdefault(v, []).append(d["name"])
    if len(agent_versions) > 1:
        latest_ver = max(agent_versions.keys())
        outdated = [(v, names) for v, names in agent_versions.items() if v != latest_ver]
        if outdated:
            lines.append("<b>🤖 Устаревшие агенты:</b>")
            for v, names in sorted(outdated):
                lines.append(f"  • v{v}: {', '.join(names[:5])}")
            lines.append("")

    await msg.answer("\n".join(lines), parse_mode="HTML", reply_markup=MAIN_KB)


# ─── Group Commands ─────────────────────────────────────────────────

@router.message(Command("run_group"))
async def cmd_run_group(msg: Message):
    """Run command on all online devices in a group: /run_group <group> [-ps] <cmd>"""
    if not is_admin(msg.from_user.id):
        return
    parts = msg.text.split(maxsplit=2)
    if len(parts) < 3:
        await msg.answer(
            "Использование:\n"
            "<code>/run_group Офис hostname</code>\n"
            "<code>/run_group Серверы -ps Get-Service</code>",
            parse_mode="HTML", reply_markup=MAIN_KB,
        )
        return

    group_name = parts[1].strip()
    rest = parts[2].strip()
    ps = False
    if rest.startswith("-ps "):
        ps = True
        rest = rest[4:].strip()

    if not rest:
        await msg.answer("❌ Не указана команда.", reply_markup=MAIN_KB)
        return

    devs = await get_full_devices()
    group_devs = [d for d in devs if d["group"].lower() == group_name.lower() and d["online"]]
    if not group_devs:
        await msg.answer(f"❌ Нет онлайн устройств в группе «{group_name}».", reply_markup=MAIN_KB)
        return

    wait_msg = await msg.answer(
        f"⏳ Выполняю на <b>{len(group_devs)}</b> устройствах в <b>{group_name}</b>...",
        parse_mode="HTML",
    )

    sem = asyncio.Semaphore(5)
    results = {}

    async def run_one(d):
        async with sem:
            result = await mc_run_command(d["id"], rest, powershell=ps)
            results[d["name"]] = result

    await asyncio.gather(*[run_one(d) for d in group_devs])

    lines = [
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📁 <b>Группа: {group_name}</b> ({len(results)} устройств)\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"<code>$ {rest}</code>\n"
    ]

    success = 0
    fail = 0
    for name, output in sorted(results.items()):
        first_line = ""
        for line in output.split("\n"):
            stripped = line.strip()
            skip_prefixes = ("Microsoft Windows", "(c) ", "C:\\Program Files\\Mesh Agent>")
            if stripped and not any(stripped.startswith(p) for p in skip_prefixes):
                if stripped != rest and stripped != "exit":
                    first_line = stripped[:60]
                    break
        if "Error" in output:
            fail += 1
            lines.append(f"❌ <b>{name}</b>: {first_line}")
        else:
            success += 1
            lines.append(f"✅ <b>{name}</b>: {first_line}")

    lines.insert(4, f"\n✅ {success} успешно, ❌ {fail} ошибок\n")

    text = "\n".join(lines)
    if len(text) > 4000:
        text = text[:4000] + "\n..."

    await wait_msg.edit_text(text, parse_mode="HTML")


# ─── Full inventory ─────────────────────────────────────────────────

@router.message(F.text == BTN_INVENTORY)
async def msg_inventory(msg: Message):
    if not is_admin(msg.from_user.id):
        return
    wait = await msg.answer("📦 Выгружаю полный инвентарь...")
    devs = await get_full_devices()
    if not devs:
        await wait.edit_text("📭 Нет устройств.")
        return
    csv_data = build_inventory_csv(devs)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    await msg.answer_document(
        BufferedInputFile(csv_data, filename=f"inventory_all_{ts}.csv"),
        caption=f"📦 <b>Полный инвентарь</b> — {len(devs)} устройств\n27 колонок • CSV (;) UTF-8 BOM для Excel",
        parse_mode="HTML",
    )
    json_data = json.dumps(devs, indent=2, ensure_ascii=False, default=str).encode()
    await msg.answer_document(
        BufferedInputFile(json_data, filename=f"inventory_all_{ts}.json"),
        caption="📋 JSON (полные данные)",
    )
    if HAS_OPENPYXL:
        xlsx_data = build_inventory_xlsx(devs)
        if xlsx_data:
            await msg.answer_document(
                BufferedInputFile(xlsx_data, filename=f"inventory_all_{ts}.xlsx"),
                caption=f"📊 <b>Excel-отчёт</b> — {len(devs)} устройств",
                parse_mode="HTML",
            )
    await wait.delete()


# ─── Health ──────────────────────────────────────────────────────────

@router.message(F.text == BTN_HEALTH)
async def msg_health(msg: Message):
    if not is_admin(msg.from_user.id):
        return
    alive = await mc_is_alive()
    svc = await mc_service_status()
    cpu = psutil.cpu_percent(interval=1)
    mem = psutil.virtual_memory()
    disk = shutil.disk_usage("/")
    l1, l5, l15 = os.getloadavg()
    net = psutil.net_io_counters()
    t = (
        "━━━━━━━━━━━━━━━━━━━━━━\n❤️  <b>Здоровье</b>\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"{'🟢' if alive else '🔴'} MC: <b>{'ALIVE' if alive else 'DOWN'}</b> ({svc})\n"
        f"🔄 Load: {l1:.1f}/{l5:.1f}/{l15:.1f}\n"
        f"🧠 CPU: {pbar(cpu)} {cpu:.0f}%\n"
        f"💾 RAM: {pbar(mem.percent)} {mem.percent:.0f}% ({fmt_bytes(mem.used)}/{fmt_bytes(mem.total)})\n"
        f"💿 Disk: {pbar(disk.used / disk.total * 100)} {disk.used / disk.total * 100:.0f}%\n\n"
        f"🌐 Net: ↓{fmt_bytes(net.bytes_recv)} ↑{fmt_bytes(net.bytes_sent)}\n\n"
        f"🛡 Автоперезапуск: Вкл  •  📦 Инвентарь: {INVENTORY_HOUR}:00 UTC  •  📋 Отчёт: {DAILY_REPORT_HOUR}:00 UTC"
    )
    await msg.answer(t, parse_mode="HTML", reply_markup=MAIN_KB)


# ─── Tools menu ──────────────────────────────────────────────────────

@router.message(F.text == BTN_TOOLS)
async def msg_tools(msg: Message):
    if not is_admin(msg.from_user.id):
        return
    buttons = [
        [InlineKeyboardButton(text="🔍 Поиск", callback_data="tool:search"),
         InlineKeyboardButton(text="⚖️ Сравнение", callback_data="tool:compare")],
        [InlineKeyboardButton(text="🔔 Алерты", callback_data="tool:alerts"),
         InlineKeyboardButton(text="📜 Изменения", callback_data="tool:changes")],
        [InlineKeyboardButton(text="🖥 Команда", callback_data="tool:run"),
         InlineKeyboardButton(text="📁 Группа CMD", callback_data="tool:run_group")],
        [InlineKeyboardButton(text="📥 Установщик", callback_data="tool:install"),
         InlineKeyboardButton(text="📝 Скрипты", callback_data="tool:scripts")],
        [InlineKeyboardButton(text="📄 PDF отчёт", callback_data="tool:pdf"),
         InlineKeyboardButton(text="📊 Excel", callback_data="tool:xlsx")],
        [InlineKeyboardButton(text="🛡 Безопасность", callback_data="tool:security"),
         InlineKeyboardButton(text="📈 Top", callback_data="tool:top")],
        [InlineKeyboardButton(text="🗺 Карта сети", callback_data="tool:netmap"),
         InlineKeyboardButton(text="🔇 Мьюты", callback_data="tool:mutes")],
        [InlineKeyboardButton(text="🏓 Пинг / Трейс", callback_data="tool:ping"),
         InlineKeyboardButton(text="📝 Заметки",    callback_data="tool:notes")],
        [InlineKeyboardButton(text="📈 Тренд дисков", callback_data="tool:disk_trend"),
         InlineKeyboardButton(text="📊 Сравнить 7д",   callback_data="tool:snap_compare")],
        [InlineKeyboardButton(text="⏰ Планировщик",    callback_data="tool:scheduler"),
         InlineKeyboardButton(text="📡 SNMP роутеры",  callback_data="tool:snmp")],
        [InlineKeyboardButton(text="🖨 Принтеры", callback_data="tool:printers"),
         InlineKeyboardButton(text="🔄 Рестарт MC", callback_data="tool:restart")],
        [InlineKeyboardButton(text="💻 Инвентарь HW",   callback_data="tool:hw_inventory"),
         InlineKeyboardButton(text="🌡 Температуры",    callback_data="tool:temperature")],
        [InlineKeyboardButton(text="📊 Доступность",    callback_data="tool:availability"),
         InlineKeyboardButton(text="🌐 Статус-страница",callback_data="tool:status_page")],
        [InlineKeyboardButton(text="💾 Бэкап MC", callback_data="tool:backup")],
        [InlineKeyboardButton(text="🗄 Полный бэкап сервера", callback_data="tool:fullbackup")],
        [InlineKeyboardButton(text="🆕 Обновления MC", callback_data="tool:update_check"),
         InlineKeyboardButton(text="🔐 SSL сертификаты", callback_data="tool:certs")],
        [InlineKeyboardButton(text="🚀 Развернуть копию", callback_data="tool:deploy")],
    ]
    await msg.answer(
        "━━━━━━━━━━━━━━━━━━━━━━\n🔧 <b>Инструменты</b>\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "🔍 /search &lt;запрос&gt; — поиск устройств\n"
        "⚖️ /compare &lt;PC1&gt; &lt;PC2&gt; — сравнение\n"
        "🖥 /run &lt;PC&gt; &lt;cmd&gt; — удалённая команда\n"
        "📁 /run_group &lt;группа&gt; &lt;cmd&gt; — команда группе\n"
        "📝 /scripts — быстрые скрипты\n"
        "📥 /install — установщик агента\n"
        "🔔 Алерты — настройка уведомлений\n"
        "🛡 Безопасность — сводка безопасности\n"
        "📈 /top — топ ресурсов\n"
        "📊 Excel — полный отчёт XLSX\n"
        "🗺 Карта сети — устройства по подсетям\n"
        "🔇 /mute &lt;цель&gt; &lt;время&gt; — тех. обслуживание\n"
        "📄 PDF — PDF-отчёт\n"
        "💾 Бэкап — конфиг MC в Telegram\n"
        "🆕 Обновления — проверка + обновление MC\n",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
    )


@router.callback_query(F.data == "tool:search")
async def cb_tool_search(cb: CallbackQuery):
    await cb.message.answer("🔍 Отправьте команду:\n<code>/search запрос</code>", parse_mode="HTML")
    await cb.answer()


@router.callback_query(F.data == "tool:compare")
async def cb_tool_compare(cb: CallbackQuery):
    await cb.message.answer("⚖️ Отправьте команду:\n<code>/compare PC1 PC2</code>", parse_mode="HTML")
    await cb.answer()


@router.callback_query(F.data == "tool:run")
async def cb_tool_run(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    devs = await get_full_devices()
    online = sorted([d for d in devs if d["online"]], key=lambda x: (x.get("group",""), x["name"]))
    if not online:
        await cb.message.answer("⚠️ Нет онлайн-устройств.")
        return
    rows = []
    cur_group = None
    for d in online:
        grp = d.get("group", "")
        if grp != cur_group:
            cur_group = grp
            rows.append([InlineKeyboardButton(text=f"── {grp} ──", callback_data="noop")])
        rows.append([InlineKeyboardButton(
            text=f"🟢 {d['name']}",
            callback_data=f"rcmd_pick:{d['name'][:55]}",
        )])
    await cb.message.answer(
        "🖥 <b>Выбери устройство для команды:</b>",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=rows),
    )


@router.callback_query(F.data == "tool:install")
async def cb_tool_install(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    groups = await _get_mesh_groups()
    if not groups:
        await cb.answer("Нет групп", show_alert=True)
        return
    buttons = []
    for name in sorted(groups.keys()):
        buttons.append([InlineKeyboardButton(text=f"📁 {name}", callback_data=f"inst:{name[:40]}")])
    buttons.append([InlineKeyboardButton(text="📁 Все группы", callback_data="inst:__all__")])
    await cb.message.answer(
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "📥 <b>Установщик агента</b>\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "Выберите группу.\n"
        "BAT для Windows 10/11 (включая 25H2).\n"
        "Обходит SmartScreen, Defender и баг KB5074105.",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
    )
    await cb.answer()


@router.callback_query(F.data == "tool:pdf")
async def cb_tool_pdf(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    wait_msg = await cb.message.answer("📄 Генерирую PDF-отчёт...")
    devs = await get_full_devices()
    if not devs:
        await wait_msg.edit_text("📭 Нет устройств.")
        await cb.answer()
        return
    pdf_data = build_inventory_pdf(devs)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    await cb.message.answer_document(
        BufferedInputFile(pdf_data, filename=f"inventory_{ts}.pdf"),
        caption=f"📄 <b>PDF-отчёт</b> — {len(devs)} устройств",
        parse_mode="HTML",
    )
    await wait_msg.delete()
    await cb.answer()


@router.callback_query(F.data == "tool:restart")
async def cb_tool_restart(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.message.answer("🔄 Перезапускаю MeshCentral...")
    await mc_restart()
    await asyncio.sleep(8)
    alive = await mc_is_alive()
    await cb.message.answer(f"{'🟢 OK' if alive else '🔴 Не отвечает (подождите 30с)'}", reply_markup=MAIN_KB)
    await cb.answer()


# ─── Update check ────────────────────────────────────────────────────

@router.callback_query(F.data == "tool:update_check")
async def cb_tool_update_check(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("\U0001f512", show_alert=True)
        return
    await cb.answer()
    wait_msg = await cb.message.answer("\u23f3 Проверяю обновления MeshCentral...")
    info = await check_mc_update()
    if info["current"] == "unknown":
        text = "\u274c Не удалось определить текущую версию MeshCentral."
    elif info["latest"] == "unknown":
        text = "\u274c Не удалось проверить npm-реестр."
    elif info["has_update"]:
        text = (
            "\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n"
            "\U0001f195 <b>Доступно обновление!</b>\n"
            "\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n\n"
            f"\U0001f4e6 Текущая: <b>{info['current']}</b>\n"
            f"\U0001f680 Новая: <b>{info['latest']}</b>"
        )
        await wait_msg.edit_text(text, parse_mode="HTML", reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔄 Обновить сейчас", callback_data="mc:update")],
        ]))
        return
    else:
        text = (
            "\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n"
            "\u2705 <b>MeshCentral актуален</b>\n"
            "\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n\n"
            f"\U0001f4e6 Версия: <b>{info['current']}</b>\n"
            "Обновлений нет."
        )
    await wait_msg.edit_text(text, parse_mode="HTML")


# ─── Alerts config ───────────────────────────────────────────────────

@router.callback_query(F.data == "tool:alerts")
async def cb_tool_alerts(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    cfg = load_alerts_cfg()
    t = (
        "━━━━━━━━━━━━━━━━━━━━━━\n🔔 <b>Алерты</b>\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"💿 Диск заполнен ≥ <b>{cfg['disk_pct']}%</b>\n"
        f"🛡 AV выключен: <b>{'Да' if cfg['av_off'] else 'Нет'}</b>\n"
        f"⏰ Офлайн > <b>{cfg['offline_hours']}ч</b>\n"
        f"🆕 Новое устройство: <b>{'Да' if cfg['new_device'] else 'Нет'}</b>\n"
    )
    buttons = [
        [InlineKeyboardButton(text=f"💿 Порог: {cfg['disk_pct']}%", callback_data="alert:disk_cycle"),
         InlineKeyboardButton(text="❓", callback_data="help:disk")],
        [InlineKeyboardButton(text=f"🛡 AV: {'ON' if cfg['av_off'] else 'OFF'}", callback_data="alert:av_toggle"),
         InlineKeyboardButton(text="❓", callback_data="help:av")],
        [InlineKeyboardButton(text=f"⏰ Офлайн: {cfg['offline_hours']}ч", callback_data="alert:offline_cycle"),
         InlineKeyboardButton(text="❓", callback_data="help:offline")],
        [InlineKeyboardButton(text=f"🆕 Новое: {'ON' if cfg['new_device'] else 'OFF'}", callback_data="alert:new_toggle"),
         InlineKeyboardButton(text="❓", callback_data="help:new_device")],
    ]
    await cb.message.answer(t, parse_mode="HTML", reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    await cb.answer()


@router.callback_query(F.data == "alert:disk_cycle")
async def cb_alert_disk(cb: CallbackQuery):
    cfg = load_alerts_cfg()
    cycle = [80, 85, 90, 95]
    idx = cycle.index(cfg["disk_pct"]) if cfg["disk_pct"] in cycle else 0
    cfg["disk_pct"] = cycle[(idx + 1) % len(cycle)]
    save_alerts_cfg(cfg)
    await cb.answer(f"Порог диска: {cfg['disk_pct']}%")
    # refresh
    await cb_tool_alerts(cb)


@router.callback_query(F.data == "alert:av_toggle")
async def cb_alert_av(cb: CallbackQuery):
    cfg = load_alerts_cfg()
    cfg["av_off"] = not cfg["av_off"]
    save_alerts_cfg(cfg)
    await cb.answer(f"AV алерт: {'ON' if cfg['av_off'] else 'OFF'}")
    await cb_tool_alerts(cb)


@router.callback_query(F.data == "alert:offline_cycle")
async def cb_alert_offline(cb: CallbackQuery):
    cfg = load_alerts_cfg()
    cycle = [6, 12, 24, 48, 72]
    idx = cycle.index(cfg["offline_hours"]) if cfg["offline_hours"] in cycle else 0
    cfg["offline_hours"] = cycle[(idx + 1) % len(cycle)]
    save_alerts_cfg(cfg)
    await cb.answer(f"Офлайн порог: {cfg['offline_hours']}ч")
    await cb_tool_alerts(cb)


@router.callback_query(F.data == "alert:new_toggle")
async def cb_alert_new(cb: CallbackQuery):
    cfg = load_alerts_cfg()
    cfg["new_device"] = not cfg["new_device"]
    save_alerts_cfg(cfg)
    await cb.answer(f"Новое устройство: {'ON' if cfg['new_device'] else 'OFF'}")
    await cb_tool_alerts(cb)


# ─── Changes ─────────────────────────────────────────────────────────

@router.callback_query(F.data == "tool:changes")
async def cb_tool_changes(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    devs = await get_full_devices()
    changes = detect_changes(devs)
    if not changes:
        await cb.message.answer("📜 Изменений не обнаружено с последнего снимка.", reply_markup=MAIN_KB)
    else:
        t = "━━━━━━━━━━━━━━━━━━━━━━\n📜 <b>Изменения</b>\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
        t += "\n".join(changes[:30])
        await cb.message.answer(t, parse_mode="HTML", reply_markup=MAIN_KB)
    # update snapshot after showing
    save_snapshot(devs)
    save_snap_history(devs)
    await cb.answer()


# ─── Backups ─────────────────────────────────────────────────────────

@router.callback_query(F.data == "tool:backup")
async def cb_tool_backup(cb: CallbackQuery):
    """Send latest MeshCentral auto-backup ZIP."""
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    backup_dir = Path("/opt/meshcentral/meshcentral-backups")
    zips = sorted(backup_dir.glob("*.zip"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not zips:
        await cb.message.answer("❌ Бэкапов MeshCentral не найдено.")
        return
    latest = zips[0]
    size_mb = latest.stat().st_size / 1_048_576
    wait_msg = await cb.message.answer(f"📦 Отправляю бэкап MC ({size_mb:.1f} MB)...")
    try:
        with open(latest, "rb") as f:
            data = f.read()
        await cb.message.answer_document(
            BufferedInputFile(data, filename=latest.name),
            caption=(
                f"💾 <b>Бэкап MeshCentral</b>\n"
                f"📁 {latest.name}\n"
                f"📦 {size_mb:.1f} MB\n"
                f"🔒 Пароль: в config.json → zipPassword\n\n"
                f"<i>Всего бэкапов: {len(zips)}</i>"
            ),
            parse_mode="HTML",
        )
    except Exception as e:
        await cb.message.answer(f"❌ Ошибка: {e}")
    await wait_msg.delete()


@router.callback_query(F.data == "tool:fullbackup")
async def cb_tool_fullbackup(cb: CallbackQuery):
    """Create and send full server backup: bots + WG configs + services + nginx."""
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    wait_msg = await cb.message.answer("🗄 Создаю полный бэкап сервера...")

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    zip_buf = io.BytesIO()

    # Files to include: (archive_path, real_path)
    entries: list[tuple[str, str]] = []

    # MeshCentral data (DB + config, skip signedagents and temp files)
    mc_data_dir = Path(MC_DATA)
    for p in mc_data_dir.iterdir():
        if p.is_file() and not p.name.endswith(".db~") and p.name != "mesherrors.txt":
            entries.append((f"meshcentral-data/{p.name}", str(p)))

    # meshcentral-bot
    mc_bot_dir = Path("/opt/meshcentral-bot")
    for fname in ["bot.py", ".env", "admin.json", "alerts_cfg.json",
                  "scripts.json", "snapshots.json", "mute.json",
                  "vis-network.min.js"]:
        p = mc_bot_dir / fname
        if p.exists():
            entries.append((f"meshcentral-bot/{fname}", str(p)))

    # awg-bot
    awg_bot_dir = Path("/opt/awg-bot")
    for fname in ["bot.py", ".env", "admin.json", "clients.json",
                  "history.json", "traffic_daily.json"]:
        p = awg_bot_dir / fname
        if p.exists():
            entries.append((f"awg-bot/{fname}", str(p)))
    # awg-bot backups (WG configs only)
    for p in (awg_bot_dir / "backups").glob("*.conf"):
        entries.append((f"awg-bot/backups/{p.name}", str(p)))

    # WireGuard / AmneziaWG configs
    for conf in ["/etc/amnezia/amneziawg/awg0.conf",
                 "/etc/amnezia/amneziawg/wg0.conf"]:
        if Path(conf).exists():
            entries.append((f"wireguard/{Path(conf).name}", conf))

    # systemd services
    for svc in ["meshcentral.service", "meshcentral-bot.service", "awg-bot.service",
                "awg-quick@awg0.service", "wg-quick@wg0.service"]:
        p = Path(f"/etc/systemd/system/{svc}")
        if p.exists():
            entries.append((f"services/{svc}", str(p)))

    # nginx
    p = Path("/etc/nginx/sites-enabled/meshcentral")
    if p.exists():
        entries.append(("nginx/meshcentral.conf", str(p)))
    entries.append(("nginx/nginx.conf", "/etc/nginx/nginx.conf"))

    # RESTORE.md
    restore_md = f"""# Инструкция восстановления сервера
Создан: {datetime.now().strftime('%d.%m.%Y %H:%M')}

## 1. Зависимости
```bash
apt update && apt install -y nodejs npm nginx python3 python3-venv python3-pip certbot python3-certbot-nginx
npm install -g meshcentral
# AmneziaWG: https://github.com/amnezia-vpn/amneziawg-linux-kernel-module
```

## 2. MeshCentral
```bash
mkdir -p /opt/meshcentral
cp -r meshcentral-data/ /opt/meshcentral/meshcentral-data/
cp services/meshcentral.service /etc/systemd/system/
systemctl daemon-reload && systemctl enable --now meshcentral
```

## 3. Боты
```bash
# meshcentral-bot
mkdir -p /opt/meshcentral-bot
cp meshcentral-bot/* /opt/meshcentral-bot/
cd /opt/meshcentral-bot && python3 -m venv venv
venv/bin/pip install aiogram aiohttp python-dotenv fpdf2 matplotlib psutil pyzipper openpyxl
cp services/meshcentral-bot.service /etc/systemd/system/
systemctl daemon-reload && systemctl enable --now meshcentral-bot

# awg-bot
mkdir -p /opt/awg-bot
cp awg-bot/* /opt/awg-bot/
cd /opt/awg-bot && python3 -m venv venv
venv/bin/pip install aiogram python-dotenv
cp services/awg-bot.service /etc/systemd/system/
systemctl daemon-reload && systemctl enable --now awg-bot
```

## 4. WireGuard / AmneziaWG
```bash
mkdir -p /etc/amnezia/amneziawg
cp wireguard/awg0.conf /etc/amnezia/amneziawg/
cp wireguard/wg0.conf /etc/amnezia/amneziawg/
systemctl enable --now awg-quick@awg0 wg-quick@wg0
```

## 5. Nginx
```bash
cp nginx/meshcentral.conf /etc/nginx/sites-enabled/meshcentral
cp nginx/nginx.conf /etc/nginx/nginx.conf
certbot --nginx -d hub.office.mooo.com
nginx -t && systemctl reload nginx
```

## 6. Firewall
```bash
ufw allow 22/tcp && ufw allow 80/tcp && ufw allow 443/tcp
ufw allow 443/udp && ufw allow 9443/udp && ufw allow 51820/udp
ufw route allow in on wg0 out on ens3
ufw route allow in on ens3 out on wg0
ufw route allow in on awg0 out on ens3
ufw route allow in on ens3 out on awg0
ufw enable
```

## Важно
- Обновите IP в WG конфигах если он изменился
- Обновите DNS запись hub.office.mooo.com на новый IP
- Пароль бэкапов MeshCentral: см config.json → zipPassword
"""
    entries.append(("RESTORE.md", None))  # special — write from string

    try:
        with pyzipper.AESZipFile(zip_buf, "w",
                                  compression=pyzipper.ZIP_DEFLATED,
                                  encryption=pyzipper.WZ_AES) as zf:
            zf.setpassword(b"Kh@mzat88712Pass")
            for arc_path, real_path in entries:
                if real_path is None:
                    zf.writestr(arc_path, restore_md)
                elif Path(real_path).exists():
                    zf.write(real_path, arc_path)
        zip_buf.seek(0)
        size_mb = zip_buf.getbuffer().nbytes / 1_048_576
        await cb.message.answer_document(
            BufferedInputFile(zip_buf.read(), filename=f"server_backup_{ts}.zip"),
            caption=(
                f"🗄 <b>Полный бэкап сервера</b>\n"
                f"📅 {ts}\n"
                f"📦 {size_mb:.1f} MB\n"
                f"🔒 Пароль: <code>Kh@mzat88712Pass</code>\n\n"
                f"<b>Содержимое:</b>\n"
                f"• meshcentral-data/ — БД и конфиги MC\n"
                f"• meshcentral-bot/ — бот + данные\n"
                f"• awg-bot/ — AWG бот + клиенты\n"
                f"• wireguard/ — конфиги WG/AWG\n"
                f"• services/ — systemd сервисы\n"
                f"• nginx/ — конфиг nginx\n"
                f"• RESTORE.md — инструкция восстановления"
            ),
            parse_mode="HTML",
        )
    except Exception as e:
        await cb.message.answer(f"❌ Ошибка создания бэкапа: {e}")
    await wait_msg.delete()


# ─── Tool callbacks: Security, Top, RunGroup, Excel, NetMap, Scripts, Mutes ──

@router.callback_query(F.data == "tool:security")
async def cb_tool_security(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    wait_msg = await cb.message.answer("🛡 Анализирую безопасность...")
    devs = await get_full_devices()
    if not devs:
        await wait_msg.edit_text("📭 Нет устройств.")
        return

    lines = ["━━━━━━━━━━━━━━━━━━━━━━\n🛡 <b>Сводка безопасности</b>\n━━━━━━━━━━━━━━━━━━━━━━\n"]

    # AV disabled
    av_off = [d for d in devs if d.get("av_disabled")]
    if av_off:
        lines.append(f"<b>🛡 Антивирус выключен ({len(av_off)}):</b>")
        for d in av_off[:10]:
            lines.append(f"  • {d['name']}: {d['antivirus']}")
        lines.append("")

    # Firewall issues
    fw_issues = [d for d in devs if d.get("firewall", "").lower() not in ("", "on", "enabled", "включён")]
    fw_off = [d for d in fw_issues if "off" in d.get("firewall", "").lower() or "выкл" in d.get("firewall", "").lower()]
    if fw_off:
        lines.append(f"<b>🔥 Firewall выключен ({len(fw_off)}):</b>")
        for d in fw_off[:10]:
            lines.append(f"  • {d['name']}: {d['firewall']}")
        lines.append("")

    # No TPM 2.0
    no_tpm = [d for d in devs if d.get("tpm", "") and "2.0" not in d.get("tpm", "")]
    if no_tpm:
        lines.append(f"<b>🔐 Нет TPM 2.0 ({len(no_tpm)}):</b>")
        for d in no_tpm[:10]:
            lines.append(f"  • {d['name']}: {d['tpm'] or 'не обнаружен'}")
        lines.append("")

    # Outdated agents
    agent_versions = {}
    for d in devs:
        v = d.get("agent_ver", "")
        if v:
            agent_versions.setdefault(v, []).append(d["name"])
    if len(agent_versions) > 1:
        latest_ver = max(agent_versions.keys())
        outdated = {v: names for v, names in agent_versions.items() if v != latest_ver}
        if outdated:
            total_outdated = sum(len(n) for n in outdated.values())
            lines.append(f"<b>🤖 Устаревшие агенты ({total_outdated}):</b>")
            for v, names in sorted(outdated.items()):
                lines.append(f"  • v{v}: {', '.join(names[:5])}")
            lines.append("")

    # Disk alerts
    disk_alerts = [d for d in devs if d.get("vol_alerts")]
    if disk_alerts:
        lines.append(f"<b>💿 Диск заполнен ({len(disk_alerts)}):</b>")
        for d in disk_alerts[:10]:
            lines.append(f"  • {d['name']}: {', '.join(d['vol_alerts'])}")
        lines.append("")

    if len(lines) == 1:
        lines.append("✅ Проблем безопасности не обнаружено!")

    await wait_msg.edit_text("\n".join(lines), parse_mode="HTML")


@router.callback_query(F.data == "tool:top")
async def cb_tool_top(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    # Reuse cmd_top logic by creating a fake message-like call
    devs = await get_full_devices()
    if not devs:
        await cb.message.answer("📭 Нет устройств.", reply_markup=MAIN_KB)
        return

    lines = ["━━━━━━━━━━━━━━━━━━━━━━\n📊 <b>Top Resources</b>\n━━━━━━━━━━━━━━━━━━━━━━\n"]

    disk_devs = []
    for d in devs:
        for va in d.get("vol_alerts", []):
            disk_devs.append((d["name"], va))
    if disk_devs:
        lines.append("<b>💿 Диски заполнены:</b>")
        for name, alert in disk_devs[:5]:
            lines.append(f"  • {name}: {alert}")
        lines.append("")

    offline_devs = sorted(
        [d for d in devs if not d["online"] and d.get("offline_hours", 0) > 0],
        key=lambda x: x["offline_hours"], reverse=True,
    )
    if offline_devs:
        lines.append("<b>⏰ Долго офлайн:</b>")
        for d in offline_devs[:5]:
            lines.append(f"  • {d['name']}: {fmt_offline(d['offline_hours'])}")
        lines.append("")

    def _parse_ram(ram_str: str) -> float:
        ram_str = ram_str.strip()
        if "TB" in ram_str:
            return float(ram_str.replace("TB", "").strip()) * 1024
        if "GB" in ram_str:
            return float(ram_str.replace("GB", "").strip())
        if "MB" in ram_str:
            return float(ram_str.replace("MB", "").strip()) / 1024
        return 0

    ram_devs = sorted(
        [d for d in devs if _parse_ram(d["ram_total"]) > 0],
        key=lambda x: _parse_ram(x["ram_total"]),
    )
    if ram_devs:
        lines.append("<b>💾 Наименьшая RAM:</b>")
        for d in ram_devs[:5]:
            lines.append(f"  • {d['name']}: {d['ram_total']}")
        lines.append("")

    agent_versions = {}
    for d in devs:
        v = d.get("agent_ver", "")
        if v:
            agent_versions.setdefault(v, []).append(d["name"])
    if len(agent_versions) > 1:
        latest_ver = max(agent_versions.keys())
        outdated = [(v, names) for v, names in agent_versions.items() if v != latest_ver]
        if outdated:
            lines.append("<b>🤖 Устаревшие агенты:</b>")
            for v, names in sorted(outdated):
                lines.append(f"  • v{v}: {', '.join(names[:5])}")
            lines.append("")

    await cb.message.answer("\n".join(lines), parse_mode="HTML", reply_markup=MAIN_KB)


@router.callback_query(F.data == "tool:run_group")
async def cb_tool_run_group(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    devs = await get_full_devices()
    groups = sorted(set(d.get("group", "") for d in devs if d.get("group")))
    if not groups:
        await cb.message.answer("⚠️ Нет групп.")
        return
    rows = [[InlineKeyboardButton(text=f"📁 {g}", callback_data=f"rgrp_pick:{g[:55]}")] for g in groups]
    await cb.message.answer(
        "📁 <b>Выбери группу для команды:</b>",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=rows),
    )


@router.callback_query(F.data.startswith("rcmd_pick:"))
async def cb_rcmd_pick(cb: CallbackQuery, state: FSMContext):
    """User selected a device — show preset commands."""
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    dev_name = cb.data[len("rcmd_pick:"):]
    await state.update_data(rcmd_device=dev_name)
    preset_text, preset_btns = _rcmd_preset_keyboard("rcmd_ps", "rcmd_cancel")
    await cb.message.answer(
        f"🖥 <b>{dev_name}</b> — выбери команду:{preset_text}",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=preset_btns),
    )


@router.callback_query(F.data.startswith("rgrp_pick:"))
async def cb_rgrp_pick(cb: CallbackQuery, state: FSMContext):
    """User selected a group — show preset commands."""
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    group_name = cb.data[len("rgrp_pick:"):]
    await state.update_data(rcmd_group=group_name)
    preset_text, preset_btns = _rcmd_preset_keyboard("rgrp_ps", "rcmd_cancel")
    await cb.message.answer(
        f"📁 <b>Группа: {group_name}</b> — выбери команду:{preset_text}",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=preset_btns),
    )


@router.callback_query(F.data.startswith("rcmd_ps:"))
async def cb_rcmd_ps(cb: CallbackQuery, state: FSMContext):
    """Run a preset script on the previously selected device."""
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    script_name = cb.data[len("rcmd_ps:"):]
    data = await state.get_data()
    dev_name = data.get("rcmd_device", "")
    if not dev_name:
        await cb.answer("Устройство не выбрано. Начни заново.", show_alert=True)
        return
    scripts = load_scripts()
    if script_name not in scripts:
        await cb.answer("Скрипт не найден", show_alert=True)
        return
    await cb.answer()
    devs = await get_full_devices()
    d = next((x for x in devs if x["name"] == dev_name), None)
    if not d or not d["online"]:
        await cb.message.answer(f"⚪ <b>{dev_name}</b> офлайн.", parse_mode="HTML")
        return
    s = scripts[script_name]
    wait_msg = await cb.message.answer(
        f"⏳ <b>{script_name}</b> на <b>{d['name']}</b>...", parse_mode="HTML",
    )
    result = await mc_run_command(d["id"], s["cmd"], powershell=s.get("ps", False))
    lines = result.split("\n")
    clean = []
    skip_prefixes = ("Microsoft Windows", "(c) ", "C:\\Program Files\\Mesh Agent>")
    for line in lines:
        stripped = line.strip()
        if stripped and not any(stripped.startswith(p) for p in skip_prefixes):
            if stripped != s["cmd"] and stripped != "exit":
                clean.append(line)
    output = "\n".join(clean).strip() or "(пустой ответ)"
    if len(output) > 3800:
        output = output[:3800] + "\n..."
    lang = "PS" if s.get("ps") else "CMD"
    await wait_msg.edit_text(
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🖥 <b>{d['name']}</b>  📝 <b>{script_name}</b> [{lang}]\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"<pre>{output}</pre>",
        parse_mode="HTML",
    )


@router.callback_query(F.data.startswith("rgrp_ps:"))
async def cb_rgrp_ps(cb: CallbackQuery, state: FSMContext):
    """Run a preset script on all online devices in the previously selected group."""
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    script_name = cb.data[len("rgrp_ps:"):]
    data = await state.get_data()
    group_name = data.get("rcmd_group", "")
    if not group_name:
        await cb.answer("Группа не выбрана. Начни заново.", show_alert=True)
        return
    scripts = load_scripts()
    if script_name not in scripts:
        await cb.answer("Скрипт не найден", show_alert=True)
        return
    await cb.answer()
    devs = await get_full_devices()
    group_devs = [d for d in devs if d.get("group", "").lower() == group_name.lower() and d["online"]]
    if not group_devs:
        await cb.message.answer(f"❌ Нет онлайн устройств в группе «{group_name}».")
        return
    s = scripts[script_name]
    wait_msg = await cb.message.answer(
        f"⏳ <b>{script_name}</b> на <b>{len(group_devs)}</b> уст. в <b>{group_name}</b>...",
        parse_mode="HTML",
    )
    sem = asyncio.Semaphore(5)
    results = {}
    async def run_one(d):
        async with sem:
            results[d["name"]] = await mc_run_command(d["id"], s["cmd"], powershell=s.get("ps", False))
    await asyncio.gather(*[run_one(d) for d in group_devs])

    skip_prefixes = ("Microsoft Windows", "(c) ", "C:\\Program Files\\Mesh Agent>")
    lang = "PS" if s.get("ps") else "CMD"
    lines = [
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📁 <b>{group_name}</b>  📝 <b>{script_name}</b> [{lang}]  ({len(results)} уст.)\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
    ]
    success = fail = 0
    for name, output in sorted(results.items()):
        first_line = ""
        for line in output.split("\n"):
            stripped = line.strip()
            if stripped and not any(stripped.startswith(p) for p in skip_prefixes):
                if stripped != s["cmd"] and stripped != "exit":
                    first_line = stripped[:60]
                    break
        if "Error" in output:
            fail += 1
            lines.append(f"❌ <b>{name}</b>: {first_line}")
        else:
            success += 1
            lines.append(f"✅ <b>{name}</b>: {first_line}")
    lines.append(f"\n✅ {success}  ❌ {fail}")
    text = "\n".join(lines)
    if len(text) > 4000:
        text = text[:4000] + "\n..."
    await wait_msg.edit_text(text, parse_mode="HTML")


@router.callback_query(F.data == "rcmd_custom")
async def cb_rcmd_custom(cb: CallbackQuery, state: FSMContext):
    """Switch to manual command input for device."""
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    data = await state.get_data()
    dev_name = data.get("rcmd_device", "")
    await state.set_state(RemoteCmdState.entering_cmd)
    await cb.answer()
    await cb.message.answer(
        f"🖥 <b>{dev_name}</b>\n\n"
        "Введи команду (добавь <code>-ps</code> в начале для PowerShell):\n"
        "<i>Пример: ipconfig /all\n"
        "Пример PS: -ps Get-Process | Sort CPU -Desc | Select -First 5</i>",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Отмена", callback_data="rcmd_cancel")],
        ]),
    )


@router.callback_query(F.data == "rgrp_custom")
async def cb_rgrp_custom(cb: CallbackQuery, state: FSMContext):
    """Switch to manual command input for group."""
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    data = await state.get_data()
    group_name = data.get("rcmd_group", "")
    await state.set_state(RemoteCmdState.entering_gcmd)
    await cb.answer()
    await cb.message.answer(
        f"📁 <b>Группа: {group_name}</b>\n\n"
        "Введи команду для всех онлайн-устройств группы\n"
        "(<code>-ps</code> в начале — PowerShell):\n"
        "<i>Пример: hostname\nПример PS: -ps (Get-Date).ToString()</i>",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Отмена", callback_data="rcmd_cancel")],
        ]),
    )


@router.callback_query(F.data == "rcmd_cancel")
async def cb_rcmd_cancel(cb: CallbackQuery, state: FSMContext):
    await state.clear()
    await cb.answer("Отменено")
    await cb.message.delete()


@router.message(RemoteCmdState.entering_cmd)
async def fsm_rcmd_entering_cmd(msg: Message, state: FSMContext):
    """User typed a command for a specific device."""
    if not is_admin(msg.from_user.id):
        return
    data = await state.get_data()
    dev_name = data.get("rcmd_device", "")
    await state.clear()

    command = msg.text.strip()
    powershell = False
    if command.lower().startswith("-ps "):
        powershell = True
        command = command[4:].strip()
    if not command:
        await msg.answer("❌ Пустая команда.", reply_markup=MAIN_KB)
        return

    devs = await get_full_devices()
    d = next((x for x in devs if x["name"] == dev_name), None)
    if not d:
        await msg.answer(f"❌ Устройство «{dev_name}» не найдено.", reply_markup=MAIN_KB)
        return
    if not d["online"]:
        await msg.answer(f"⚪ <b>{d['name']}</b> офлайн.", parse_mode="HTML", reply_markup=MAIN_KB)
        return

    wait_msg = await msg.answer(
        f"⏳ Выполняю на <b>{d['name']}</b>:\n<code>{command}</code>",
        parse_mode="HTML",
    )
    result = await mc_run_command(d["id"], command, powershell=powershell)

    lines = result.split("\n")
    clean = []
    skip_prefixes = ("Microsoft Windows", "(c) ", "C:\\Program Files\\Mesh Agent>")
    for line in lines:
        stripped = line.strip()
        if stripped and not any(stripped.startswith(p) for p in skip_prefixes):
            if stripped != command and stripped != "exit":
                clean.append(line)
    output = "\n".join(clean).strip() or "(пустой ответ)"
    if len(output) > 3800:
        output = output[:3800] + "\n..."

    lang = "PS" if powershell else "CMD"
    await wait_msg.edit_text(
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🖥 <b>{d['name']}</b> [{lang}]\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"<code>$ {command}</code>\n\n"
        f"<code>{output}</code>",
        parse_mode="HTML",
    )


@router.message(RemoteCmdState.entering_gcmd)
async def fsm_rcmd_entering_gcmd(msg: Message, state: FSMContext):
    """User typed a command for an entire group."""
    if not is_admin(msg.from_user.id):
        return
    data = await state.get_data()
    group_name = data.get("rcmd_group", "")
    await state.clear()

    command = msg.text.strip()
    powershell = False
    if command.lower().startswith("-ps "):
        powershell = True
        command = command[4:].strip()
    if not command:
        await msg.answer("❌ Пустая команда.", reply_markup=MAIN_KB)
        return

    devs = await get_full_devices()
    group_devs = [d for d in devs if d.get("group", "").lower() == group_name.lower() and d["online"]]
    if not group_devs:
        await msg.answer(f"❌ Нет онлайн устройств в группе «{group_name}».", reply_markup=MAIN_KB)
        return

    wait_msg = await msg.answer(
        f"⏳ Выполняю на <b>{len(group_devs)}</b> устройствах в <b>{group_name}</b>...",
        parse_mode="HTML",
    )

    sem = asyncio.Semaphore(5)
    results = {}

    async def run_one(d):
        async with sem:
            result = await mc_run_command(d["id"], command, powershell=powershell)
            results[d["name"]] = result

    await asyncio.gather(*[run_one(d) for d in group_devs])

    lang = "PS" if powershell else "CMD"
    lines = [
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📁 <b>Группа: {group_name}</b> ({len(results)} уст.) [{lang}]\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"<code>$ {command}</code>\n"
    ]

    skip_prefixes = ("Microsoft Windows", "(c) ", "C:\\Program Files\\Mesh Agent>")
    success = 0
    fail = 0
    for name, output in sorted(results.items()):
        first_line = ""
        for line in output.split("\n"):
            stripped = line.strip()
            if stripped and not any(stripped.startswith(p) for p in skip_prefixes):
                if stripped != command and stripped != "exit":
                    first_line = stripped[:60]
                    break
        if "Error" in output:
            fail += 1
            lines.append(f"❌ <b>{name}</b>: {first_line}")
        else:
            success += 1
            lines.append(f"✅ <b>{name}</b>: {first_line}")

    lines.append(f"\n✅ {success}  ❌ {fail}")
    text = "\n".join(lines)
    if len(text) > 4000:
        text = text[:4000] + "\n..."
    await wait_msg.edit_text(text, parse_mode="HTML")


@router.callback_query(F.data == "tool:xlsx")
async def cb_tool_xlsx(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    if not HAS_OPENPYXL:
        await cb.answer("openpyxl не установлен", show_alert=True)
        return
    await cb.answer()
    wait_msg = await cb.message.answer("📊 Генерирую Excel-отчёт...")
    devs = await get_full_devices()
    if not devs:
        await wait_msg.edit_text("📭 Нет устройств.")
        return
    xlsx_data = build_inventory_xlsx(devs)
    if not xlsx_data:
        await wait_msg.edit_text("❌ Ошибка генерации XLSX.")
        return
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    await cb.message.answer_document(
        BufferedInputFile(xlsx_data, filename=f"inventory_{ts}.xlsx"),
        caption=f"📊 <b>Excel-отчёт</b> — {len(devs)} устройств",
        parse_mode="HTML",
    )
    await wait_msg.delete()


@router.callback_query(F.data == "tool:netmap")
async def cb_tool_netmap(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    wait_msg = await cb.message.answer("🗺 Строю интерактивную карту сети...")
    devs = await get_full_devices()
    if not devs:
        await wait_msg.edit_text("📭 Нет устройств.")
        return
    # Send interactive HTML map
    html = build_network_map_html(devs)
    if html:
        await cb.message.answer_document(
            BufferedInputFile(html.encode("utf-8"), filename="network_map.html"),
            caption="🗺 <b>Интерактивная карта сети</b>\n\n"
                    "Откройте файл в браузере:\n"
                    "• Scroll — масштаб\n"
                    "• Перетаскивание — перемещение\n"
                    "• Клик на устройство — подробности",
            parse_mode="HTML",
        )
    # Also send PNG as preview
    png = build_network_map(devs)
    if png:
        await cb.message.answer_photo(
            BufferedInputFile(png, filename="network_map.png"),
            caption="🗺 <b>Превью карты сети</b> (PNG)",
            parse_mode="HTML",
        )
    await wait_msg.delete()


SCRIPT_CATEGORIES = {
    "system":      "🖥 Система",
    "network":     "🌐 Сеть",
    "maintenance": "🔧 Обслуживание",
    "printers":    "🖨 Принтеры",
    "security":    "🛡 Безопасность",
    "custom":      "⚙️ Пользовательские",
}


def _scripts_message(scripts: dict) -> tuple[str, list]:
    """Build text + buttons for scripts list, grouped by category."""
    cat_scripts: dict[str, list] = {}
    for name, s in sorted(scripts.items()):
        cat = s.get("cat", "custom")
        cat_scripts.setdefault(cat, []).append((name, s))
    lines = ["━━━━━━━━━━━━━━━━━━━━━━\n📝 <b>Скрипты</b>\n━━━━━━━━━━━━━━━━━━━━━━"]
    buttons = []
    for cat_key in ["system", "network", "maintenance", "printers", "security", "custom"]:
        entries = cat_scripts.get(cat_key)
        if not entries:
            continue
        cat_label = SCRIPT_CATEGORIES.get(cat_key, cat_key)
        lines.append(f"\n{cat_label}")
        for name, s in entries:
            lang = "⚡PS" if s.get("ps") else "🖥CMD"
            desc = s.get("desc", "")
            lines.append(f"  <b>{name}</b> [{lang}]  <i>{desc}</i>")
            buttons.append([
                InlineKeyboardButton(text=f"▶️ {name}", callback_data=f"srun:{name[:55]}"),
                InlineKeyboardButton(text="🗑",          callback_data=f"sdel:{name[:55]}"),
            ])
    return "\n".join(lines), buttons


def _rcmd_preset_keyboard(prefix: str, cancel_cb: str) -> tuple[str, list]:
    """Build text + buttons for preset command picker (rcmd_ps / rgrp_ps).

    prefix: 'rcmd_ps' or 'rgrp_ps'
    cancel_cb: callback_data for cancel button
    """
    scripts = load_scripts()
    cat_scripts: dict[str, list] = {}
    for name, s in sorted(scripts.items()):
        cat = s.get("cat", "custom")
        cat_scripts.setdefault(cat, []).append((name, s))

    lines = []
    buttons = []
    for cat_key in ["system", "network", "maintenance", "printers", "security", "custom"]:
        entries = cat_scripts.get(cat_key)
        if not entries:
            continue
        cat_label = SCRIPT_CATEGORIES.get(cat_key, cat_key)
        lines.append(f"\n{cat_label}")
        # Two scripts per row for compact display
        row = []
        for name, s in entries:
            desc = s.get("desc", "")
            lang = "⚡" if s.get("ps") else "🖥"
            row.append(InlineKeyboardButton(
                text=f"{lang} {name}",
                callback_data=f"{prefix}:{name[:50]}",
            ))
            if len(row) == 2:
                buttons.append(row)
                row = []
            lines.append(f"  <b>{name}</b>  <i>{desc}</i>")
        if row:
            buttons.append(row)

    buttons.append([
        InlineKeyboardButton(text="✏️ Своя команда", callback_data=cancel_cb.replace("cancel", "custom")),
        InlineKeyboardButton(text="❌ Отмена",        callback_data=cancel_cb),
    ])
    return "\n".join(lines), buttons


@router.callback_query(F.data == "tool:scripts")
async def cb_tool_scripts(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    scripts = load_scripts()
    if not scripts:
        await cb.message.answer("📝 Нет скриптов.", parse_mode="HTML")
        return
    text, buttons = _scripts_message(scripts)
    await cb.message.answer(text, parse_mode="HTML",
                             reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))


@router.callback_query(F.data == "tool:mutes")
async def cb_tool_mutes(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    cleanup_expired_mutes()
    mutes = load_mutes()
    add_btn = [InlineKeyboardButton(text="➕ Добавить мьют", callback_data="mute:add")]
    header = (
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "🔇 <b>Мьюты алертов</b>\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "<b>Что такое мьют?</b>\n"
        "Бот следит за устройствами и присылает алерты когда ПК уходит в офлайн, "
        "заканчивается место на диске или происходят другие события.\n\n"
        "Мьют — это временное отключение алертов для конкретного устройства, "
        "группы или всех устройств сразу.\n\n"
        "<b>Когда нужен мьют?</b>\n"
        "• Плановое обслуживание (апдейты, перезагрузка)\n"
        "• ПК временно выключен (выходные, праздники)\n"
        "• Тест оборудования\n\n"
    )
    if not mutes:
        await cb.message.answer(
            header + "Активных мьютов нет.",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[add_btn]),
        )
        return
    lines = [header + "<b>Активные мьюты:</b>\n"]
    buttons = []
    for target, info in sorted(mutes.items()):
        until = datetime.fromtimestamp(info["until"], tz=timezone.utc).strftime("%d.%m %H:%M UTC")
        reason = info.get("reason", "")
        disp = "🌐 Все устройства" if target == "__all__" else target
        lines.append(f"• <b>{disp}</b> до {until}{' — ' + reason if reason else ''}")
        buttons.append([InlineKeyboardButton(text=f"🔊 Включить алерты: {disp[:30]}", callback_data=f"unmute:{target[:40]}")])
    buttons.append(add_btn)
    await cb.message.answer(
        "\n".join(lines),
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
    )


@router.callback_query(F.data.startswith("unmute:"))
async def cb_unmute(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    target = cb.data.split(":", 1)[1]
    mutes = load_mutes()
    if target in mutes:
        del mutes[target]
        save_mutes(mutes)
        await cb.answer(f"{target} — алерты включены")
        await cb.message.delete()
    else:
        await cb.answer("Не найден", show_alert=True)


# ─── Printer discovery ───────────────────────────────────────────────

# PowerShell command: get all installed printers with details
_PRINTER_SCAN_CMD = (
    "Get-Printer | Select-Object Name, DriverName, PortName, PrinterStatus, Shared, "
    "@{N='Default';E={$_.Attributes -band 4 -gt 0}} | ConvertTo-Json -Compress"
)


def _load_printers() -> dict:
    return _load_json(PRINTERS_FILE, {})


def _save_printers(data: dict):
    _save_json(PRINTERS_FILE, data)


def _printer_status_str(status) -> str:
    statuses = {0: "✅ Готов", 1: "⚠️ Пауза", 2: "❌ Ошибка", 3: "⏳ Удаление",
                4: "🔌 Нет бумаги", 5: "⚠️ Занят", 6: "🖨 Печать", 7: "⏸ Офлайн"}
    try:
        return statuses.get(int(status), f"❓ {status}")
    except Exception:
        return str(status)


def _supply_emoji(desc: str) -> str:
    d = desc.lower()
    if any(w in d for w in ("black", "noir", "schwarz", "negro", "nero", "чёрн", "черн", "bk", " k")):
        return "🖤"
    if "cyan" in d or " c" == d[-2:]:
        return "🔵"
    if any(w in d for w in ("magenta", "пурпур", " m")):
        return "🔴"
    if any(w in d for w in ("yellow", "jaune", "gelb", "жёлт", "желт", " y")):
        return "🟡"
    if "drum" in d or "барабан" in d or "imaging" in d:
        return "🥁"
    if "fuser" in d or "фьюзер" in d:
        return "🔥"
    if "waste" in d or "отход" in d:
        return "🗑"
    return "🖨"


def _ink_bar(pct: int) -> str:
    """Return visual progress bar for ink level."""
    if pct < 0:
        return "❓ N/A"
    filled = round(pct / 10)
    bar = "█" * filled + "░" * (10 - filled)
    if pct >= 40:
        lvl = "🟢"
    elif pct >= INK_WARN_PCT:
        lvl = "🟡"
    else:
        lvl = "🔴"
    return f"{lvl} {bar} {pct}%"


def _load_ink_alerts() -> dict:
    return _load_json(INK_ALERTS_FILE, {})


def _save_ink_alerts(data: dict):
    _save_json(INK_ALERTS_FILE, data)


def _get_printer_scan_cmd() -> str:
    """Return PowerShell command: use printer_ink.ps1 if available, else basic fallback."""
    if PRINTER_INK_PS1.exists():
        return PRINTER_INK_PS1.read_text(encoding="utf-8")
    return (
        "Get-Printer | Select-Object Name, DriverName, PortName, PrinterStatus, Shared, "
        "@{N='Default';E={$_.Attributes -band 4 -gt 0}} | ConvertTo-Json -Compress"
    )


def _format_printer_card(p: dict, detailed: bool = False) -> str:
    """Format one printer entry with ink bars."""
    default_mark = " ⭐" if p.get("default") else ""
    shared_mark  = " 🔗" if p.get("shared") else ""
    status       = _printer_status_str(p.get("status", 0))
    ip_str       = f" | IP: <code>{p['printer_ip']}</code>" if p.get("printer_ip") else ""
    lines        = [f"  🖨 <b>{p['name']}</b>{default_mark}{shared_mark}"]
    if detailed:
        lines.append(f"     {status}{ip_str}")
        lines.append(f"     Драйвер: <i>{p.get('driver', '—')}</i>")
    else:
        lines.append(f"     {status}{ip_str}")
    supplies = p.get("supplies", [])
    if supplies:
        for s in supplies:
            pct  = s.get("pct", -1)
            desc = s.get("desc", "")
            icon = _supply_emoji(desc)
            bar  = _ink_bar(pct)
            lines.append(f"     {icon} <b>{desc}:</b>  {bar}")
    else:
        lines.append("     <i>Уровень чернил: нет данных</i>")
    return "\n".join(lines)


async def _send_ink_alerts(printers_db: dict):
    """Send Telegram alert for low ink supplies (< INK_WARN_PCT). Deduplicates by 24h."""
    aid = get_admin_id()
    if not aid:
        return
    now_ts = datetime.now(timezone.utc).timestamp()
    alerts  = _load_ink_alerts()
    changed = False
    msgs    = []
    for dev_name, info in printers_db.items():
        for p in info.get("printers", []):
            if p.get("is_virtual"):
                continue
            for s in p.get("supplies", []):
                pct = s.get("pct", -1)
                if not (0 <= pct < INK_WARN_PCT):
                    continue
                key      = f"{dev_name}::{p['name']}::{s['desc']}"
                last_sent = alerts.get(key, 0)
                if now_ts - last_sent < 86400:   # 24 h cooldown
                    continue
                alerts[key] = now_ts
                changed = True
                bar = _ink_bar(pct)
                msgs.append(
                    f"🖥 <b>{dev_name}</b> — {p['name']}\n"
                    f"   {_supply_emoji(s['desc'])} <b>{s['desc']}:</b> {bar}"
                )
    if changed:
        _save_ink_alerts(alerts)
    if msgs:
        header = f"🔴 <b>Мало чернил!</b> (менее {INK_WARN_PCT}%)\n\n"
        await bot.send_message(aid, header + "\n\n".join(msgs), parse_mode="HTML")


@router.callback_query(F.data == "tool:printers")
async def cb_tool_printers(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    # Show saved results + option to scan
    printers_db = _load_printers()
    devs = await get_full_devices()
    groups = sorted(set(d.get("group", "") for d in devs if d.get("group")))

    lines = ["━━━━━━━━━━━━━━━━━━━━━━\n🖨 <b>Принтеры</b>\n━━━━━━━━━━━━━━━━━━━━━━\n"]
    if printers_db:
        lines.append("<b>Последнее сканирование:</b>")
        for dev_name, info in sorted(printers_db.items()):
            scanned = info.get("scanned_at", "")[:16].replace("T", " ")
            prlist = info.get("printers", [])
            real    = [p for p in prlist if not p.get("is_virtual")]
            virtual = [p for p in prlist if p.get("is_virtual")]
            if not real:
                vnames = ", ".join(p.get("name", "?")[:28] for p in virtual[:4])
                more   = f" +{len(virtual)-4}" if len(virtual) > 4 else ""
                lines.append(f"\n🖥 <b>{dev_name}</b>  <i>({scanned})</i>")
                lines.append(f"  ⚪ только виртуальные: <i>{vnames}{more}</i>")
            else:
                lines.append(f"\n🖥 <b>{dev_name}</b>  <i>({scanned})</i>")
                for p in real:
                    lines.append(_format_printer_card(p, detailed=False))
                if virtual:
                    vnames = ", ".join(p.get("name", "?")[:22] for p in virtual[:3])
                    more   = f" +{len(virtual)-3}" if len(virtual) > 3 else ""
                    lines.append(f"     <i>⚪ скрыто: {vnames}{more}</i>")
    else:
        lines.append("Сканирований ещё не было.\n\nНажми <b>Сканировать офис</b> — "
                     "бот запустит <code>Get-Printer</code> на всех онлайн-ПК группы.")

    rows = []
    for g in groups:
        rows.append([InlineKeyboardButton(text=f"🔍 Сканировать: {g}", callback_data=f"prn:scan:{g[:45]}")])
    if printers_db:
        rows.append([InlineKeyboardButton(text="🗑 Очистить результаты", callback_data="prn:clear")])
    await cb.message.answer(
        "\n".join(lines),
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=rows) if rows else None,
    )


@router.callback_query(F.data.startswith("prn:scan:"))
async def cb_prn_scan(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    group_name = cb.data[len("prn:scan:"):]

    devs = await get_full_devices()
    group_devs = [d for d in devs if d.get("group", "").lower() == group_name.lower() and d["online"]]
    if not group_devs:
        await cb.message.answer(f"❌ Нет онлайн устройств в группе «{group_name}».")
        return

    wait_msg = await cb.message.answer(
        f"⏳ Сканирую принтеры на <b>{len(group_devs)}</b> ПК в <b>{group_name}</b>...\n"
        f"<i>(Get-Printer на каждом устройстве)</i>",
        parse_mode="HTML",
    )

    sem = asyncio.Semaphore(5)
    results: dict[str, list] = {}

    async def scan_one(d):
        async with sem:
            raw = await mc_run_command(d["id"], _get_printer_scan_cmd(), powershell=True)
            printers = []
            try:
                data = json.loads(raw)
                if isinstance(data, dict):
                    # PS5 wraps arrays as {"value": [...], "Count": N} — unwrap it
                    if "value" in data and isinstance(data.get("value"), list):
                        data = data["value"]
                    else:
                        data = [data]
                if not isinstance(data, list):
                    data = []
                for p in data:
                    pname = (p.get("Name") or "").strip()
                    # Skip empty names and garbled names (OEM CP437 artifacts U+0080–U+00FF)
                    if not pname or any('\u0080' <= c <= '\u00FF' for c in pname):
                        continue

                    # Python-side virtual detection (fallback when PS1 couldn't detect)
                    _VIRT_KW = ("anydesk", "pdf", "xps", "microsoft", "onenote",
                                "fax", "cutepdf", "adobe", "bullzip", "nitro",
                                "biztalk", "generic / text only", "send to onenote",
                                "pdfcreator", "doro", "docuworks")
                    drv_low = p.get("DriverName", "").lower()
                    is_virtual = bool(p.get("IsVirtual", False)) or any(kw in drv_low for kw in _VIRT_KW)

                    # Parse supplies from SNMP result
                    raw_supplies = p.get("Supplies") or []
                    supplies = []
                    for s in raw_supplies:
                        if not isinstance(s, dict):
                            continue
                        pct = s.get("pct", -1)
                        if pct is None:
                            pct = -1
                        supplies.append({
                            "desc": str(s.get("desc", "")),
                            "cur":  int(s.get("cur", 0)),
                            "max":  int(s.get("max", 0)),
                            "pct":  int(pct),
                        })
                    printers.append({
                        "name":       pname,
                        "driver":     p.get("DriverName", ""),
                        "port":       p.get("PortName", ""),
                        "status":     p.get("PrinterStatus", 0),
                        "shared":     bool(p.get("Shared")),
                        "default":    bool(p.get("Default")),
                        "is_virtual": is_virtual,
                        "printer_ip": p.get("PrinterIP", ""),
                        "supplies":   supplies,
                    })
            except Exception:
                pass  # device returned error / no printers
            results[d["name"]] = printers

    await asyncio.gather(*[scan_one(d) for d in group_devs])

    # Save to db
    printers_db = _load_printers()
    now = datetime.now(timezone.utc).isoformat()
    for dev_name, prlist in results.items():
        printers_db[dev_name] = {
            "group": group_name,
            "scanned_at": now,
            "printers": prlist,
        }
    _save_printers(printers_db)

    # Send low-ink alerts
    asyncio.create_task(_send_ink_alerts(printers_db))

    # Build report
    real_total = sum(1 for prlist in results.values() for p in prlist if not p.get("is_virtual"))
    lines = [
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🖨 <b>Принтеры — {group_name}</b>\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"Устройств проверено: {len(results)}  |  Принтеров найдено: <b>{real_total}</b>\n"
    ]
    for dev_name, prlist in sorted(results.items()):
        real = [p for p in prlist if not p.get("is_virtual")]
        virtual = [p for p in prlist if p.get("is_virtual")]
        if not prlist:
            lines.append(f"\n🖥 <b>{dev_name}</b>  — нет принтеров")
        elif not real:
            vnames = ", ".join(p.get("name", "?")[:28] for p in virtual[:4])
            more   = f" +{len(virtual)-4}" if len(virtual) > 4 else ""
            lines.append(f"\n🖥 <b>{dev_name}</b>  ⚪ только виртуальные:")
            lines.append(f"     <i>{vnames}{more}</i>")
        else:
            lines.append(f"\n🖥 <b>{dev_name}</b>")
            for p in real:
                lines.append(_format_printer_card(p, detailed=True))
            if virtual:
                vnames = ", ".join(p.get("name", "?")[:22] for p in virtual[:3])
                more   = f" +{len(virtual)-3}" if len(virtual) > 3 else ""
                lines.append(f"     <i>⚪ скрыто: {vnames}{more}</i>")

    text = "\n".join(lines)
    if len(text) > 4000:
        text = text[:4000] + "\n..."
    await wait_msg.edit_text(text, parse_mode="HTML")


@router.callback_query(F.data == "prn:clear")
async def cb_prn_clear(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    _save_printers({})
    await cb.answer("Результаты очищены")
    await cb.message.delete()


# ─── Mute Add FSM ────────────────────────────────────────────────────

@router.callback_query(F.data == "mute:add")
async def cb_mute_add(cb: CallbackQuery, state: FSMContext):
    """Start adding a mute: show device/group picker."""
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    wait = await cb.message.answer("⏳ Загружаю список...")
    devs = await get_full_devices()
    await wait.delete()
    await state.set_state(MuteSetup.picking_target)

    groups = sorted(set(d.get("group", "") for d in devs if d.get("group")))
    online = sorted([d for d in devs if d["online"]], key=lambda x: (x.get("group", ""), x["name"]))

    rows = []
    # All devices button
    rows.append([InlineKeyboardButton(text="🌐 Все устройства", callback_data="mute:tgt:__all__")])
    # Group buttons
    for g in groups:
        rows.append([InlineKeyboardButton(text=f"📁 {g}", callback_data=f"mute:tgt:{g[:50]}")])
    # Separator + online devices
    if online:
        rows.append([InlineKeyboardButton(text="── Устройства ──", callback_data="noop")])
        for d in online:
            rows.append([InlineKeyboardButton(
                text=f"🟢 {d['name']}",
                callback_data=f"mute:tgt:{d['name'][:50]}",
            )])
    rows.append([InlineKeyboardButton(text="❌ Отмена", callback_data="mute:cancel")])
    await cb.message.answer(
        "🔇 <b>Новый мьют — шаг 1/2</b>\n\nВыбери устройство или группу:",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=rows),
    )


@router.callback_query(F.data.startswith("mute:tgt:"), MuteSetup.picking_target)
async def cb_mute_pick_target(cb: CallbackQuery, state: FSMContext):
    target = cb.data[len("mute:tgt:"):]
    await state.update_data(mute_target=target)
    await state.set_state(MuteSetup.picking_duration)
    await cb.answer()

    label = {"__all__": "🌐 Все устройства"}.get(target, f"<b>{target}</b>")
    durations = [
        ("30 мин", "30m"), ("1 час", "1h"), ("2 часа", "2h"),
        ("4 часа", "4h"), ("8 часов", "8h"), ("24 часа", "24h"),
        ("3 дня", "3d"), ("7 дней", "7d"),
    ]
    rows = []
    for label_d, val in durations:
        rows.append([InlineKeyboardButton(text=label_d, callback_data=f"mute:dur:{val}")])
    rows.append([InlineKeyboardButton(text="❌ Отмена", callback_data="mute:cancel")])

    await cb.message.answer(
        f"🔇 <b>Новый мьют — шаг 2/2</b>\n\n"
        f"Цель: {label}\n\n"
        "Выбери длительность (или напиши вручную, например <code>3h</code>, <code>2d</code>):",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=rows),
    )


@router.callback_query(F.data.startswith("mute:dur:"), MuteSetup.picking_duration)
async def cb_mute_pick_dur(cb: CallbackQuery, state: FSMContext):
    dur_str = cb.data[len("mute:dur:"):]
    data = await state.get_data()
    target = data.get("mute_target", "")
    await state.clear()
    await cb.answer()

    seconds = parse_duration(dur_str)
    if not seconds:
        await cb.message.answer("❌ Ошибка длительности.", reply_markup=MAIN_KB)
        return

    mutes = load_mutes()
    until = time.time() + seconds
    mutes[target] = {"until": until, "reason": "", "muted_at": datetime.now(timezone.utc).isoformat()}
    save_mutes(mutes)

    until_str = datetime.fromtimestamp(until, tz=timezone.utc).strftime("%d.%m %H:%M UTC")
    disp = "Все устройства" if target == "__all__" else target
    await cb.message.answer(
        f"🔇 <b>Мьют добавлен</b>\n\n"
        f"• <b>{disp}</b> до {until_str}",
        parse_mode="HTML",
    )


@router.message(MuteSetup.picking_duration)
async def fsm_mute_duration_text(msg: Message, state: FSMContext):
    """User typed a custom duration like 3h."""
    if not is_admin(msg.from_user.id):
        return
    data = await state.get_data()
    target = data.get("mute_target", "")
    dur_str = msg.text.strip()
    seconds = parse_duration(dur_str)
    if not seconds:
        await msg.answer("❌ Неверный формат. Примеры: 30m, 2h, 3d")
        return
    await state.clear()

    mutes = load_mutes()
    until = time.time() + seconds
    mutes[target] = {"until": until, "reason": "", "muted_at": datetime.now(timezone.utc).isoformat()}
    save_mutes(mutes)

    until_str = datetime.fromtimestamp(until, tz=timezone.utc).strftime("%d.%m %H:%M UTC")
    disp = "Все устройства" if target == "__all__" else target
    await msg.answer(
        f"🔇 <b>Мьют добавлен</b>\n\n"
        f"• <b>{disp}</b> до {until_str}",
        parse_mode="HTML",
        reply_markup=MAIN_KB,
    )


@router.callback_query(F.data == "mute:cancel")
async def cb_mute_cancel(cb: CallbackQuery, state: FSMContext):
    await state.clear()
    await cb.answer("Отменено")
    await cb.message.delete()


# ─── Help system for alerts ─────────────────────────────────────────

@router.callback_query(F.data.startswith("help:"))
async def cb_help(cb: CallbackQuery):
    alert_type = cb.data.split(":", 1)[1]
    info = ALERT_HELP.get(alert_type)
    if not info:
        await cb.answer("Нет справки", show_alert=True)
        return
    await cb.answer()
    cfg = load_alerts_cfg()
    text = (
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"{info['title']}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"<b>Что это:</b>\n{info['what']}\n\n"
        f"<b>Когда срабатывает:</b>\n{info['when'].format(interval=DEVICE_CHECK_INTERVAL, threshold=cfg.get('disk_pct', 90), offline_hours=cfg.get('offline_hours', 24))}\n\n"
        f"<b>Настройка:</b>\n{info['config'].format(threshold=cfg.get('disk_pct', 90), offline_hours=cfg.get('offline_hours', 24))}\n\n"
        f"<b>Что делать:</b>\n{info['action']}\n\n"
        f"<b>Пример:</b>\n<i>{info['example']}</i>"
    )
    await cb.message.answer(text, parse_mode="HTML")


# ─── MC Auto-Update callbacks ───────────────────────────────────────

@router.callback_query(F.data == "mc:update")
async def cb_mc_update(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    await cb.message.answer(
        "⚠️ <b>Обновление MeshCentral</b>\n\n"
        "Будет выполнено:\n"
        "1. Бэкап config.json\n"
        "2. npm update meshcentral\n"
        "3. Перезапуск MeshCentral\n"
        "4. Проверка работоспособности\n\n"
        "⏱ Время: ~2-3 минуты. MC будет недоступен.",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ Обновить сейчас", callback_data="mc:update_go")],
            [InlineKeyboardButton(text="❌ Отмена", callback_data="noop")],
        ]),
    )


@router.callback_query(F.data == "mc:update_go")
async def cb_mc_update_go(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer("Запускаю обновление...")
    aid = cb.from_user.id
    asyncio.create_task(perform_mc_update(aid))


# ─── Background tasks ───────────────────────────────────────────────

async def health_loop():
    global _mc_was_down, _http_down
    await asyncio.sleep(15)
    while not _shutdown_event.is_set():
        try:
            aid = get_admin_id()
            alive = await mc_is_alive()
            if not alive and not _mc_was_down:
                _mc_was_down = True
                await mc_restart()
                if aid:
                    try:
                        await bot.send_message(aid, "🔴 <b>MeshCentral упал!</b> Перезапуск...", parse_mode="HTML")
                    except Exception:
                        pass
                await asyncio.sleep(20)
                if await mc_is_alive():
                    _mc_was_down = False
                    if aid:
                        try:
                            await bot.send_message(aid, "🟢 MeshCentral восстановлен.", parse_mode="HTML")
                        except Exception:
                            pass
            elif alive and _mc_was_down:
                _mc_was_down = False
                if aid:
                    try:
                        await bot.send_message(aid, "🟢 MeshCentral работает.", parse_mode="HTML")
                    except Exception:
                        pass

            # ── HTTP services healthcheck ──
            if aid:
                http_results = await check_all_http_services()
                for r in http_results:
                    name = r["name"]
                    ok   = r["ok"]
                    was_down = _http_down.get(name, False)
                    if not ok and not was_down:
                        _http_down[name] = True
                        status_str = f" (HTTP {r['status']})" if r["status"] else " (недоступен)"
                        try:
                            await bot.send_message(
                                aid,
                                f"🔴 <b>{name}</b> недоступен!{status_str}\n"
                                f"<code>{r['url']}</code>",
                                parse_mode="HTML",
                            )
                        except Exception:
                            pass
                    elif ok and was_down:
                        _http_down[name] = False
                        try:
                            await bot.send_message(
                                aid,
                                f"🟢 <b>{name}</b> восстановлен.",
                                parse_mode="HTML",
                            )
                        except Exception:
                            pass
                    else:
                        _http_down[name] = not ok
        except Exception as e:
            log.error(f"Health: {e}")
        try:
            await asyncio.wait_for(_shutdown_event.wait(), timeout=HEALTH_CHECK_INTERVAL)
            break
        except asyncio.TimeoutError:
            pass


async def device_loop():
    global _known_devices
    await asyncio.sleep(25)
    try:
        for d in await get_full_devices():
            _known_devices[d["id"]] = {"name": d["name"], "online": d["online"]}
    except Exception:
        pass

    while not _shutdown_event.is_set():
        try:
            aid = get_admin_id()
            if not aid:
                await asyncio.sleep(DEVICE_CHECK_INTERVAL)
                continue

            devs = await get_full_devices()
            cfg = load_alerts_cfg()
            cur = {d["id"]: d for d in devs}

            for did, d in cur.items():
                prev = _known_devices.get(did)
                if is_muted(d["name"], d.get("group", "")):
                    continue
                if prev is None:
                    if cfg.get("new_device", True):
                        try:
                            await bot.send_message(
                                aid,
                                f"🆕 <b>Новое устройство:</b> {d['name']}\n💻 {d['os']}\n🌐 {d['ip']}",
                                parse_mode="HTML",
                                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                                    [InlineKeyboardButton(text="❓ Что делать?", callback_data="help:new_device")],
                                ]),
                            )
                        except Exception:
                            pass
                elif d["online"] and not prev["online"]:
                    try:
                        await bot.send_message(
                            aid,
                            f"🟢 <b>{d['name']}</b> подключился\n   <code>{d['ip']}</code>",
                            parse_mode="HTML",
                        )
                    except Exception:
                        pass
                elif not d["online"] and prev["online"]:
                    try:
                        await bot.send_message(
                            aid, f"⚪ <b>{d['name']}</b> отключился", parse_mode="HTML",
                        )
                    except Exception:
                        pass

            # ─ Condition alerts ─
            alerts_sent = _load_json(DATA_DIR / "alerts_sent.json", {})
            today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

            for d in devs:
                if is_muted(d["name"], d.get("group", "")):
                    continue
                alert_key = f"{d['name']}_{today}"

                # Disk alert
                if d.get("vol_alerts") and cfg.get("disk_pct"):
                    dk = f"disk_{alert_key}"
                    if dk not in alerts_sent:
                        alerts_sent[dk] = True
                        try:
                            await bot.send_message(
                                aid,
                                f"💿 <b>Диск заполнен:</b> {d['name']}\n{', '.join(d['vol_alerts'])}",
                                parse_mode="HTML",
                                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                                    [InlineKeyboardButton(text="❓ Что делать?", callback_data="help:disk")],
                                ]),
                            )
                        except Exception:
                            pass

                # AV disabled alert
                if d.get("av_disabled") and cfg.get("av_off"):
                    ak = f"av_{alert_key}"
                    if ak not in alerts_sent:
                        alerts_sent[ak] = True
                        try:
                            await bot.send_message(
                                aid,
                                f"🛡 <b>Антивирус выключен:</b> {d['name']}\n{d['antivirus']}",
                                parse_mode="HTML",
                                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                                    [InlineKeyboardButton(text="❓ Что делать?", callback_data="help:av")],
                                ]),
                            )
                        except Exception:
                            pass

                # Long offline alert
                if d.get("offline_hours", 0) >= cfg.get("offline_hours", 24):
                    ok = f"offline_{alert_key}"
                    if ok not in alerts_sent:
                        alerts_sent[ok] = True
                        try:
                            await bot.send_message(
                                aid,
                                f"⏰ <b>Долго офлайн:</b> {d['name']} ({fmt_offline(d['offline_hours'])})",
                                parse_mode="HTML",
                                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                                    [InlineKeyboardButton(text="❓ Что делать?", callback_data="help:offline")],
                                ]),
                            )
                        except Exception:
                            pass

            # cleanup old alerts (keep only today)
            alerts_sent = {k: v for k, v in alerts_sent.items() if today in k}
            _save_json(DATA_DIR / "alerts_sent.json", alerts_sent)

            # record uptime
            record_uptime(devs)

            _known_devices = {did: {"name": d["name"], "online": d["online"]} for did, d in cur.items()}
        except Exception as e:
            log.error(f"DevMon: {e}")
        try:
            await asyncio.wait_for(_shutdown_event.wait(), timeout=DEVICE_CHECK_INTERVAL)
            break
        except asyncio.TimeoutError:
            pass


async def scheduled_loop():
    global _last_inventory_date, _last_daily_report, _last_weekly_digest, _last_update_check
    await asyncio.sleep(30)
    while not _shutdown_event.is_set():
        try:
            aid = get_admin_id()
            now = datetime.now(timezone.utc)
            today = now.strftime("%Y-%m-%d")

            if now.hour == INVENTORY_HOUR and _last_inventory_date != today and aid:
                _last_inventory_date = today
                devs = await get_full_devices()
                if devs:
                    # save snapshot for change tracking
                    save_snapshot(devs)
                    save_snap_history(devs)
                    save_disk_snapshot(devs)
                    try:
                        await bot.send_document(
                            aid,
                            BufferedInputFile(build_inventory_csv(devs), filename=f"inventory_{today}.csv"),
                            caption=f"📦 <b>Авто-инвентарь</b> {today} • {len(devs)} устройств",
                            parse_mode="HTML",
                        )
                    except Exception:
                        pass

            if now.hour == DAILY_REPORT_HOUR and _last_daily_report != today and aid:
                _last_daily_report = today
                devs = await get_full_devices()
                online = sum(1 for d in devs if d["online"])
                cpu = psutil.cpu_percent(interval=0.5)
                mem = psutil.virtual_memory()

                # detect changes
                changes = detect_changes(devs)
                changes_str = ""
                if changes:
                    changes_str = "\n\n📜 <b>Изменения:</b>\n" + "\n".join(changes[:10])

                # device alerts summary
                alert_lines = []
                cfg = load_alerts_cfg()
                for d in devs:
                    if d.get("vol_alerts"):
                        alert_lines.append(f"  💿 {d['name']}: {', '.join(d['vol_alerts'])}")
                    if d.get("av_disabled"):
                        alert_lines.append(f"  🛡 {d['name']}: AV выключен")
                    if d.get("offline_hours", 0) >= cfg.get("offline_hours", 24):
                        alert_lines.append(f"  ⏰ {d['name']}: офлайн {fmt_offline(d['offline_hours'])}")
                alerts_str = ""
                if alert_lines:
                    alerts_str = "\n\n⚠️ <b>Проблемы:</b>\n" + "\n".join(alert_lines[:10])

                # SSL cert status in daily report
                ssl_str = ""
                if _ssl_cache:
                    ssl_warn = [r for r in _ssl_cache if not r["ok"] or r["days_left"] <= SSL_WARN_DAYS]
                    if ssl_warn:
                        ssl_str = "\n\n🔐 <b>SSL:</b>\n" + ssl_status_text(ssl_warn)

                try:
                    await bot.send_message(
                        aid,
                        f"━━━━━━━━━━━━━━━━━━━━━━\n📋 <b>Отчёт {today}</b>\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
                        f"📱 Устройств: {len(devs)} (🟢 {online})\n"
                        f"🧠 CPU: {cpu:.0f}% 💾 RAM: {mem.percent:.0f}%\n"
                        f"🛡 MC: {'🟢' if await mc_is_alive() else '🔴'}"
                        f"{changes_str}{alerts_str}{ssl_str}",
                        parse_mode="HTML",
                    )
                except Exception:
                    pass
                # save snapshot after report
                save_snapshot(devs)
                save_snap_history(devs)

            # ── Weekly digest (Sunday) ──
            if now.weekday() == 6 and now.hour == WEEKLY_DIGEST_HOUR and _last_weekly_digest != today and aid:
                _last_weekly_digest = today
                devs = await get_full_devices()
                if devs:
                    await _send_weekly_digest(aid, devs)

            # ── Daily update check ──
            if now.hour == UPDATE_CHECK_HOUR and _last_update_check != today and aid:
                _last_update_check = today
                info = await check_mc_update()
                if info["has_update"]:
                    try:
                        await bot.send_message(
                            aid,
                            f"\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n"
                            f"\U0001f195 <b>Обновление MeshCentral!</b>\n"
                            f"\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n\n"
                            f"\U0001f4e6 Текущая: <b>{info['current']}</b>\n"
                            f"\U0001f680 Новая: <b>{info['latest']}</b>",
                            parse_mode="HTML",
                            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                                [InlineKeyboardButton(text="🔄 Обновить сейчас", callback_data="mc:update")],
                            ]),
                        )
                    except Exception:
                        pass
        except Exception as e:
            log.error(f"Sched: {e}")
        try:
            await asyncio.wait_for(_shutdown_event.wait(), timeout=60)
            break
        except asyncio.TimeoutError:
            pass


async def ssl_check_loop():
    global _last_ssl_check, _ssl_cache
    await asyncio.sleep(60)  # дать боту запустится
    while not _shutdown_event.is_set():
        try:
            aid = get_admin_id()
            now = datetime.now(timezone.utc)
            today = now.strftime("%Y-%m-%d")
            if (now.hour == SSL_CHECK_HOUR and _last_ssl_check != today) or not _ssl_cache:
                _last_ssl_check = today
                results = await check_all_ssl()
                _ssl_cache = results
                # отправить алерт только если есть проблемы
                problems = [r for r in results if not r["ok"] or r["days_left"] <= SSL_WARN_DAYS]
                if problems and aid:
                    lines = ssl_status_text(problems)
                    crit = any(not r["ok"] or r["days_left"] <= SSL_CRIT_DAYS for r in problems)
                    header = "🔴 <b>SSL КРИТИЧНО</b>" if crit else "🟡 <b>SSL предупреждение</b>"
                    try:
                        await bot.send_message(
                            aid,
                            f"━━━━━━━━━━━━━━━━━━━━━━\n{header}\n━━━━━━━━━━━━━━━━━━━━━━\n\n{lines}\n\n"
                            f"🔐 /certs — проверить все сертификаты",
                            parse_mode="HTML",
                        )
                    except Exception:
                        pass
        except Exception as e:
            log.error(f"ssl_check_loop: {e}")
        try:
            await asyncio.wait_for(_shutdown_event.wait(), timeout=3600)
            break
        except asyncio.TimeoutError:
            pass


async def on_startup():
    _background_tasks.append(asyncio.create_task(health_loop()))
    _background_tasks.append(asyncio.create_task(device_loop()))
    _background_tasks.append(asyncio.create_task(scheduled_loop()))
    _background_tasks.append(asyncio.create_task(wifi_poll_loop()))
    _background_tasks.append(asyncio.create_task(netmap_loop()))
    _background_tasks.append(asyncio.create_task(ssl_check_loop()))
    _background_tasks.append(asyncio.create_task(cmd_scheduler_loop()))
    _background_tasks.append(asyncio.create_task(snmp_poll_loop()))
    _background_tasks.append(asyncio.create_task(hw_inventory_loop()))
    _background_tasks.append(asyncio.create_task(temp_loop()))
    log.info("Background tasks started")


async def shutdown():
    log.info("Shutting down gracefully...")
    _shutdown_event.set()
    for t in _background_tasks:
        t.cancel()
    await asyncio.gather(*_background_tasks, return_exceptions=True)
    await bot.session.close()
    log.info("Shutdown complete.")


# ─── Deploy guide ────────────────────────────────────────────────────

GITHUB_REPO = "https://github.com/mr-khamzat/mc-stack"

@router.callback_query(F.data == "tool:deploy")
async def cb_tool_deploy(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()

    text = (
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "🚀 <b>Развернуть копию стека</b>\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "Стек включает:\n"
        "• 🤖 Telegram бот (этот)\n"
        "• 🗄 RackViz — визуализатор стойки\n"
        "• 🗺 NetMap — карта сети\n"
        "• 📡 WiFi зонды (Keenetic)\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "📋 <b>Что нужно для установки:</b>\n\n"
        "1. 🖥 Сервер с Ubuntu 22/24 LTS\n"
        "   (2 CPU, 2 GB RAM, 20 GB disk)\n\n"
        "2. 🤖 Telegram Bot Token\n"
        "   → @BotFather → /newbot\n\n"
        "3. 👤 Твой Telegram Chat ID\n"
        "   → написать @userinfobot\n\n"
        "4. 🔐 Данные MeshCentral сервера\n"
        "   URL, логин, пароль\n\n"
        "5. 🔑 MC Login Token Key\n"
        "   (с MC сервера: <code>node meshcentral.js --logintokenkey</code>)\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "⚡ <b>Установка (3 команды):</b>\n\n"
        f"<code>git clone {GITHUB_REPO}</code>\n"
        "<code>cd mc-stack</code>\n"
        "<code>sudo bash deploy.sh</code>\n\n"
        "Скрипт задаст вопросы и всё настроит сам.\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "📂 <b>GitHub репозиторий:</b>\n"
        f"{GITHUB_REPO}\n\n"
        "📖 <b>Подробная инструкция:</b>\n"
        f"{GITHUB_REPO}/blob/main/README.md"
    )

    buttons = [
        [InlineKeyboardButton(text="📂 GitHub репо", url=GITHUB_REPO)],
        [InlineKeyboardButton(text="📦 Упаковать с этого сервера", callback_data="tool:deploy_pack")],
    ]
    await cb.message.answer(text, parse_mode="HTML",
                            reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))


@router.callback_query(F.data == "tool:deploy_pack")
async def cb_tool_deploy_pack(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer("Упаковываю...", show_alert=False)

    pack_script = Path("/opt/deploy-kit/pack.sh")
    if not pack_script.exists():
        await cb.message.answer("❌ Скрипт /opt/deploy-kit/pack.sh не найден")
        return

    msg = await cb.message.answer("⏳ Упаковываю проект, это займёт 30-60 секунд...")
    try:
        proc = await asyncio.create_subprocess_exec(
            "bash", str(pack_script),
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT,
        )
        out, _ = await asyncio.wait_for(proc.communicate(), timeout=120)
        output = out.decode(errors="replace")

        # Найти путь к архиву
        archive = None
        for line in output.splitlines():
            if line.strip().endswith(".tar.gz") and "/tmp/" in line:
                for part in line.split():
                    if part.endswith(".tar.gz"):
                        archive = Path(part.strip())
                        break

        if archive and archive.exists():
            size = archive.stat().st_size / (1024 * 1024)
            await msg.edit_text(
                f"✅ Архив создан: <code>{archive.name}</code> ({size:.1f} MB)\n\n"
                "📤 Скопируй на новый сервер:\n"
                f"<code>scp root@{MC_URL.replace('https://', '').split(':')[0]}:{archive} /tmp/</code>\n\n"
                "Затем на новом сервере:\n"
                f"<code>cd /tmp && tar xzf {archive.name}</code>\n"
                f"<code>cd mc-stack-deploy/deploy-kit && bash deploy.sh</code>",
                parse_mode="HTML",
            )
        else:
            await msg.edit_text(f"⚠️ Архив не найден. Вывод:\n<code>{output[-500:]}</code>",
                                parse_mode="HTML")
    except asyncio.TimeoutError:
        await msg.edit_text("⏱ Таймаут упаковки (>120 сек)")
    except Exception as e:
        await msg.edit_text(f"❌ Ошибка: {e}")


@router.message(Command("certs"))
@router.callback_query(F.data == "tool:certs")
async def cmd_certs(event):
    msg = event if isinstance(event, Message) else event.message
    if not is_admin(event.from_user.id):
        if isinstance(event, CallbackQuery):
            await event.answer("🔒", show_alert=True)
        return
    if isinstance(event, CallbackQuery):
        await event.answer()
    wait = await msg.answer("⏳ Проверяю SSL сертификаты...")
    results = await check_all_ssl()
    global _ssl_cache
    _ssl_cache = results
    text = ssl_status_text(results)
    ok_count = sum(1 for r in results if r["ok"] and r["days_left"] > SSL_WARN_DAYS)
    warn_count = sum(1 for r in results if r["ok"] and SSL_CRIT_DAYS < r["days_left"] <= SSL_WARN_DAYS)
    crit_count = sum(1 for r in results if not r["ok"] or r["days_left"] <= SSL_CRIT_DAYS)
    summary = f"🟢 {ok_count}  🟡 {warn_count}  🔴 {crit_count}"
    await wait.edit_text(
        f"━━━━━━━━━━━━━━━━━━━━━━\n🔐 <b>SSL Сертификаты</b>\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"{text}\n\n{summary}",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔄 Обновить статус", callback_data="tool:certs"),
             InlineKeyboardButton(text="🔁 Продлить certbot",  callback_data="tool:ssl_renew")],
        ]),
    )


# ─── SSL Renew ───────────────────────────────────────────────────────────────

@router.callback_query(F.data == "tool:ssl_renew")
async def cb_ssl_renew(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True); return
    await cb.answer()
    wait = await cb.message.answer("⏳ Запускаю <code>certbot renew</code>…\n<i>Это может занять 30–60 секунд.</i>",
                                   parse_mode="HTML")
    try:
        proc = await asyncio.create_subprocess_exec(
            "certbot", "renew", "--non-interactive", "--quiet",
            stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE)
        out, err = await asyncio.wait_for(proc.communicate(), timeout=120)
        rc   = proc.returncode
        text = (out or b"").decode(errors="replace").strip()
        errt = (err or b"").decode(errors="replace").strip()
        combined = (text + "\n" + errt).strip() or "(certbot не вывел ничего — сертификаты актуальны)"
        icon = "✅" if rc == 0 else "⚠️"
        msg_text = (f"{icon} <b>certbot renew</b>  (exit {rc})\n\n"
                    f"<pre>{combined[:3000]}</pre>")
    except asyncio.TimeoutError:
        msg_text = "⏱ Timeout: certbot не ответил за 120 секунд."
    except Exception as e:
        msg_text = f"❌ Ошибка запуска certbot: {e}"

    await wait.edit_text(msg_text, parse_mode="HTML",
                         reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                             [InlineKeyboardButton(text="🔐 Проверить SSL", callback_data="tool:certs")],
                             [InlineKeyboardButton(text="🔁 Повторить",     callback_data="tool:ssl_renew")],
                         ]))


# ─── Ping / Traceroute ───────────────────────────────────────────────────────

def _ping_target_kb(ip: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🏓 Ping (4 пакета)",  callback_data=f"ping:run:ping:{ip}"),
         InlineKeyboardButton(text="🌐 Traceroute",       callback_data=f"ping:run:trace:{ip}")],
        [InlineKeyboardButton(text="📊 MTR (10 циклов)",  callback_data=f"ping:run:mtr:{ip}")],
        [InlineKeyboardButton(text="◀️ Выбрать другой",   callback_data="tool:ping")],
    ])


@router.callback_query(F.data == "tool:ping")
async def cb_tool_ping(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True); return
    await cb.answer()
    await state.clear()
    devs = await get_full_devices()
    # Collect unique WAN IPs per group
    seen: dict[str, str] = {}   # ip → group
    for d in sorted(devs, key=lambda x: x.get("group", "")):
        wan = (d.get("ip") or "").strip()
        grp = d.get("group", "?")
        if wan and wan not in seen:
            seen[wan] = grp
    rows = []
    for ip, grp in seen.items():
        rows.append([InlineKeyboardButton(
            text=f"🏢 {grp}  —  {ip}", callback_data=f"ping:ip:{ip}")])
    rows.append([InlineKeyboardButton(text="✏️ Ввести IP / hostname вручную",
                                      callback_data="ping:manual")])
    rows.append([InlineKeyboardButton(text="◀️ Инструменты", callback_data="ping:back")])
    await cb.message.answer(
        "━━━━━━━━━━━━━━━━━━━━━━\n🏓 <b>Ping / Traceroute</b>\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "Пинг выполняется <b>с сервера</b> (144.31.89.167).\n"
        "Выбери цель или введи адрес вручную:",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=rows),
    )


@router.callback_query(F.data.startswith("ping:ip:"))
async def cb_ping_ip(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True); return
    await cb.answer()
    ip = cb.data[len("ping:ip:"):]
    await cb.message.answer(
        f"🎯 Цель: <code>{ip}</code>\n\nВыбери действие:",
        parse_mode="HTML", reply_markup=_ping_target_kb(ip))


@router.callback_query(F.data == "ping:manual")
async def cb_ping_manual(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True); return
    await cb.answer()
    await state.set_state(PingState.waiting_ip)
    await cb.message.answer("✏️ Введи IP-адрес или hostname:")


@router.message(PingState.waiting_ip)
async def ping_manual_ip(msg: Message, state: FSMContext):
    if not is_admin(msg.from_user.id): return
    ip = msg.text.strip()
    await state.clear()
    await msg.answer(
        f"🎯 Цель: <code>{ip}</code>\n\nВыбери действие:",
        parse_mode="HTML", reply_markup=_ping_target_kb(ip))


@router.callback_query(F.data.startswith("ping:run:"))
async def cb_ping_run(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True); return
    await cb.answer()
    # Format: ping:run:TYPE:IP  (IP may contain colons for IPv6)
    parts = cb.data.split(":", 3)
    action, ip = parts[2], parts[3]

    icons = {"ping": "🏓", "trace": "🌐", "mtr": "📊"}
    wait = await cb.message.answer(
        f"{icons.get(action,'🔍')} <b>{action}</b> → <code>{ip}</code> …",
        parse_mode="HTML")

    if action == "ping":
        cmd = ["ping", "-c", "4", "-W", "3", ip]; timeout = 20
    elif action == "trace":
        cmd = ["traceroute", "-n", "-m", "20", "-w", "2", ip]; timeout = 60
    else:  # mtr
        cmd = ["mtr", "--report", "--report-cycles", "10", "-n", ip]; timeout = 90

    try:
        proc = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE)
        out, err = await asyncio.wait_for(proc.communicate(), timeout=timeout)
        result = (out or b"").decode(errors="replace").strip() or \
                 (err or b"").decode(errors="replace").strip() or "(нет вывода)"
    except asyncio.TimeoutError:
        result = f"⏱ Timeout ({timeout}s)"
    except Exception as e:
        result = f"❌ {e}"

    # Parse ping summary line if present
    summary = ""
    for line in result.splitlines():
        if "packet loss" in line or "packets transmitted" in line:
            summary = f"\n<i>{line.strip()}</i>"
            break

    text = (f"{icons.get(action,'🔍')} <b>{action}</b> → <code>{ip}</code>{summary}\n\n"
            f"<pre>{result[:3400]}</pre>")
    rows = [
        [InlineKeyboardButton(text="🔄 Повторить", callback_data=cb.data)],
    ]
    if action == "ping":
        rows.append([InlineKeyboardButton(text="🌐 Traceroute", callback_data=f"ping:run:trace:{ip}"),
                     InlineKeyboardButton(text="📊 MTR",        callback_data=f"ping:run:mtr:{ip}")])
    rows.append([InlineKeyboardButton(text="◀️ Выбрать другой", callback_data="tool:ping")])
    await wait.edit_text(text, parse_mode="HTML",
                         reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))


@router.callback_query(F.data == "ping:back")
async def cb_ping_back(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True); return
    await cb.answer()
    # Re-open tools menu
    await cb.message.delete()


# ─── Notes ───────────────────────────────────────────────────────────────────

def _load_notes() -> dict:
    try:
        if NOTES_FILE.exists():
            return json.loads(NOTES_FILE.read_text())
    except Exception:
        pass
    return {}


def _save_notes(notes: dict) -> None:
    NOTES_FILE.write_text(json.dumps(notes, ensure_ascii=False, indent=2))


def _notes_device_kb(device_name: str, has_note: bool) -> InlineKeyboardMarkup:
    rows = [
        [InlineKeyboardButton(text="✏️ Написать заметку",  callback_data=f"note:edit:{device_name[:50]}")],
    ]
    if has_note:
        rows.append([InlineKeyboardButton(text="🗑 Удалить заметку", callback_data=f"note:del:{device_name[:50]}")])
    rows.append([InlineKeyboardButton(text="◀️ К заметкам", callback_data="tool:notes")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


@router.callback_query(F.data == "tool:notes")
async def cb_tool_notes(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True); return
    await cb.answer()
    await state.clear()
    notes = _load_notes()
    devs  = await get_full_devices()
    # Sort: devices with notes first, then alphabetically
    rows = []
    with_note    = sorted([d for d in devs if d["name"] in notes], key=lambda d: d["name"])
    without_note = sorted([d for d in devs if d["name"] not in notes], key=lambda d: d["name"])
    for d in with_note + without_note:
        has = d["name"] in notes
        icon = "📝" if has else ("🟢" if d.get("online") else "⚫")
        rows.append([InlineKeyboardButton(
            text=f"{icon} {d['name'][:35]}",
            callback_data=f"note:dev:{d['name'][:50]}")])
    note_lines = [f"📝 <b>Заметки к устройствам</b>  ({len(notes)} записей)\n"]
    if notes:
        for name, note in sorted(notes.items()):
            note_lines.append(f"• <b>{name}</b>: {note[:80]}{'…' if len(note)>80 else ''}")
    await cb.message.answer(
        "━━━━━━━━━━━━━━━━━━━━━━\n" + "\n".join(note_lines) + "\n━━━━━━━━━━━━━━━━━━━━━━\n\nВыбери устройство:",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=rows),
    )


@router.callback_query(F.data.startswith("note:dev:"))
async def cb_note_device(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True); return
    await cb.answer()
    device_name = cb.data[len("note:dev:"):]
    notes = _load_notes()
    note  = notes.get(device_name, "")
    text  = (f"📝 <b>{device_name}</b>\n\n"
             f"Текущая заметка:\n<blockquote>{note}</blockquote>" if note
             else f"📝 <b>{device_name}</b>\n\nЗаметок нет.")
    await cb.message.answer(text, parse_mode="HTML",
                             reply_markup=_notes_device_kb(device_name, bool(note)))


@router.callback_query(F.data.startswith("note:edit:"))
async def cb_note_edit(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True); return
    await cb.answer()
    device_name = cb.data[len("note:edit:"):]
    await state.set_state(NotesState.writing_note)
    await state.update_data(device_name=device_name)
    notes = _load_notes()
    hint  = f"\nТекущая: <i>{notes[device_name][:80]}</i>" if device_name in notes else ""
    await cb.message.answer(
        f"✏️ Введи заметку для <b>{device_name}</b>:{hint}\n\n"
        "<i>(отправь '-' чтобы отменить)</i>",
        parse_mode="HTML")


@router.message(NotesState.writing_note)
async def notes_write(msg: Message, state: FSMContext):
    if not is_admin(msg.from_user.id): return
    data = await state.get_data()
    device_name = data.get("device_name", "")
    await state.clear()
    if msg.text.strip() == "-":
        await msg.answer("❌ Отменено.", reply_markup=MAIN_KB); return
    notes = _load_notes()
    notes[device_name] = msg.text.strip()[:500]
    _save_notes(notes)
    await msg.answer(f"✅ Заметка сохранена для <b>{device_name}</b>:\n<blockquote>{notes[device_name]}</blockquote>",
                     parse_mode="HTML", reply_markup=MAIN_KB)


@router.callback_query(F.data.startswith("note:del:"))
async def cb_note_del(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True); return
    device_name = cb.data[len("note:del:"):]
    notes = _load_notes()
    notes.pop(device_name, None)
    _save_notes(notes)
    await cb.answer(f"🗑 Заметка удалена", show_alert=False)
    await cb.message.answer(f"🗑 Заметка для <b>{device_name}</b> удалена.",
                             parse_mode="HTML", reply_markup=MAIN_KB)


# ─── Disk trend handler ───────────────────────────────────────────────

@router.callback_query(F.data == "tool:disk_trend")
async def cb_tool_disk_trend(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()

    trends = get_disk_trends()
    hist   = _load_json(DISK_HISTORY_FILE, {})
    n_devices = len(hist)
    n_points  = sum(len(v) for v in hist.values())

    header = (
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "📈 <b>Тренд заполнения дисков</b>\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"📊 История: {n_devices} устройств, {n_points} замеров\n"
    )

    if not trends:
        text = (
            header +
            "\n<i>Данных пока недостаточно. Тренд считается по данным\n"
            "ежедневного инвентаря (08:00 UTC). Нужно минимум 2 дня.</i>\n\n"
            "💡 Чтобы собрать данные сейчас:\n"
            "🔧 Инструменты → 📈 Тренд дисков → кнопка 🔄 Собрать сейчас"
        )
        rows = [
            [InlineKeyboardButton(text="🔄 Собрать сейчас", callback_data="disk_trend:collect")],
            [InlineKeyboardButton(text="◀️ Инструменты",    callback_data="disk_trend:back")],
        ]
        await cb.message.answer(text, parse_mode="HTML",
                                reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
        return

    lines = [header]
    crit  = [t for t in trends if t["days_to_full"] is not None and t["days_to_full"] <= 30]
    warn  = [t for t in trends if t["days_to_full"] is not None and 30 < t["days_to_full"] <= 90]
    ok    = [t for t in trends if t["days_to_full"] is None]

    def _trend_line(t: dict) -> str:
        d2f = t["days_to_full"]
        if d2f is not None:
            urgency = "🔴" if d2f <= 30 else "🟡"
            eta = f"⚠️ заполнится через ~{int(d2f)} д."
        else:
            urgency = "🟢"
            eta = "стабильно"
        rate = f"+{t['fill_rate_gb_day']:.2f}" if t['fill_rate_gb_day'] >= 0 else f"{t['fill_rate_gb_day']:.2f}"
        return (
            f"{urgency} <b>{t['device']}</b>  {t['letter']}:\n"
            f"   {t['used_gb']:.1f}/{t['total_gb']:.1f} ГБ ({t['used_pct']:.0f}%)  "
            f"{rate} ГБ/д  {eta}"
        )

    if crit:
        lines.append("🔴 <b>Критично (заполнится &lt;30 дней):</b>")
        for t in crit:
            lines.append(_trend_line(t))
        lines.append("")
    if warn:
        lines.append("🟡 <b>Внимание (30–90 дней):</b>")
        for t in warn[:5]:
            lines.append(_trend_line(t))
        if len(warn) > 5:
            lines.append(f"   <i>... и ещё {len(warn)-5}</i>")
        lines.append("")
    if ok:
        lines.append(f"🟢 <b>Стабильны:</b> {len(ok)} дисков")

    text = "\n".join(lines)
    if len(text) > 4000:
        text = text[:4000] + "\n<i>... (обрезано)</i>"

    rows = [
        [InlineKeyboardButton(text="🔄 Собрать сейчас", callback_data="disk_trend:collect")],
        [InlineKeyboardButton(text="◀️ Инструменты",    callback_data="disk_trend:back")],
    ]
    await cb.message.answer(text, parse_mode="HTML",
                            reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))


@router.callback_query(F.data == "disk_trend:collect")
async def cb_disk_trend_collect(cb: CallbackQuery):
    """Collect disk snapshot right now (don't wait for daily inventory)."""
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer("⏳ Собираю данные...")
    await cb.message.answer("⏳ Опрашиваю устройства...", parse_mode="HTML")
    devs = await get_full_devices()
    if devs:
        save_disk_snapshot(devs)
        n = sum(1 for d in devs if d.get("volumes_raw"))
        await cb.message.answer(f"✅ Сохранён снапшот дисков: {n} устройств с данными.\n"
                                 f"Нажми 📈 Тренд дисков снова для просмотра.",
                                 parse_mode="HTML")
    else:
        await cb.message.answer("⚠️ Не удалось получить данные устройств.", parse_mode="HTML")


@router.callback_query(F.data == "disk_trend:back")
async def cb_disk_trend_back(cb: CallbackQuery):
    await cb.answer()
    await cb.message.answer("🔧 Выберите инструмент:", reply_markup=MAIN_KB)


# ─── Snapshot 7-day compare ──────────────────────────────────────────

@router.callback_query(F.data == "tool:snap_compare")
async def cb_snap_compare(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    wait = await cb.message.answer("⏳ Получаю данные…")
    devs = await get_full_devices()
    text = compare_snap_history(devs, days=7)
    await wait.delete()
    for chunk in [text[i:i+3800] for i in range(0, len(text), 3800)]:
        await cb.message.answer(chunk, parse_mode="HTML")


# ─── Group WoL ───────────────────────────────────────────────────────

@router.callback_query(F.data.startswith("wol_grp:"))
async def cb_wol_group(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    group = cb.data.split(":", 1)[1]
    devs = [d for d in await get_full_devices() if d["group"] == group and not d["online"]]
    sent, skipped = [], []
    for d in devs:
        macs = [nic["mac"] for nic in d.get("nic_details", [])
                if nic.get("mac", "") not in ("", "00:00:00:00:00:00")]
        if not macs:
            skipped.append(d["name"])
            continue
        ok = send_wol(macs[0])
        sent.append(f"{'✅' if ok else '❌'} {d['name']} ({macs[0]})")
    lines = [f"📡 <b>Group WoL: {group}</b>", ""]
    if sent:
        lines += ["<b>Отправлено:</b>"] + sent
    if skipped:
        lines += ["", "<b>Пропущено (нет MAC):</b>"] + [f"  ⚪ {n}" for n in skipped]
    if not sent and not skipped:
        lines.append("✅ Все устройства уже онлайн.")
    await cb.message.answer("\n".join(lines), parse_mode="HTML")
    await cb.answer()


# ─── Command Scheduler ───────────────────────────────────────────────

def _sched_load() -> list[dict]:
    return _load_json(SCHEDULER_FILE, [])

def _sched_save(tasks: list[dict]) -> None:
    _save_json(SCHEDULER_FILE, tasks)

def _sched_add(devices: list[str], command: str, run_at: datetime) -> dict:
    tasks = _sched_load()
    task = {
        "id": int(time.time() * 1000) % 100000,
        "devices": devices,
        "command": command,
        "run_at": run_at.isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "status": "pending",
    }
    tasks.append(task)
    _sched_save(tasks)
    return task

def _parse_run_time(text: str) -> datetime | None:
    """Parse 'через 30' / '30' (minutes from now) or 'HH:MM' (today/tomorrow UTC)."""
    text = text.strip()
    now = datetime.now(timezone.utc)
    m = re.match(r'^(?:через\s+)?(\d+)$', text, re.IGNORECASE)
    if m:
        return now + timedelta(minutes=int(m.group(1)))
    m = re.match(r'^(\d{1,2}):(\d{2})$', text)
    if m:
        h, mn = int(m.group(1)), int(m.group(2))
        candidate = now.replace(hour=h, minute=mn, second=0, microsecond=0)
        if candidate <= now:
            candidate += timedelta(days=1)
        return candidate
    return None


async def cmd_scheduler_loop():
    """Background task: run scheduled commands when their time comes."""
    while not _shutdown_event.is_set():
        try:
            tasks = _sched_load()
            now = datetime.now(timezone.utc)
            changed = False
            for t in tasks:
                if t["status"] != "pending":
                    continue
                run_at = datetime.fromisoformat(t["run_at"])
                if run_at.tzinfo is None:
                    run_at = run_at.replace(tzinfo=timezone.utc)
                if now < run_at:
                    continue
                devs = await get_full_devices()
                name_to_id = {d["name"]: d["id"] for d in devs}
                results = []
                for dev_name in t["devices"]:
                    dev_id = name_to_id.get(dev_name)
                    if not dev_id:
                        results.append(f"<b>{dev_name}</b>: ⚠️ не найдено")
                        continue
                    try:
                        out = await mc_run_command(dev_id, t["command"])
                        out_short = (out or "(нет вывода)")[:200]
                        results.append(f"<b>{dev_name}</b>:\n<code>{out_short}</code>")
                    except Exception as ex:
                        results.append(f"<b>{dev_name}</b>: ❌ {ex}")
                t["status"] = "done"
                t["result"] = results
                t["done_at"] = now.isoformat()
                changed = True
                admin_id = _load_json(ADMIN_FILE, {}).get("admin_id") or _load_json(ADMIN_FILE, {}).get("id")
                if admin_id:
                    msg_lines = [f"⏰ <b>Планировщик</b> — задача #{t['id']} выполнена",
                                 f"Команда: <code>{t['command'][:100]}</code>", ""] + results[:10]
                    try:
                        await bot.send_message(admin_id, "\n".join(msg_lines), parse_mode="HTML")
                    except Exception:
                        pass
            if changed:
                _sched_save(tasks)
        except Exception as e:
            log.error(f"cmd_scheduler_loop: {e}")
        try:
            await asyncio.wait_for(_shutdown_event.wait(), timeout=30)
            break
        except asyncio.TimeoutError:
            pass


@router.callback_query(F.data == "tool:scheduler")
async def cb_tool_scheduler(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    tasks = _sched_load()
    pending = [t for t in tasks if t["status"] == "pending"]
    done_recent = sorted([t for t in tasks if t["status"] == "done"],
                          key=lambda t: t.get("done_at", ""), reverse=True)[:5]
    lines = ["⏰ <b>Планировщик команд</b>", ""]
    if pending:
        lines.append(f"<b>Ожидают выполнения ({len(pending)}):</b>")
        for t in pending:
            run_at = datetime.fromisoformat(t["run_at"])
            devs_s = ", ".join(t["devices"][:3])
            if len(t["devices"]) > 3:
                devs_s += f" (+{len(t['devices'])-3})"
            lines.append(f"  #{t['id']} — {t['command'][:40]}\n"
                         f"  → {devs_s}\n"
                         f"  ⏱ {run_at.strftime('%d.%m %H:%M')} UTC")
    else:
        lines.append("Нет запланированных задач.")
    if done_recent:
        lines += ["", "<b>Недавно выполнены:</b>"]
        for t in done_recent:
            lines.append(f"  #{t['id']} ✅ {t['command'][:40]}")
    kb_rows = [[InlineKeyboardButton(text="➕ Новая задача", callback_data="sched:new")]]
    if pending:
        kb_rows.append([InlineKeyboardButton(text="🗑 Очистить pending", callback_data="sched:clear")])
    await cb.message.answer("\n".join(lines), parse_mode="HTML",
                            reply_markup=InlineKeyboardMarkup(inline_keyboard=kb_rows))


@router.callback_query(F.data == "sched:new")
async def cb_sched_new(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    devs = await get_full_devices()
    groups = sorted(set(d["group"] for d in devs if d.get("group")))
    rows = [[InlineKeyboardButton(text=f"📁 {g}", callback_data=f"sched:grp:{g[:40]}")] for g in groups]
    rows.append([InlineKeyboardButton(text="❌ Отмена", callback_data="sched:cancel")])
    await state.set_state(SchedulerFSM.picking_group)
    await cb.message.answer("⏰ <b>Новая задача — шаг 1/4</b>\nВыберите группу:",
                            parse_mode="HTML",
                            reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))


@router.callback_query(F.data.startswith("sched:grp:"), SchedulerFSM.picking_group)
async def cb_sched_pick_group(cb: CallbackQuery, state: FSMContext):
    group = cb.data[len("sched:grp:"):]
    devs = await get_full_devices()
    group_devs = [d for d in devs if d["group"] == group]
    await state.update_data(group=group, selected=[], group_devs=[d["name"] for d in group_devs])
    await state.set_state(SchedulerFSM.picking_devices)
    rows = []
    for d in group_devs:
        st = "🟢" if d["online"] else "🔴"
        rows.append([InlineKeyboardButton(text=f"{st} {d['name']}", callback_data=f"sched:dev:{d['name'][:50]}")])
    rows.append([
        InlineKeyboardButton(text="☑️ Выбрать все", callback_data="sched:all"),
        InlineKeyboardButton(text="➡️ Далее", callback_data="sched:next_cmd"),
    ])
    rows.append([InlineKeyboardButton(text="❌ Отмена", callback_data="sched:cancel")])
    await cb.answer()
    await cb.message.answer(
        f"⏰ <b>Шаг 2/4</b> — выберите устройства (тап = вкл/выкл)\nГруппа: <b>{group}</b>",
        parse_mode="HTML", reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))


@router.callback_query(F.data.startswith("sched:dev:"), SchedulerFSM.picking_devices)
async def cb_sched_toggle_dev(cb: CallbackQuery, state: FSMContext):
    name = cb.data[len("sched:dev:"):]
    data = await state.get_data()
    selected: list = list(data.get("selected", []))
    if name in selected:
        selected.remove(name)
        note = f"❌ {name}"
    else:
        selected.append(name)
        note = f"✅ {name}"
    await state.update_data(selected=selected)
    await cb.answer(note, show_alert=False)


@router.callback_query(F.data == "sched:all", SchedulerFSM.picking_devices)
async def cb_sched_all(cb: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    all_devs = data.get("group_devs", [])
    await state.update_data(selected=list(all_devs))
    await cb.answer(f"Выбраны все: {len(all_devs)}", show_alert=False)


@router.callback_query(F.data == "sched:next_cmd", SchedulerFSM.picking_devices)
async def cb_sched_next_cmd(cb: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    selected = data.get("selected", [])
    if not selected:
        await cb.answer("Выберите хотя бы одно устройство!", show_alert=True)
        return
    await state.set_state(SchedulerFSM.entering_cmd)
    await cb.answer()
    sel_s = ", ".join(selected[:5])
    if len(selected) > 5:
        sel_s += f" (+{len(selected)-5})"
    await cb.message.answer(
        f"⏰ <b>Шаг 3/4</b>\nУстройств: {len(selected)} ({sel_s})\n\nВведите команду:",
        parse_mode="HTML")


@router.message(SchedulerFSM.entering_cmd)
async def sched_enter_cmd(msg: Message, state: FSMContext):
    if not is_admin(msg.from_user.id):
        return
    cmd = msg.text.strip() if msg.text else ""
    if not cmd:
        await msg.answer("Введите непустую команду.")
        return
    await state.update_data(command=cmd)
    await state.set_state(SchedulerFSM.entering_time)
    await msg.answer(
        "⏰ <b>Шаг 4/4</b> — введите время запуска:\n"
        "• <code>30</code> — через 30 минут\n"
        "• <code>14:30</code> — в 14:30 UTC (сегодня или завтра)",
        parse_mode="HTML")


@router.message(SchedulerFSM.entering_time)
async def sched_enter_time(msg: Message, state: FSMContext):
    if not is_admin(msg.from_user.id):
        return
    run_at = _parse_run_time(msg.text.strip() if msg.text else "")
    if not run_at:
        await msg.answer(
            "❌ Не понял время. Примеры:\n"
            "<code>30</code> — через 30 мин\n"
            "<code>14:30</code> — в 14:30 UTC", parse_mode="HTML")
        return
    data = await state.get_data()
    await state.clear()
    task = _sched_add(data["selected"], data["command"], run_at)
    devs_s = ", ".join(data["selected"][:5])
    if len(data["selected"]) > 5:
        devs_s += f" (+{len(data['selected'])-5})"
    await msg.answer(
        f"✅ <b>Задача #{task['id']} запланирована</b>\n\n"
        f"Команда: <code>{data['command'][:100]}</code>\n"
        f"Устройства: {devs_s}\n"
        f"Запуск: {run_at.strftime('%d.%m.%Y %H:%M')} UTC",
        parse_mode="HTML")


@router.callback_query(F.data == "sched:cancel")
async def cb_sched_cancel(cb: CallbackQuery, state: FSMContext):
    await state.clear()
    await cb.answer("Отменено")
    await cb.message.answer("❌ Создание задачи отменено.")


@router.callback_query(F.data == "sched:clear")
async def cb_sched_clear(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    tasks = _sched_load()
    tasks = [t for t in tasks if t["status"] != "pending"]
    _sched_save(tasks)
    await cb.answer("Pending задачи очищены", show_alert=True)


# ─── SNMP Router Monitoring ───────────────────────────────────────────

async def run_snmp_probe(device_id: str, probe: dict) -> dict | None:
    """Run snmp_probe.ps1 on the remote PC; return parsed JSON or None."""
    try:
        script = SNMP_PROBE_SCRIPT.read_text(encoding="utf-8")
    except Exception as e:
        log.error(f"snmp_probe: cannot read script: {e}")
        return None
    community = probe.get("snmp_community", "public") or "public"
    script = script.replace("SNMP_COMMUNITY_PLACEHOLDER", community)
    login_key = await _get_login_key()
    if not login_key:
        return None
    args = [
        "node", MESHCTRL, "RunCommand",
        "--url", MC_WSS,
        "--loginkey", login_key,
        "--id", device_id,
        "--run", script,
        "--reply", "--powershell",
    ]
    try:
        proc = await asyncio.create_subprocess_exec(
            *args,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
            cwd=MC_DIR,
        )
        stdout, _ = await asyncio.wait_for(proc.communicate(), timeout=60)
        raw = stdout.decode(errors="replace").strip()
        brace = raw.find("{")
        if brace == -1:
            return None
        obj, _ = json.JSONDecoder().raw_decode(raw, brace)
        return obj
    except asyncio.TimeoutError:
        log.warning(f"snmp_probe timeout for device {device_id}")
    except Exception as e:
        log.error(f"snmp_probe error: {e}")
    return None


def _snmp_fmt_rate(bps: float) -> str:
    """Format bytes/sec as human-readable rate."""
    if bps < 0:
        return "?"
    if bps < 1024:
        return f"{bps:.0f} B/s"
    if bps < 1024 * 1024:
        return f"{bps/1024:.1f} KB/s"
    return f"{bps/1024/1024:.2f} MB/s"


async def snmp_poll_loop():
    """Background task: poll SNMP on router via MeshCentral agent every 5 min."""
    global _snmp_data
    # Load cached data
    try:
        _snmp_data = json.loads(SNMP_DATA_FILE.read_text())
    except Exception:
        _snmp_data = {}

    while not _shutdown_event.is_set():
        try:
            probes = _load_json(KEENETIC_PROBES_FILE, [])
            devs = await get_full_devices()
            name_to_id = {d["name"]: d["id"] for d in devs}

            for probe in probes:
                if not probe.get("snmp_community"):
                    continue
                agent_name = probe.get("agent_name", "")
                dev_id = name_to_id.get(agent_name)
                if not dev_id:
                    continue
                # Check if agent is online
                dev = next((d for d in devs if d["name"] == agent_name), None)
                if not dev or not dev.get("online"):
                    continue

                result = await run_snmp_probe(dev_id, probe)
                location = probe.get("location", agent_name)
                now_ts = time.time()

                if not result or result.get("error"):
                    err = result.get("error", "timeout") if result else "timeout"
                    _snmp_data[agent_name] = {
                        "ok": False, "location": location,
                        "error": err, "updated": now_ts,
                    }
                    log.warning(f"snmp_poll: {agent_name}: {err}")
                    continue

                # Calculate traffic rates from previous sample
                prev = _snmp_data.get(agent_name, {}).get("data")
                prev_ts = _snmp_data.get(agent_name, {}).get("updated", 0)
                rates = {}
                if prev and prev_ts and (now_ts - prev_ts) > 5:
                    dt = now_ts - prev_ts
                    for iface in ("if1", "if2", "if3"):
                        old_in  = prev.get(f"{iface}_in", -1)
                        old_out = prev.get(f"{iface}_out", -1)
                        new_in  = result.get(f"{iface}_in", -1)
                        new_out = result.get(f"{iface}_out", -1)
                        if old_in >= 0 and new_in >= 0 and new_in >= old_in:
                            rates[f"{iface}_rate_in"]  = (new_in  - old_in)  / dt
                            rates[f"{iface}_rate_out"] = (new_out - old_out) / dt

                _snmp_data[agent_name] = {
                    "ok": True, "location": location,
                    "router": result.get("router", ""),
                    "updated": now_ts,
                    "data": result,
                    "rates": rates,
                }
                log.info(f"snmp_poll: {agent_name} ({location}) CPU={result.get('cpu_pct',-1)}%"
                         f" uptime={result.get('uptime','?')}")

            _save_json(SNMP_DATA_FILE, _snmp_data)
        except Exception as e:
            log.error(f"snmp_poll_loop: {e}")
        try:
            await asyncio.wait_for(_shutdown_event.wait(), timeout=SNMP_POLL_INTERVAL)
            break
        except asyncio.TimeoutError:
            pass


def _snmp_status_text() -> str:
    """Build SNMP status summary string."""
    if not _snmp_data:
        return "Нет данных. Добавьте <code>snmp_community</code> в keenetic_probes.json."
    lines = ["📡 <b>SNMP роутеры</b>", ""]
    for agent, entry in _snmp_data.items():
        loc = entry.get("location", agent)
        if not entry.get("ok"):
            lines.append(f"🔴 <b>{loc}</b> — ❌ {entry.get('error','?')}")
            continue
        d = entry.get("data", {})
        r = entry.get("rates", {})
        cpu    = d.get("cpu_pct", -1)
        uptime = d.get("uptime", "?")
        router = d.get("router", "?")
        name   = d.get("sys_name", "") or d.get("sys_descr", "")[:40] or "?"
        cpu_s  = f"{cpu}%" if cpu >= 0 else "?"
        upd    = datetime.fromtimestamp(entry["updated"]).strftime("%H:%M")
        lines.append(f"🟢 <b>{loc}</b> — {router} ({name})")
        lines.append(f"   CPU: {cpu_s}  |  Uptime: {uptime}  |  ⏱ {upd}")
        # Traffic for first non-zero interface rate
        for iface in ("if1", "if2", "if3"):
            ri = r.get(f"{iface}_rate_in", -1)
            ro = r.get(f"{iface}_rate_out", -1)
            if ri >= 0 and ro >= 0:
                lines.append(f"   ↓ {_snmp_fmt_rate(ri)}  ↑ {_snmp_fmt_rate(ro)}")
                break
        lines.append("")
    return "\n".join(lines).strip()


@router.callback_query(F.data == "tool:snmp")
async def cb_tool_snmp(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    text = _snmp_status_text()
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="🔄 Обновить", callback_data="snmp:refresh"),
        InlineKeyboardButton(text="⚙️ Настройка", callback_data="snmp:config"),
    ]])
    await cb.message.answer(text, parse_mode="HTML", reply_markup=kb)


@router.callback_query(F.data == "snmp:refresh")
async def cb_snmp_refresh(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer("Данные обновляются раз в 5 мин автоматически")
    text = _snmp_status_text()
    try:
        await cb.message.edit_text(text, parse_mode="HTML", reply_markup=cb.message.reply_markup)
    except Exception:
        await cb.message.answer(text, parse_mode="HTML", reply_markup=cb.message.reply_markup)


@router.callback_query(F.data == "snmp:config")
async def cb_snmp_config(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    probes = _load_json(KEENETIC_PROBES_FILE, [])
    lines = ["⚙️ <b>SNMP — настройка</b>", "",
             "Добавьте поле <code>snmp_community</code> в keenetic_probes.json для нужных зондов:",
             "", "<pre>"]
    for p in probes:
        has = "✅" if p.get("snmp_community") else "❌"
        comm = p.get("snmp_community", "не задано")
        lines.append(f'{has} {p.get("location", p.get("agent_name","?"))}: community="{comm}"')
    lines += ["</pre>", "", "Пример: добавьте в probe объект:",
              '<code>"snmp_community": "public"</code>',
              "", "Стандартный community string на Keenetic: <b>public</b>"]
    await cb.message.answer("\n".join(lines), parse_mode="HTML")
    await cb.answer()


# ─── HW Inventory handler ────────────────────────────────────────────

@router.callback_query(F.data == "tool:hw_inventory")
async def cb_tool_hw_inventory(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    global _hw_inventory
    if not _hw_inventory:
        _hw_inventory = _load_json(HW_INVENTORY_FILE, {})

    if not _hw_inventory:
        await cb.message.answer(
            "💻 <b>HW инвентарь</b>\n\nДанные ещё собираются. Сбор запускается автоматически "
            "каждые 4 часа для онлайн-устройств.\n\n"
            "Нажмите кнопку ещё раз через несколько минут.",
            parse_mode="HTML",
        )
        return

    # Summary table
    lines = ["💻 <b>Аппаратный инвентарь</b>\n"]
    for name, inv in sorted(_hw_inventory.items()):
        cpu = inv.get("cpu_name", "?")[:30]
        ram = inv.get("ram_total_gb", "?")
        disks = inv.get("disks", [])
        disk_str = " ".join(f"{d.get('letter','?')}:{d.get('size_gb','?')}GB[{d.get('dtype','?')}]" for d in disks[:3])
        lines.append(f"<b>{name}</b>")
        lines.append(f"  ⚡ {cpu}")
        lines.append(f"  🧠 {ram} GB RAM")
        if disk_str:
            lines.append(f"  💾 {disk_str}")
        lines.append("")

    text = "\n".join(lines)

    # Offer per-device detail buttons
    dev_buttons = []
    for name in sorted(_hw_inventory.keys()):
        safe = name[:30]
        dev_buttons.append([InlineKeyboardButton(text=f"💻 {safe}", callback_data=f"hw_detail:{safe}")])
    dev_buttons.append([InlineKeyboardButton(text="🔄 Собрать сейчас", callback_data="hw_collect_now")])

    await cb.message.answer(
        text[:4000],
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=dev_buttons),
    )


@router.callback_query(F.data.startswith("hw_detail:"))
async def cb_hw_detail(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    device_name = cb.data.split(":", 1)[1]
    inv = _hw_inventory.get(device_name)
    text = _hw_inventory_text(inv, device_name)
    await cb.message.answer(text, parse_mode="HTML")
    await cb.answer()


@router.callback_query(F.data == "hw_collect_now")
async def cb_hw_collect_now(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer("⏳ Запускаю сбор…", show_alert=True)
    await cb.message.answer("⏳ Запускаю сбор HW данных для онлайн-устройств…\nЭто займёт 1-3 минуты.", parse_mode="HTML")
    asyncio.create_task(_hw_collect_now_task(cb.from_user.id))


async def _hw_collect_now_task(aid: int):
    try:
        devs = await get_full_devices()
        online = [d for d in devs if d.get("online")]
        for d in online:
            await _collect_hw_for_device(d["id"], d["name"])
            await asyncio.sleep(3)
        await bot.send_message(
            aid,
            f"✅ HW инвентарь обновлён: {len(online)} устройств\n"
            f"Используйте 💻 Инвентарь HW чтобы посмотреть результаты.",
            parse_mode="HTML",
        )
    except Exception as e:
        await bot.send_message(aid, f"❌ Ошибка сбора HW: {e}")


# ─── Temperature handler ─────────────────────────────────────────────

@router.callback_query(F.data == "tool:temperature")
async def cb_tool_temperature(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    global _temp_data
    if not _temp_data:
        _temp_data = _load_json(TEMP_DATA_FILE, {})

    if not _temp_data:
        await cb.message.answer(
            "🌡 <b>Температуры</b>\n\nДанные ещё собираются. "
            "Сбор запускается каждые 15 минут для онлайн-устройств.\n"
            "Попробуйте через несколько минут.",
            parse_mode="HTML",
        )
        return

    lines = ["🌡 <b>Температуры и нагрузка CPU</b>\n"]
    for name, data in sorted(_temp_data.items()):
        load = data.get("cpu_load_pct", 0)
        temps = data.get("temps", [])
        updated = data.get("updated", "")
        if data.get("no_sensor"):
            lines.append(f"<b>{name}</b>: датчики недоступны, CPU {load}%")
        elif not temps:
            lines.append(f"<b>{name}</b>: нет данных, CPU {load}%")
        else:
            # Show max temp
            max_t = max(t.get("temp_c", 0) for t in temps)
            warn = "🔴" if max_t >= TEMP_WARN_C else ("🟡" if max_t >= 60 else "🟢")
            lines.append(f"{warn} <b>{name}</b>: {max_t}°C  CPU {load}%")
            for sensor in temps[:3]:
                lines.append(f"   · {sensor.get('zone','?')[:40]}: {sensor.get('temp_c','?')}°C")
        if updated:
            lines.append(f"   <i>обновлено {updated}</i>")
        lines.append("")

    lines.append(f"🔴 ≥{TEMP_WARN_C}°C критично  🟡 ≥60°C повышено  🟢 норма")

    await cb.message.answer(
        "\n".join(lines)[:4000],
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔄 Обновить сейчас", callback_data="temp_refresh_now")],
        ]),
    )


@router.callback_query(F.data == "temp_refresh_now")
async def cb_temp_refresh_now(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer("⏳ Собираю температуры…", show_alert=True)
    asyncio.create_task(_temp_collect_now_task(cb.from_user.id))


async def _temp_collect_now_task(aid: int):
    global _temp_data
    try:
        devs = await get_full_devices()
        online = [d for d in devs if d.get("online")]
        for d in online:
            result = await _collect_temp_for_device(d["id"], d["name"])
            if result:
                result["updated"] = datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M")
                _temp_data[d["name"]] = result
                _save_json(TEMP_DATA_FILE, _temp_data)
            await asyncio.sleep(2)
        await bot.send_message(aid, f"✅ Температуры собраны: {len(online)} устройств. Нажмите 🌡 Температуры снова.")
    except Exception as e:
        await bot.send_message(aid, f"❌ Ошибка: {e}")


# ─── Availability heatmap handler ─────────────────────────────────────

@router.callback_query(F.data == "tool:availability")
async def cb_tool_availability(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    # Show device picker
    devs = await get_full_devices()
    buttons = []
    row = []
    for d in sorted(devs, key=lambda x: x["name"]):
        row.append(InlineKeyboardButton(text=d["name"][:20], callback_data=f"avail_dev:{d['name'][:40]}"))
        if len(row) == 2:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    await cb.message.answer(
        "📊 <b>Доступность устройств</b>\n\nВыберите устройство для просмотра тепловой карты 7 дней:",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
    )


@router.callback_query(F.data.startswith("avail_dev:"))
async def cb_avail_device(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    device_name = cb.data.split(":", 1)[1]
    heatmap = build_availability_heatmap(device_name)
    await cb.message.answer(heatmap, parse_mode="HTML")
    await cb.answer()


# ─── Status page handler ─────────────────────────────────────────────

@router.callback_query(F.data == "tool:status_page")
async def cb_tool_status_page(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("🔒", show_alert=True)
        return
    await cb.answer()
    status_url = f"{MC_URL}/status"
    if not STATUS_HTML_FILE.exists():
        await cb.message.answer(
            "🌐 Страница статуса ещё не сгенерирована. Подождите 60 секунд.",
            parse_mode="HTML",
        )
        return
    await cb.message.answer(
        f"🌐 <b>Страница статуса сети</b>\n\n"
        f"Публичная страница (без авторизации):\n"
        f"<a href='{status_url}'>{status_url}</a>\n\n"
        f"Обновляется каждые 60 секунд автоматически.\n"
        f"Показывает онлайн/офлайн по каждой локации.",
        parse_mode="HTML",
        disable_web_page_preview=True,
    )


async def main():
    log.info("MeshCentral Monitor Bot v4 starting...")
    loop = asyncio.get_running_loop()
    for sig in (signal.SIGTERM, signal.SIGINT):
        loop.add_signal_handler(sig, lambda: asyncio.create_task(shutdown()))
    dp.startup.register(on_startup)
    try:
        await dp.start_polling(bot)
    finally:
        if not _shutdown_event.is_set():
            await shutdown()


if __name__ == "__main__":
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    asyncio.run(main())
